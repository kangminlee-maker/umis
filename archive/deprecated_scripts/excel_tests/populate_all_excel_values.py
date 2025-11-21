#!/usr/bin/env python3
"""
모든 Excel 파일에 계산된 값 직접 입력
수식 대신 값을 하드코딩하여 즉시 확인 가능
"""

import sys
import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def populate_financial_projection():
    """Financial Projection 값 입력"""
    
    filepath = project_root / 'examples' / 'excel' / 'financial_projection_korean_adult_education_example_20251104.xlsx'
    
    print("\n" + "="*70)
    print("📊 Financial Projection 값 입력")
    print("="*70 + "\n")
    
    wb = load_workbook(filepath)
    
    # 계산
    y0 = 1250_0000_0000
    growth = 1.28
    
    years_revenue = {
        0: y0,
        1: y0 * growth,
        2: y0 * (growth ** 2),
        3: y0 * (growth ** 3),
        4: y0 * (growth ** 4),
        5: y0 * (growth ** 5),
    }
    
    # Revenue_Buildup
    if 'Revenue_Buildup' in wb.sheetnames:
        ws = wb['Revenue_Buildup']
        
        # Total Revenue 행에 값 입력 (Row 9)
        for year in range(6):
            col_letter = chr(66 + year)  # B, C, D, E, F, G
            ws[f'{col_letter}9'] = years_revenue[year]
            ws[f'{col_letter}9'].number_format = '#,##0'
        
        print(f"Revenue_Buildup (Row 9):")
        print(f"  Year 0: ₩{years_revenue[0]/1_0000_0000:.0f}억")
        print(f"  Year 5: ₩{years_revenue[5]/1_0000_0000:.0f}억\n")
    
    # Dashboard
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        
        ws['B5'] = years_revenue[5]
        ws['B5'].number_format = '₩#,##0'
        
        net_income_y5 = years_revenue[5] * 0.10
        ws['B6'] = net_income_y5
        ws['B6'].number_format = '₩#,##0'
        
        cagr = (years_revenue[5] / years_revenue[0]) ** (1/5) - 1
        ws['B7'] = cagr
        ws['B7'].number_format = '0.0%'
        
        print(f"Dashboard:")
        print(f"  Revenue Y5 (B5): ₩{years_revenue[5]/1_0000_0000:.0f}억")
        print(f"  Net Income Y5 (B6): ₩{net_income_y5/1_0000_0000:.0f}억")
        print(f"  CAGR (B7): {cagr*100:.0f}%\n")
    
    # 저장
    output_path = project_root / 'examples' / 'excel' / 'financial_projection_CALCULATED_20251104.xlsx'
    wb.save(output_path)
    
    print(f"✅ 저장: {output_path.name}\n")
    
    return output_path


def populate_unit_economics():
    """Unit Economics 값 입력"""
    
    filepath = project_root / 'examples' / 'excel' / 'unit_economics_music_streaming_example_20251104.xlsx'
    
    print("="*70)
    print("📊 Unit Economics 값 입력")
    print("="*70 + "\n")
    
    wb = load_workbook(filepath)
    
    # 계산
    arpu = 9000
    lifetime = 25
    margin = 0.35
    cac = 25000
    churn = 0.04
    
    ltv1 = arpu * lifetime * margin  # 78,750
    ltv2 = arpu * margin / churn  # 78,750
    ltv = (ltv1 + ltv2) / 2  # 78,750
    ratio = ltv / cac  # 3.15
    payback = cac / (arpu * margin)  # 7.94
    
    # LTV_Calculation
    if 'LTV_Calculation' in wb.sheetnames:
        ws = wb['LTV_Calculation']
        
        # LTV 방법 1 (Row 9)
        ws['B9'] = ltv1
        ws['B9'].number_format = '#,##0'
        
        # LTV 방법 2 (Row 16)
        ws['B16'] = ltv2
        ws['B16'].number_format = '#,##0'
        
        # LTV 평균 (Row 18)
        ws['B18'] = ltv
        ws['B18'].number_format = '#,##0'
        ws['B18'].font = Font(bold=True)
        
        print(f"LTV_Calculation:")
        print(f"  LTV 방법 1 (B9): ₩{ltv1:,.0f}")
        print(f"  LTV 평균 (B18): ₩{ltv:,.0f}\n")
    
    # LTV_CAC_Ratio
    if 'LTV_CAC_Ratio' in wb.sheetnames:
        ws = wb['LTV_CAC_Ratio']
        
        # Ratio (Row 7)
        ws['B7'] = ratio
        ws['B7'].number_format = '0.00'
        
        print(f"LTV_CAC_Ratio:")
        print(f"  Ratio (B7): {ratio:.2f} ✅\n")
    
    # Payback_Period
    if 'Payback_Period' in wb.sheetnames:
        ws = wb['Payback_Period']
        
        # Payback (Row 11)
        ws['B11'] = payback
        ws['B11'].number_format = '0.0'
        
        print(f"Payback_Period:")
        print(f"  Payback (B11): {payback:.1f}개월 ✅\n")
    
    # Dashboard
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        
        ws['B5'] = ltv
        ws['B5'].number_format = '₩#,##0'
        
        ws['B6'] = cac
        ws['B6'].number_format = '₩#,##0'
        
        ws['B7'] = ratio
        ws['B7'].number_format = '0.00'
        
        ws['B8'] = payback
        ws['B8'].number_format = '0.0'
        
        print(f"Dashboard:")
        print(f"  LTV (B5): ₩{ltv:,.0f}")
        print(f"  Ratio (B7): {ratio:.2f} (Good) ✅\n")
    
    # 저장
    output_path = project_root / 'examples' / 'excel' / 'unit_economics_CALCULATED_20251104.xlsx'
    wb.save(output_path)
    
    print(f"✅ 저장: {output_path.name}\n")
    
    return output_path


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🎯 모든 Excel 파일에 값 입력")
    print("="*70)
    print("\n목적: 수식 대신 계산된 값을 직접 입력하여 즉시 확인\n")
    
    # 1. Financial Projection
    fp_path = populate_financial_projection()
    
    # 2. Unit Economics
    ue_path = populate_unit_economics()
    
    # 3. Market Sizing (이미 완료)
    ms_path = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_CALCULATED_20251104.xlsx'
    
    print("="*70)
    print("🎉 모든 파일 완료!")
    print("="*70)
    
    print("\n📁 생성된 파일 (값이 입력된 버전):")
    print(f"  1. {ms_path.name}")
    print(f"  2. {ue_path.name}")
    print(f"  3. {fp_path.name}")
    
    print("\n💡 사용 방법:")
    print("  이 파일들을 Excel에서 열면 모든 값이 즉시 표시됩니다!")
    print("  수식은 주석으로 확인 가능합니다.")
    
    sys.exit(0)

