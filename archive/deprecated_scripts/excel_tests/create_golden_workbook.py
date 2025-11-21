#!/usr/bin/env python3
"""
Golden Workbook 생성 (정답지)
모든 값을 하드코딩한 정답 Excel 파일
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def create_golden_financial_projection():
    """
    Financial Projection Golden Workbook
    
    성인 교육 케이스:
    - Year 0: ₩1,250억
    - Growth: 28% YoY
    - Year 1 = 1,250 × 1.28 = ₩1,600억
    - Year 3 = 1,250 × 1.28^3 = ₩2,621억
    - Year 5 = 1,250 × 1.28^5 = ₩4,295억
    """
    
    print("\n📊 Golden Workbook 생성: Financial Projection")
    print("="*70)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Golden_Values"
    
    # 제목
    ws['A1'] = "Golden Workbook - Financial Projection"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = "정답지: 성인 교육 시장 (CAGR 28%)"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells('A2:E2')
    
    # 헤더
    ws['A4'] = "Metric"
    ws['B4'] = "Year 0"
    ws['C4'] = "Year 1"
    ws['D4'] = "Year 3"
    ws['E4'] = "Year 5"
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'{col}4'].font = Font(bold=True, color="FFFFFF")
    
    # 컬럼 폭
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # === 정답 값 (하드코딩) ===
    
    # Revenue
    y0 = 1250_0000_0000
    y1 = y0 * 1.28
    y3 = y0 * (1.28 ** 3)
    y5 = y0 * (1.28 ** 5)
    
    row = 5
    ws[f'A{row}'] = "Revenue"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    ws[f'B{row}'] = y0
    ws[f'C{row}'] = y1
    ws[f'D{row}'] = y3
    ws[f'E{row}'] = y5
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].font = Font(bold=True)
    
    # COGS (30% of Revenue, 1-Gross Margin)
    row += 1
    ws[f'A{row}'] = "COGS"
    ws[f'B{row}'] = y0 * 0.30
    ws[f'C{row}'] = y1 * 0.30
    ws[f'D{row}'] = y3 * 0.30
    ws[f'E{row}'] = y5 * 0.30
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
    
    # Gross Profit
    row += 1
    ws[f'A{row}'] = "Gross Profit"
    ws[f'B{row}'] = y0 * 0.70
    ws[f'C{row}'] = y1 * 0.70
    ws[f'D{row}'] = y3 * 0.70
    ws[f'E{row}'] = y5 * 0.70
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # S&M (30%)
    row += 1
    ws[f'A{row}'] = "S&M (30%)"
    ws[f'B{row}'] = y0 * 0.30
    ws[f'C{row}'] = y1 * 0.30
    ws[f'D{row}'] = y3 * 0.30
    ws[f'E{row}'] = y5 * 0.30
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
    
    # R&D (15%)
    row += 1
    ws[f'A{row}'] = "R&D (15%)"
    ws[f'B{row}'] = y0 * 0.15
    ws[f'C{row}'] = y1 * 0.15
    ws[f'D{row}'] = y3 * 0.15
    ws[f'E{row}'] = y5 * 0.15
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
    
    # G&A (10%)
    row += 1
    ws[f'A{row}'] = "G&A (10%)"
    ws[f'B{row}'] = y0 * 0.10
    ws[f'C{row}'] = y1 * 0.10
    ws[f'D{row}'] = y3 * 0.10
    ws[f'E{row}'] = y5 * 0.10
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
    
    # Total OPEX (55%)
    row += 1
    ws[f'A{row}'] = "Total OPEX (55%)"
    ws[f'B{row}'] = y0 * 0.55
    ws[f'C{row}'] = y1 * 0.55
    ws[f'D{row}'] = y3 * 0.55
    ws[f'E{row}'] = y5 * 0.55
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    # EBITDA (Gross Profit - OPEX = 70% - 55% = 15%)
    row += 1
    ws[f'A{row}'] = "EBITDA (15%)"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    ws[f'B{row}'] = y0 * 0.15
    ws[f'C{row}'] = y1 * 0.15
    ws[f'D{row}'] = y3 * 0.15
    ws[f'E{row}'] = y5 * 0.15
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws[f'{col}{row}'].font = Font(bold=True)
    
    # Net Income (10% of Revenue)
    row += 1
    ws[f'A{row}'] = "Net Income (10%)"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'B{row}'] = y0 * 0.10
    ws[f'C{row}'] = y1 * 0.10
    ws[f'D{row}'] = y3 * 0.10
    ws[f'E{row}'] = y5 * 0.10
    
    for col in ['B', 'C', 'D', 'E']:
        ws[f'{col}{row}'].number_format = '#,##0'
        ws[f'{col}{row}'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws[f'{col}{row}'].font = Font(bold=True, color="FFFFFF")
    
    # === 가이드 ===
    row += 2
    ws[f'A{row}'] = "📋 검증 방법"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    
    row += 1
    ws[f'A{row}'] = "1. 생성된 Excel 파일 열기"
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "2. Revenue_Buildup 시트 → Total Revenue 행 확인"
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "3. 이 Golden Workbook의 값과 비교"
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    ws[f'A{row}'] = "4. 오차 < 1% 이면 정상"
    ws.merge_cells(f'A{row}:E{row}')
    
    # 저장
    output_dir = project_root / 'examples' / 'excel'
    filepath = output_dir / 'golden_financial_projection.xlsx'
    wb.save(filepath)
    
    print(f"✅ Golden Workbook 생성: {filepath}")
    print(f"📏 크기: {filepath.stat().st_size / 1024:.1f} KB\n")
    
    print("📊 정답 (하드코딩된 값):")
    print(f"   Year 0 Revenue: ₩{y0/1_0000_0000:.0f}억")
    print(f"   Year 1 Revenue: ₩{y1/1_0000_0000:.0f}억 (×1.28)")
    print(f"   Year 3 Revenue: ₩{y3/1_0000_0000:.0f}억 (×1.28^3)")
    print(f"   Year 5 Revenue: ₩{y5/1_0000_0000:.0f}억 (×1.28^5)")
    print(f"   Year 5 Net Income: ₩{y5*0.10/1_0000_0000:.0f}억 (10%)")
    
    return filepath


def create_golden_unit_economics():
    """
    Unit Economics Golden Workbook
    
    음악 스트리밍 케이스:
    - ARPU: ₩9,000
    - Lifetime: 25개월
    - Margin: 35%
    - LTV = 9,000 × 25 × 0.35 = ₩78,750
    - CAC: ₩25,000
    - LTV/CAC = 78,750 / 25,000 = 3.15
    """
    
    print("\n📊 Golden Workbook 생성: Unit Economics")
    print("="*70)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Golden_Values"
    
    # 제목
    ws['A1'] = "Golden Workbook - Unit Economics"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "정답지: 음악 스트리밍 (LTV/CAC 3.15)"
    ws['A2'].font = Font(size=10, italic=True, color="666666")
    ws.merge_cells('A2:D2')
    
    # 헤더
    ws['A4'] = "Metric"
    ws['B4'] = "Value"
    ws['C4'] = "Unit"
    ws['D4'] = "Calculation"
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'{col}4'].font = Font(bold=True, color="FFFFFF")
    
    # 컬럼 폭
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 35
    
    # === 정답 값 ===
    
    arpu = 9000
    lifetime = 25
    margin = 0.35
    churn = 0.04
    cac = 25000
    
    # LTV (방법 1)
    ltv1 = arpu * lifetime * margin
    
    # LTV (방법 2)
    ltv2 = arpu * margin / churn
    
    # LTV (평균)
    ltv = (ltv1 + ltv2) / 2
    
    # LTV/CAC
    ratio = ltv / cac
    
    # Payback
    payback = cac / (arpu * margin)
    
    row = 5
    data = [
        ("ARPU", arpu, "원/월", "입력값"),
        ("CAC", cac, "원", "입력값"),
        ("Gross Margin", margin, "%", "입력값"),
        ("Monthly Churn", churn, "%", "입력값"),
        ("Customer Lifetime", lifetime, "months", "입력값"),
        ("", "", "", ""),
        ("LTV (방법 1)", ltv1, "원", "ARPU × Lifetime × Margin"),
        ("LTV (방법 2)", ltv2, "원", "ARPU × Margin / Churn"),
        ("LTV (평균)", ltv, "원", "(방법1 + 방법2) / 2"),
        ("", "", "", ""),
        ("LTV/CAC Ratio", ratio, "배", "LTV / CAC"),
        ("Payback Period", payback, "개월", "CAC / (ARPU × Margin)"),
    ]
    
    for metric, value, unit, calc in data:
        ws[f'A{row}'] = metric
        
        if metric:
            ws[f'A{row}'].font = Font(size=10, bold=("LTV" in metric or "Ratio" in metric or "Payback" in metric))
            
            if value != "":
                ws[f'B{row}'] = value
                
                # 숫자 포맷
                if unit == "%":
                    ws[f'B{row}'].number_format = '0.0%'
                elif unit in ["원", "원/월"]:
                    ws[f'B{row}'].number_format = '#,##0'
                else:
                    ws[f'B{row}'].number_format = '0.00'
                
                # 강조
                if "LTV" in metric or "Ratio" in metric:
                    ws[f'B{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    ws[f'B{row}'].font = Font(bold=True)
            
            ws[f'C{row}'] = unit
            ws[f'D{row}'] = calc
            ws[f'D{row}'].font = Font(size=9, italic=True)
        
        row += 1
    
    # === 검증 기준 ===
    row += 1
    ws[f'A{row}'] = "✅ 검증 기준"
    ws[f'A{row}'].font = Font(size=11, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = f"LTV/CAC = {ratio:.2f} (Good, 목표 > 3.0)"
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws[f'A{row}'] = f"Payback = {payback:.1f}개월 (Good, 목표 < 12개월)"
    ws.merge_cells(f'A{row}:D{row}')
    
    # 저장
    output_dir = project_root / 'examples' / 'excel'
    filepath = output_dir / 'golden_unit_economics.xlsx'
    wb.save(filepath)
    
    print(f"✅ Golden Workbook 생성: {filepath}")
    print(f"📏 크기: {filepath.stat().st_size / 1024:.1f} KB\n")
    
    print("📊 정답 (하드코딩된 값):")
    print(f"   LTV (방법 1): ₩{ltv1:,.0f}")
    print(f"   LTV (방법 2): ₩{ltv2:,.0f}")
    print(f"   LTV (평균): ₩{ltv:,.0f}")
    print(f"   LTV/CAC: {ratio:.2f}")
    print(f"   Payback: {payback:.1f}개월")
    
    return filepath


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🎯 Golden Workbook 생성 (정답지)")
    print("="*70)
    print("\n목적: 실제 계산 결과를 하드코딩하여 생성된 Excel과 비교\n")
    
    # 1. Financial Projection Golden
    fp_golden = create_golden_financial_projection()
    
    # 2. Unit Economics Golden
    ue_golden = create_golden_unit_economics()
    
    print("\n" + "="*70)
    print("✅ Golden Workbook 생성 완료")
    print("="*70)
    print("\n📁 생성된 파일:")
    print(f"   - {fp_golden.name}")
    print(f"   - {ue_golden.name}")
    
    print("\n💡 사용 방법:")
    print("   1. Golden Workbook 열기 (정답)")
    print("   2. 생성된 예제 Excel 열기")
    print("   3. Revenue, Net Income, LTV, LTV/CAC 비교")
    print("   4. 오차 < 1%이면 정상")
    
    print("\n📋 다음 단계:")
    print("   1. 두 파일을 나란히 열기")
    print("   2. 주요 셀 값 비교")
    print("   3. 차이 확인")
    print("   4. 오류 발견 시 Generator 수정")
    
    sys.exit(0)

