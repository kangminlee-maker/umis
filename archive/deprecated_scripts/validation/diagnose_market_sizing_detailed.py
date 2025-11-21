#!/usr/bin/env python3
"""
Market Sizing 상세 진단
전체 시트 구조와 수식 확인
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'

if not filepath.exists():
    print(f"❌ 파일 없음")
    sys.exit(1)

wb = load_workbook(filepath, data_only=False)

print("\n" + "="*70)
print("🔍 Method_1_TopDown 상세 진단")
print("="*70)

if 'Method_1_TopDown' in wb.sheetnames:
    ws = wb['Method_1_TopDown']
    
    # Row 1-15 전체 출력
    print("\nRow 1-15 전체 데이터:")
    print("-"*70)
    
    for row_idx in range(1, 16):
        print(f"\nRow {row_idx}:")
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row_idx}']
            if cell.value:
                print(f"  {col}{row_idx}: {cell.value}")

print("\n" + "="*70)
print("🔍 Convergence_Analysis 상세 진단")
print("="*70)

if 'Convergence_Analysis' in wb.sheetnames:
    ws = wb['Convergence_Analysis']
    
    # Row 1-20 전체 출력
    print("\nRow 1-20 전체 데이터:")
    print("-"*70)
    
    for row_idx in range(1, 21):
        print(f"\nRow {row_idx}:")
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row_idx}']
            if cell.value:
                print(f"  {col}{row_idx}: {cell.value}")

