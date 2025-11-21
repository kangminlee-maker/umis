#!/usr/bin/env python3
"""
생성된 Excel vs Golden Workbook 비교
수식 패턴과 계산 결과 검증
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def compare_financial_projection():
    """Financial Projection 비교"""
    
    print("\n" + "="*70)
    print("📊 Financial Projection 비교 검증")
    print("="*70 + "\n")
    
    # 파일 경로
    generated = project_root / 'examples' / 'excel' / 'financial_projection_korean_adult_education_example_20251104.xlsx'
    golden = project_root / 'examples' / 'excel' / 'golden_financial_projection.xlsx'
    
    if not generated.exists():
        print(f"❌ 생성 파일 없음: {generated.name}")
        return False
    
    if not golden.exists():
        print(f"❌ Golden 파일 없음: {golden.name}")
        return False
    
    print(f"📁 생성 파일: {generated.name}")
    print(f"📁 Golden 파일: {golden.name}\n")
    
    # Golden 값 로드
    wb_golden = load_workbook(golden, data_only=True)
    ws_golden = wb_golden['Golden_Values']
    
    golden_values = {
        'revenue_y0': ws_golden['B5'].value,
        'revenue_y1': ws_golden['C5'].value,
        'revenue_y3': ws_golden['D5'].value,
        'revenue_y5': ws_golden['E5'].value,
        'net_income_y5': ws_golden['E13'].value,
    }
    
    print("📊 Golden 값 (정답):")
    print(f"   Year 0 Revenue: ₩{golden_values['revenue_y0']/1_0000_0000:.0f}억")
    print(f"   Year 1 Revenue: ₩{golden_values['revenue_y1']/1_0000_0000:.0f}억")
    print(f"   Year 3 Revenue: ₩{golden_values['revenue_y3']/1_0000_0000:.0f}억")
    print(f"   Year 5 Revenue: ₩{golden_values['revenue_y5']/1_0000_0000:.0f}억")
    print(f"   Year 5 Net Income: ₩{golden_values['net_income_y5']/1_0000_0000:.0f}억\n")
    
    # 생성 파일 로드 (수식 확인)
    wb_gen = load_workbook(generated, data_only=False)
    
    print("🔍 생성 파일 수식 검증:")
    print("-"*70)
    
    results = []
    
    # 1. Revenue_Buildup 수식 확인
    if 'Revenue_Buildup' in wb_gen.sheetnames:
        ws_rev = wb_gen['Revenue_Buildup']
        
        # Total Revenue 행 찾기 (보통 Row 9-10)
        total_row = None
        for row_idx in range(8, 12):
            cell = ws_rev[f'A{row_idx}']
            if cell.value and 'Total Revenue' in str(cell.value):
                total_row = row_idx
                break
        
        if total_row:
            print(f"\n1. Revenue_Buildup (Row {total_row})")
            
            # Year 0 수식 확인
            cell_y0 = ws_rev[f'B{total_row}']
            print(f"   Year 0 (B{total_row}): {cell_y0.value}")
            
            if cell_y0.value and '=SUM' in str(cell_y0.value):
                results.append(('Revenue Y0 수식', True, "SUM 패턴 정상"))
            else:
                results.append(('Revenue Y0 수식', False, f"SUM 아님: {cell_y0.value}"))
            
            # Year 1 수식 확인
            cell_y1 = ws_rev[f'C{total_row}']
            print(f"   Year 1 (C{total_row}): {cell_y1.value}")
            
            if cell_y1.value and '=SUM' in str(cell_y1.value):
                results.append(('Revenue Y1 수식', True, "SUM 패턴 정상"))
            else:
                results.append(('Revenue Y1 수식', False, f"SUM 아님: {cell_y1.value}"))
            
            # 세그먼트 수식 확인 (Row 5)
            seg_y1 = ws_rev['C5']
            print(f"\n   세그먼트 Year 1 (C5): {seg_y1.value}")
            
            if seg_y1.value and isinstance(seg_y1.value, str):
                # =B5*(1+$H$5) 패턴인지 확인
                if 'B5' in seg_y1.value and 'H5' in seg_y1.value:
                    results.append(('세그먼트 성장 수식', True, "B5*(1+$H$5) 패턴 정상 ✅"))
                elif 'C5' in seg_y1.value:
                    results.append(('세그먼트 성장 수식', False, "자기 참조 발견! ❌"))
                else:
                    results.append(('세그먼트 성장 수식', False, f"패턴 불명: {seg_y1.value}"))
    
    # 2. Cost_Structure 수식 확인
    if 'Cost_Structure' in wb_gen.sheetnames:
        ws_cost = wb_gen['Cost_Structure']
        
        print(f"\n2. Cost_Structure")
        
        # COGS 행 찾기
        cogs_row = None
        for row_idx in range(5, 10):
            cell = ws_cost[f'A{row_idx}']
            if cell.value and 'COGS' in str(cell.value) and 'Revenue' not in str(cell.value):
                cogs_row = row_idx
                break
        
        if cogs_row:
            # Revenue 행 (COGS 위)
            revenue_row = cogs_row - 1
            
            # COGS Year 0 수식
            cogs_y0 = ws_cost[f'B{cogs_row}']
            print(f"   COGS Year 0 (B{cogs_row}): {cogs_y0.value}")
            
            if cogs_y0.value and isinstance(cogs_y0.value, str):
                # =B{revenue_row}*(1-GrossMarginTarget) 패턴인지
                if f'B{revenue_row}' in cogs_y0.value or 'B5' in cogs_y0.value:
                    results.append(('COGS Year 0 수식', True, f"Revenue Row {revenue_row} 참조 정상 ✅"))
                elif f'C{revenue_row}' in cogs_y0.value or 'C5' in cogs_y0.value:
                    results.append(('COGS Year 0 수식', False, "잘못된 컬럼 참조 (다음 해) ❌"))
                else:
                    results.append(('COGS Year 0 수식', False, f"패턴 불명: {cogs_y0.value}"))
            
            # COGS Year 1 수식
            cogs_y1 = ws_cost[f'C{cogs_row}']
            print(f"   COGS Year 1 (C{cogs_row}): {cogs_y1.value}")
            
            if cogs_y1.value and isinstance(cogs_y1.value, str):
                if f'C{revenue_row}' in cogs_y1.value or 'C5' in cogs_y1.value:
                    results.append(('COGS Year 1 수식', True, f"Revenue Row {revenue_row} 참조 정상 ✅"))
                elif f'D{revenue_row}' in cogs_y1.value or 'D5' in cogs_y1.value:
                    results.append(('COGS Year 1 수식', False, "잘못된 컬럼 참조 (다음 해) ❌"))
    
    # 3. Dashboard 값 확인 (실제 계산 필요)
    print(f"\n3. Dashboard")
    print("   ⚠️ Excel에서 한 번 열고 저장해야 값 계산됨")
    print("   ⚠️ 현재는 수식만 있고 계산값 없음")
    
    # 결과 요약
    print("\n" + "="*70)
    print("📊 비교 검증 결과")
    print("="*70 + "\n")
    
    for name, passed, message in results:
        status = "✅" if passed else "❌"
        print(f"{status} {name}: {message}")
    
    passed_count = sum(1 for _, p, _ in results if p)
    total_count = len(results)
    
    print(f"\n총 {total_count}개 검증")
    print(f"통과: {passed_count}개")
    print(f"실패: {total_count - passed_count}개")
    
    all_passed = all(p for _, p, _ in results)
    
    if all_passed:
        print("\n✅ 수식 패턴 검증 통과!")
        print("\n💡 다음 단계:")
        print("   1. Excel에서 생성 파일 열기")
        print("   2. 파일 저장 (자동 계산)")
        print("   3. Golden과 값 비교")
    else:
        print("\n❌ 수식 패턴 오류 발견!")
        print("\n📋 수정 필요:")
        for name, passed, message in results:
            if not passed:
                print(f"   - {name}: {message}")
    
    return all_passed


def compare_unit_economics():
    """Unit Economics 비교"""
    
    print("\n" + "="*70)
    print("📊 Unit Economics 비교 검증")
    print("="*70 + "\n")
    
    generated = project_root / 'examples' / 'excel' / 'unit_economics_music_streaming_example_20251104.xlsx'
    golden = project_root / 'examples' / 'excel' / 'golden_unit_economics.xlsx'
    
    if not generated.exists() or not golden.exists():
        print("❌ 파일 없음")
        return False
    
    print(f"📁 생성 파일: {generated.name}")
    print(f"📁 Golden 파일: {golden.name}\n")
    
    # Golden 값
    wb_golden = load_workbook(golden, data_only=True)
    ws_golden = wb_golden['Golden_Values']
    
    ltv_expected = ws_golden['B11'].value  # LTV (평균)
    ratio_expected = ws_golden['B13'].value  # LTV/CAC
    
    print("📊 Golden 값 (정답):")
    print(f"   LTV: ₩{ltv_expected:,.0f}")
    print(f"   LTV/CAC: {ratio_expected:.2f}\n")
    
    # 생성 파일 수식 확인
    wb_gen = load_workbook(generated, data_only=False)
    
    results = []
    
    # LTV_Calculation 시트
    if 'LTV_Calculation' in wb_gen.sheetnames:
        ws_ltv = wb_gen['LTV_Calculation']
        
        print("🔍 생성 파일 수식 검증:")
        print("-"*70)
        print("\n1. LTV_Calculation")
        
        # LTV 평균 셀 찾기 (보통 B18)
        for row_idx in range(15, 25):
            cell = ws_ltv[f'A{row_idx}']
            if cell.value and '최종 LTV' in str(cell.value):
                ltv_cell = ws_ltv[f'B{row_idx}']
                print(f"   최종 LTV (B{row_idx}): {ltv_cell.value}")
                
                if ltv_cell.value and '=AVERAGE' in str(ltv_cell.value):
                    results.append(('LTV 평균 수식', True, "AVERAGE 패턴 정상 ✅"))
                else:
                    results.append(('LTV 평균 수식', False, f"AVERAGE 아님: {ltv_cell.value}"))
                break
    
    # LTV_CAC_Ratio 시트
    if 'LTV_CAC_Ratio' in wb_gen.sheetnames:
        ws_ratio = wb_gen['LTV_CAC_Ratio']
        
        print("\n2. LTV_CAC_Ratio")
        
        # Ratio 셀 찾기
        for row_idx in range(6, 12):
            cell = ws_ratio[f'A{row_idx}']
            if cell.value and 'LTV/CAC Ratio' in str(cell.value):
                ratio_cell = ws_ratio[f'B{row_idx}']
                print(f"   LTV/CAC Ratio (B{row_idx}): {ratio_cell.value}")
                
                if ratio_cell.value and '=IFERROR' in str(ratio_cell.value) and 'LTV' in str(ratio_cell.value) and 'CAC' in str(ratio_cell.value):
                    results.append(('LTV/CAC 수식', True, "LTV/CAC 패턴 정상 ✅"))
                else:
                    results.append(('LTV/CAC 수식', False, f"패턴 오류: {ratio_cell.value}"))
                break
    
    # 결과
    print("\n" + "="*70)
    print("📊 비교 검증 결과")
    print("="*70 + "\n")
    
    for name, passed, message in results:
        status = "✅" if passed else "❌"
        print(f"{status} {name}: {message}")
    
    all_passed = all(p for _, p, _ in results)
    return all_passed


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🎯 생성 Excel vs Golden Workbook 비교")
    print("="*70)
    print("\n전략: 수식 패턴 검증 (값은 Excel에서 열어야 계산됨)\n")
    
    results = []
    
    # 1. Financial Projection
    results.append(('Financial Projection', compare_financial_projection()))
    
    # 2. Unit Economics  
    results.append(('Unit Economics', compare_unit_economics()))
    
    # 최종
    print("\n" + "="*70)
    print("🏁 최종 비교 결과")
    print("="*70 + "\n")
    
    for name, passed in results:
        status = "✅" if passed else "❌"
        print(f"{status} {name}: {'통과' if passed else '실패'}")
    
    if all(p for _, p in results):
        print("\n✅ 수식 패턴 검증 모두 통과!")
        print("\n📋 실제 계산 값 확인:")
        print("   1. 생성된 Excel 파일 열기")
        print("   2. 파일 저장 (Ctrl+S)")
        print("   3. Golden Workbook과 값 비교")
        print("   4. 오차 < 1%이면 완전 정상")
        sys.exit(0)
    else:
        print("\n❌ 수식 패턴 오류 발견!")
        sys.exit(1)

