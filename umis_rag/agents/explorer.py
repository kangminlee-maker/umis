"""
Explorer RAG Agent Module

Explorer (Explorer) ì—ì´ì „íŠ¸ì˜ RAG ê¸°ë°˜ ê¸°íšŒ ë°œêµ´ ì‹œìŠ¤í…œì…ë‹ˆë‹¤.

í•µì‹¬ ê°œë…:
-----------
1. **Pattern Matching**: Observer ê´€ì°° â†’ ì‚¬ì—…ëª¨ë¸ íŒ¨í„´ ë§¤ì¹­
2. **Case Retrieval**: ìœ ì‚¬ ì‚°ì—… ì„±ê³µ ì‚¬ë¡€ ê²€ìƒ‰
3. **Multi-Stage Search**: ë‹¨ê³„ë³„ ì •ë°€ ê²€ìƒ‰
4. **Agent Collaboration**: Quantifier/Validatorê³¼ ìì—°ìŠ¤ëŸ¬ìš´ í˜‘ì—…

Explorerì˜ 7ë‹¨ê³„ í”„ë¡œì„¸ìŠ¤:
-----------------------
Phase 1: íŠ¸ë¦¬ê±° ì¸ì‹ (Observer ê´€ì°°ì—ì„œ ì‹œê·¸ë„ ì¶”ì¶œ)
Phase 2: íŒ¨í„´ ë§¤ì¹­ (ì‚¬ì—…ëª¨ë¸ + Disruption)
Phase 3: ì‚¬ë¡€ ê²€ìƒ‰ (ìœ ì‚¬ ì‚°ì—…/êµ¬ì¡°)
Phase 4: ì •ëŸ‰ ê²€ì¦ (Quantifier í˜‘ì—…)
Phase 5: ë°ì´í„° ê²€ì¦ (Validator í˜‘ì—…)
Phase 6: ê°€ì„¤ ìƒì„±
Phase 7: Guardian ê²€ì¦
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

from langchain_openai import OpenAIEmbeddings, ChatOpenAI
from langchain_community.vectorstores import Chroma
from langchain_core.documents import Document
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser

import sys
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from umis_rag.core.config import settings
from umis_rag.core.llm_provider import LLMProvider
from umis_rag.utils.logger import logger
from umis_rag.graph.hybrid_search import HybridSearch, HybridResult


class ExplorerRAG:
    """
    Explorer (Explorer) RAG Agent
    
    ì—­í• :
    -----
    - ì‹œì¥ ê¸°íšŒ ë°œêµ´
    - ì‚¬ì—…ëª¨ë¸ íŒ¨í„´ ì¸ì‹
    - ê²€ì¦ëœ ê°€ì„¤ ìƒì„±
    
    í•µì‹¬ ë©”ì„œë“œ:
    -----------
    - search_patterns(): íŠ¸ë¦¬ê±° â†’ íŒ¨í„´ ë§¤ì¹­
    - search_cases(): ìœ ì‚¬ ì‚¬ë¡€ ê²€ìƒ‰  
    - generate_hypothesis(): LLMìœ¼ë¡œ ê°€ì„¤ ìƒì„±
    - validate_with_framework(): ê²€ì¦ í”„ë ˆì„ì›Œí¬ ì ìš©
    
    í˜‘ì—…:
    -----
    - Quantifier: ì •ëŸ‰ ë°ì´í„° ìš”ì²­
    - Validator: ì¶œì²˜ ê²€ì¦ ìš”ì²­
    - Guardian: ìµœì¢… ê²€ì¦
    """
    
    def __init__(self, use_projected=False):
        """
        Explorer RAG ì—ì´ì „íŠ¸ ì´ˆê¸°í™”
        
        Args:
            use_projected: True = projected_index (v3.0 Dual-Index)
                          False = explorer_knowledge_base (ê¸°ì¡´, ê¸°ë³¸)
        """
        logger.info("Explorer RAG ì—ì´ì „íŠ¸ ì´ˆê¸°í™”")
        
        # Embeddings ì´ˆê¸°í™”
        self.embeddings = OpenAIEmbeddings(
            model=settings.embedding_model,
            openai_api_key=settings.openai_api_key
        )
        
        # ë²¡í„° ìŠ¤í† ì–´ ë¡œë“œ (v3.0 Dual-Index ì§€ì›!)
        collection_name = "projected_index" if use_projected else "explorer_knowledge_base"
        
        self.vectorstore = Chroma(
            collection_name=collection_name,
            embedding_function=self.embeddings,
            persist_directory=str(settings.chroma_persist_dir)
        )
        
        self.use_projected = use_projected
        
        # LLM ì´ˆê¸°í™” (ê°€ì„¤ ìƒì„±ìš©) - v7.8.1: llm_mode ì§€ì›
        self.llm = LLMProvider.create_llm()
        self.mode = settings.llm_mode
        
        logger.info(f"  âœ… ë²¡í„° ìŠ¤í† ì–´: {collection_name}")
        logger.info(f"  âœ… ì²­í¬ ìˆ˜: {self.vectorstore._collection.count()}ê°œ")
        logger.info(f"  ğŸ¯ LLM ëª¨ë“œ: {self.mode}")
        
        # Hybrid Search ì´ˆê¸°í™” (ì„ íƒì )
        self.hybrid_search = None
        try:
            from umis_rag.graph.connection import Neo4jConnection
            # Neo4j ì—°ê²° í…ŒìŠ¤íŠ¸
            test_conn = Neo4jConnection()
            if test_conn.verify_connection():
                self.hybrid_search = HybridSearch(graph_connection=test_conn)
                logger.info(f"  âœ… Hybrid Search í™œì„±í™” (Vector + Graph)")
            else:
                logger.warning(f"  âš ï¸  Neo4j ì—°ê²° ì‹¤íŒ¨ - Vectorë§Œ ì‚¬ìš©")
        except Exception as e:
            logger.warning(f"  âš ï¸  Hybrid Search ë¹„í™œì„± - Vectorë§Œ ì‚¬ìš©: {e}")
        logger.info(f"  âœ… LLM ëª¨ë¸: {settings.llm_model}")
    
    def search_patterns(
        self, 
        trigger_signals: str | List[str],
        top_k: int = 3,
        use_graph: bool = True  # v7.1.0: ê¸°ë³¸ê°’ True (Hybrid Search)
    ) -> List[tuple[Document, float]] | HybridResult:
        """
        v3.0: Projected Index ì§€ì›
        - use_projected=True â†’ agent_view í•„í„° ìë™
        """
        """
        íŠ¸ë¦¬ê±° ì‹œê·¸ë„ â†’ ì‚¬ì—…ëª¨ë¸ íŒ¨í„´ ë§¤ì¹­
        
        ì‚¬ìš© ì‹œì :
        ----------
        Observerê°€ ì‹œì¥ ê´€ì°°ì„ ì™„ë£Œí•˜ê³  íŠ¸ë¦¬ê±° ì‹œê·¸ë„ì„ ë°œê²¬í–ˆì„ ë•Œ
        
        ì˜ˆì‹œ:
        -----
        Input: "íŒŒí¸í™”ëœ ê³µê¸‰-ìˆ˜ìš”, ë†’ì€ ì¤‘ê°œ ë¹„ìš©"
        Output: [platform_business_model, ...]
        
        Parameters:
        -----------
        trigger_signals: íŠ¸ë¦¬ê±° ì‹œê·¸ë„ (ë¬¸ìì—´ ë˜ëŠ” ë¦¬ìŠ¤íŠ¸)
        top_k: ë°˜í™˜í•  íŒ¨í„´ ìˆ˜
        
        Returns:
        --------
        List of (Document, similarity_score)
        """
        logger.info(f"[Explorer] íŒ¨í„´ ë§¤ì¹­ ê²€ìƒ‰ ì‹œì‘")
        
        # ë¦¬ìŠ¤íŠ¸ë¥¼ ë¬¸ìì—´ë¡œ ë³€í™˜
        if isinstance(trigger_signals, list):
            query = ", ".join(trigger_signals)
        else:
            query = trigger_signals
        
        logger.info(f"  íŠ¸ë¦¬ê±°: {query[:100]}...")
        
        # v7.1.0: Hybrid Search ìš°ì„  (Knowledge Graph)
        if use_graph and self.hybrid_search:
            logger.info("  ğŸ” Hybrid Search (Vector + Graph)")
            hybrid_result = self.search_patterns_with_graph(query, top_k=top_k)
            
            # HybridResult â†’ List[tuple] ë³€í™˜ (ì¼ê´€ì„±)
            if hybrid_result and hasattr(hybrid_result, 'direct_matches'):
                # PatternMatch ê°ì²´ â†’ (Document, score) tuple ë³€í™˜
                converted = []
                for match in hybrid_result.direct_matches:
                    if hasattr(match, 'document') and hasattr(match, 'similarity'):
                        converted.append((match.document, match.similarity))
                    elif hasattr(match, 'doc') and hasattr(match, 'score'):
                        converted.append((match.doc, match.score))
                
                if converted:
                    logger.info(f"  âœ… Hybrid ê²°ê³¼ ë³€í™˜: {len(converted)}ê°œ")
                    return converted
            
            # Fallback to vector if conversion failed
            logger.warning("  âš ï¸ Hybrid ê²°ê³¼ ë³€í™˜ ì‹¤íŒ¨ â†’ Vectorë¡œ í´ë°±")
            use_graph = False
        
        # Fallback: Vectorë§Œ
        logger.info("  ğŸ” Vector Search")
        
        # íŒ¨í„´ ê°œìš”ë§Œ ê²€ìƒ‰ (íŠ¸ë¦¬ê±° ì‹œê·¸ë„ í¬í•¨ëœ ì²­í¬)
        results = self.vectorstore.similarity_search_with_score(
            query,
            k=top_k,
            filter={"chunk_type": "pattern_overview"}
        )
        
        logger.info(f"  âœ… {len(results)}ê°œ íŒ¨í„´ ë§¤ì¹­")
        for i, (doc, score) in enumerate(results, 1):
            pattern_id = doc.metadata.get("pattern_id", "N/A")
            logger.info(f"    #{i} {pattern_id} (ìœ ì‚¬ë„: {score:.4f})")
        
        return results
    
    def get_pattern_details(self, results: List[tuple]) -> List[Dict[str, Any]]:
        """
        ê²€ìƒ‰ ê²°ê³¼ tupleì„ ì‚¬ìš©í•˜ê¸° ì‰¬ìš´ dict í˜•ì‹ìœ¼ë¡œ ë³€í™˜
        
        Parameters:
        -----------
        results: search_patterns() ê²°ê³¼ List[(Document, score)]
        
        Returns:
        --------
        List[Dict] with keys: pattern_id, pattern_name, score, description, triggers, etc.
        """
        pattern_details = []
        
        for doc, score in results:
            metadata = doc.metadata
            detail = {
                'pattern_id': metadata.get('pattern_id', 'Unknown'),
                'pattern_name': metadata.get('pattern_name', 'Unknown'),
                'category': metadata.get('category', 'Unknown'),
                'score': float(score),
                'description': doc.page_content[:200] if doc.page_content else '',
                'metadata': metadata
            }
            pattern_details.append(detail)
        
        return pattern_details
    
    def search_patterns_with_graph(
        self,
        trigger_observation: str,
        top_k: int = 5,
        max_combinations: int = 10
    ) -> Optional[HybridResult]:
        """
        Hybrid Search: Vector + Graph í†µí•© ê²€ìƒ‰
        
        ì‚¬ìš© ì‹œì :
        ----------
        íŒ¨í„´ ë§¤ì¹­ê³¼ í•¨ê»˜ ê´€ë ¨ ì¡°í•©ê¹Œì§€ ë°œê²¬í•˜ê³  ì‹¶ì„ ë•Œ
        
        ì˜ˆì‹œ:
        -----
        Input: "ìŒì•… ìŠ¤íŠ¸ë¦¬ë° êµ¬ë… ì„œë¹„ìŠ¤"
        Output:
          Direct: [subscription_model, platform_model, ...]
          Combinations: [
            subscription + platform (Amazon Prime),
            subscription + licensing (Spotify),
            subscription + freemium (YouTube Premium)
          ]
        
        Parameters:
        -----------
        trigger_observation: Observer ê´€ì°° ë˜ëŠ” ì‹œì¥ ì„¤ëª…
        top_k: Vector ê²€ìƒ‰ ê²°ê³¼ ìˆ˜
        max_combinations: ìµœëŒ€ ì¡°í•© ìˆ˜
        
        Returns:
        --------
        HybridResult ë˜ëŠ” None (Hybrid Search ë¹„í™œì„± ì‹œ)
        """
        if not self.hybrid_search:
            logger.warning("  âš ï¸  Hybrid Search ë¹„í™œì„± - Vector ê²€ìƒ‰ë§Œ ì‚¬ìš©í•˜ì„¸ìš”")
            return None
        
        logger.info(f"[Explorer] Hybrid Search ì‹œì‘")
        logger.info(f"  ê´€ì°°: {trigger_observation[:100]}")
        
        # 1. Vector ê²€ìƒ‰ (use_graph=Falseë¡œ ì¬ê·€ ë°©ì§€!)
        vector_results = self.search_patterns(trigger_observation, top_k, use_graph=False)
        
        # 2. Hybrid ê²€ìƒ‰
        hybrid_result = self.hybrid_search.search(
            vector_results,
            max_combinations=max_combinations
        )
        
        # 3. ê²°ê³¼ ë¡œê¹…
        logger.info(f"  âœ… Direct matches: {len(hybrid_result.direct_matches)}")
        logger.info(f"  âœ… Combinations: {len(hybrid_result.combinations)}")
        logger.info(f"  âœ… Insights: {len(hybrid_result.insights)}")
        
        for insight in hybrid_result.insights:
            logger.info(f"    {insight}")
        
        return hybrid_result
    
    def search_cases(
        self,
        industry_or_pattern: str,
        pattern_id: Optional[str] = None,
        top_k: int = 5
    ) -> List[tuple[Document, float]]:
        """
        ìœ ì‚¬ ì‚°ì—…/êµ¬ì¡° ì„±ê³µ ì‚¬ë¡€ ê²€ìƒ‰
        
        ì‚¬ìš© ì‹œì :
        ----------
        íŒ¨í„´ ë§¤ì¹­ í›„, ì‹¤ì œ ì„±ê³µ ì‚¬ë¡€ë¥¼ ì°¾ì„ ë•Œ
        
        ì˜ˆì‹œ:
        -----
        Input: "ìŒì•… ìŠ¤íŠ¸ë¦¬ë°", pattern_id="subscription_model"
        Output: [ë„·í”Œë¦­ìŠ¤, ë©œë¡ , ìŠ¤í¬í‹°íŒŒì´, ...]
        
        Parameters:
        -----------
        industry_or_pattern: ì‚°ì—…ëª… ë˜ëŠ” ìœ ì‚¬ì„± ì„¤ëª…
        pattern_id: íŠ¹ì • íŒ¨í„´ì˜ ì‚¬ë¡€ë§Œ ê²€ìƒ‰ (ì„ íƒ)
        top_k: ë°˜í™˜í•  ì‚¬ë¡€ ìˆ˜
        """
        logger.info(f"[Explorer] ì‚¬ë¡€ ê²€ìƒ‰ ì‹œì‘")
        logger.info(f"  ì‚°ì—…/íŒ¨í„´: {industry_or_pattern[:100]}")
        
        # í•„í„° êµ¬ì„± (Chroma DB ë¬¸ë²•: AND ì—°ì‚°ì ì‚¬ìš©)
        if pattern_id:
            filter_dict = {
                "$and": [
                    {"chunk_type": "success_case"},
                    {"pattern_id": pattern_id}
                ]
            }
            logger.info(f"  í•„í„°: {pattern_id} íŒ¨í„´ì˜ ì‚¬ë¡€ë§Œ")
        else:
            filter_dict = {"chunk_type": "success_case"}
        
        results = self.vectorstore.similarity_search_with_score(
            industry_or_pattern,
            k=top_k,
            filter=filter_dict
        )
        
        logger.info(f"  âœ… {len(results)}ê°œ ì‚¬ë¡€ ë°œê²¬")
        for i, (doc, score) in enumerate(results, 1):
            company = doc.metadata.get("company", "N/A")
            logger.info(f"    #{i} {company} (ìœ ì‚¬ë„: {score:.4f})")
        
        return results
    
    def get_validation_framework(
        self,
        pattern_id: str
    ) -> Optional[Document]:
        """
        íŠ¹ì • íŒ¨í„´ì˜ ê²€ì¦ í”„ë ˆì„ì›Œí¬ ê°€ì ¸ì˜¤ê¸°
        
        ì‚¬ìš© ì‹œì :
        ----------
        ê°€ì„¤ ìƒì„± í›„, ì–´ë–»ê²Œ ê²€ì¦í• ì§€ í”„ë ˆì„ì›Œí¬ í•„ìš”í•  ë•Œ
        
        ì˜ˆì‹œ:
        -----
        Input: "subscription_model"
        Output: Quantifier/Validator/Observerì—ê²Œ ë¬¼ì–´ë³¼ ì²´í¬ë¦¬ìŠ¤íŠ¸
        """
        logger.info(f"[Explorer] ê²€ì¦ í”„ë ˆì„ì›Œí¬ ê²€ìƒ‰: {pattern_id}")
        
        results = self.vectorstore.similarity_search(
            f"{pattern_id} validation",
            k=1,
            filter={
                "$and": [
                    {"pattern_id": pattern_id},
                    {"chunk_type": "validation_framework"}
                ]
            }
        )
        
        if results:
            logger.info(f"  âœ… ê²€ì¦ í”„ë ˆì„ì›Œí¬ ë°œê²¬")
            return results[0]
        else:
            logger.warning(f"  âš ï¸  ê²€ì¦ í”„ë ˆì„ì›Œí¬ ì—†ìŒ")
            return None
    
    def generate_opportunity_hypothesis(
        self,
        observer_observation: str,
        matched_patterns: List[Document],
        success_cases: List[Document]
    ) -> str | Dict[str, Any]:
        """
        ê¸°íšŒ ê°€ì„¤ ìƒì„± (v7.7.0: Native/External ëª¨ë“œ ì§€ì›)
        
        ê°œë…:
        -----
        RAGì˜ í•µì‹¬! ê²€ìƒ‰ëœ ì •ë³´ + LLMì˜ ì¶”ë¡ 
        
        ëª¨ë“œë³„ ë™ì‘:
        -----------
        Native Mode (umis_mode='native'):
            - RAG ê²€ìƒ‰ ê²°ê³¼ë§Œ ì¤€ë¹„
            - Cursor LLMì´ ì§ì ‘ ë¶„ì„í•˜ë„ë¡ ê²°ê³¼ ë°˜í™˜
            - ë¹„ìš©: $0
        
        External Mode (umis_mode='external'):
            - RAG ê²€ìƒ‰ + OpenAI API í˜¸ì¶œ
            - ì™„ì„±ëœ ê°€ì„¤ ë°˜í™˜
            - ë¹„ìš©: ~$0.10/ìš”ì²­
        
        Parameters:
        -----------
        observer_observation: Observerì˜ ì‹œì¥ ê´€ì°° ë‚´ìš©
        matched_patterns: ë§¤ì¹­ëœ íŒ¨í„´ë“¤
        success_cases: ìœ ì‚¬ ì„±ê³µ ì‚¬ë¡€ë“¤
        
        Returns:
        --------
        Native ëª¨ë“œ: Dict (RAG ê²°ê³¼ + ì§€ì‹œì‚¬í•­)
        External ëª¨ë“œ: str (ì™„ì„±ëœ ê°€ì„¤ Markdown)
        """
        logger.info(f"[Explorer] ê°€ì„¤ ìƒì„± ì‹œì‘ (ëª¨ë“œ: {self.mode})")
        
        # ì»¨í…ìŠ¤íŠ¸ ì¡°ë¦½ (ëª¨ë“  ëª¨ë“œ ê³µí†µ)
        context = self._assemble_context(matched_patterns, success_cases)
        
        # ========================================
        # Native ëª¨ë“œ: RAG ê²°ê³¼ë§Œ ë°˜í™˜
        # ========================================
        if self.mode == "native":
            logger.info("  ğŸ¯ Native ëª¨ë“œ: RAG ê²°ê³¼ë§Œ ì¤€ë¹„ (Cursor LLMì´ ì²˜ë¦¬)")
            
            return {
                "mode": "native",
                "observer_observation": observer_observation,
                "rag_context": context,
                "matched_patterns_count": len(matched_patterns),
                "success_cases_count": len(success_cases),
                "instruction": (
                    "ìœ„ RAG ê²€ìƒ‰ ê²°ê³¼(rag_context)ë¥¼ ë°”íƒ•ìœ¼ë¡œ ê¸°íšŒ ê°€ì„¤ì„ ìƒì„±í•´ì£¼ì„¸ìš”.\n\n"
                    "í¬í•¨í•  ë‚´ìš©:\n"
                    "1. Observer ê´€ì°° ìš”ì•½\n"
                    "2. ë§¤ì¹­ëœ íŒ¨í„´ ë¶„ì„\n"
                    "3. ìœ ì‚¬ ì„±ê³µ ì‚¬ë¡€ ì‹œì‚¬ì \n"
                    "4. ê¸°íšŒ ê°€ì„¤ 3-5ê°œ (êµ¬ì¡°í™”)\n"
                    "5. ê° ê°€ì„¤ì˜ ê²€ì¦ ë°©í–¥"
                ),
                "next_step": "Cursor Composer/Chatì—ì„œ ìœ„ instructionì„ ë”°ë¼ ë¶„ì„í•˜ì„¸ìš”."
            }
        
        # ========================================
        # External ëª¨ë“œ: API í˜¸ì¶œ
        # ========================================
        else:
            logger.info("  ğŸŒ External ëª¨ë“œ: OpenAI API í˜¸ì¶œ")
        
        # Prompt êµ¬ì„±
        prompt = ChatPromptTemplate.from_messages([
            ("system", self._get_explorer_system_prompt()),
            ("user", self._get_hypothesis_generation_prompt())
        ])
        
        # LLM ì²´ì¸ êµ¬ì„±
        chain = prompt | self.llm | StrOutputParser()
        
        # ì‹¤í–‰
        logger.info("  â³ LLM ì¶”ë¡  ì¤‘...")
        hypothesis = chain.invoke({
            "observer_observation": observer_observation,
            "context": context
        })
        
        logger.info("  âœ… ê°€ì„¤ ìƒì„± ì™„ë£Œ")
        return hypothesis
    
    def _assemble_context(
        self,
        patterns: List[Document],
        cases: List[Document]
    ) -> str:
        """
        ê²€ìƒ‰ëœ ì •ë³´ë¥¼ LLM ì»¨í…ìŠ¤íŠ¸ë¡œ ì¡°ë¦½
        
        ê°œë…:
        -----
        RAG = Retrieval + Augmented Generation
        
        Retrieval (ê²€ìƒ‰):
          - ê´€ë ¨ íŒ¨í„´ 3ê°œ
          - ìœ ì‚¬ ì‚¬ë¡€ 5ê°œ
        
        Augmented (ì¦ê°•):
          - ì´ ì •ë³´ë¥¼ LLMì—ê²Œ ì»¨í…ìŠ¤íŠ¸ë¡œ ì œê³µ
          - LLMì´ ì´ë¥¼ ê¸°ë°˜ìœ¼ë¡œ ì¶”ë¡ 
        """
        context = "# ê²€ìƒ‰ëœ íŒ¨í„´\n\n"
        
        for i, doc in enumerate(patterns, 1):
            pattern_id = doc.metadata.get("pattern_id", "N/A")
            context += f"## íŒ¨í„´ {i}: {pattern_id}\n"
            context += doc.page_content[:500] + "...\n\n"
        
        context += "# ìœ ì‚¬ ì„±ê³µ ì‚¬ë¡€\n\n"
        
        for i, doc in enumerate(cases, 1):
            company = doc.metadata.get("company", "N/A")
            context += f"## ì‚¬ë¡€ {i}: {company}\n"
            context += doc.page_content[:500] + "...\n\n"
        
        return context
    
    def _get_explorer_system_prompt(self) -> str:
        """Explorer ì—ì´ì „íŠ¸ ì‹œìŠ¤í…œ í”„ë¡¬í”„íŠ¸"""
        return """ë‹¹ì‹ ì€ Explorerì…ë‹ˆë‹¤. UMISì˜ Explorer ì—ì´ì „íŠ¸ë¡œì„œ ì‹œì¥ ê¸°íšŒë¥¼ ë°œêµ´í•˜ëŠ” ì „ë¬¸ê°€ì…ë‹ˆë‹¤.

ë‹¹ì‹ ì˜ ì—­í• :
- Observerì˜ ì‹œì¥ ê´€ì°°ì„ ë°›ì•„ ê¸°íšŒ íŒ¨í„´ ì¸ì‹
- ê²€ì¦ëœ ì‚¬ì—…ëª¨ë¸ íŒ¨í„´ 7ê°œ ë³´ìœ 
- 1ë“± ì¶”ì›” íŒ¨í„´ 5ê°œ ë³´ìœ 
- 30+ ì„±ê³µ ì‚¬ë¡€ ë°ì´í„°ë² ì´ìŠ¤ í™œìš©

ë‹¹ì‹ ì˜ ê°•ì :
- êµ¬ì¡°ì  ì‚¬ê³  (íŒ¨í„´ ì¸ì‹)
- ì°½ì˜ì  ì‘ìš© (íŒ¨í„´ â†’ ìš°ë¦¬ ì‹œì¥ ì ìš©)
- ê²€ì¦ ì¤‘ì‹¬ (ê·¼ê±° ì—†ëŠ” ê°€ì„¤ ì•ˆ ë§Œë“¦)

ì‘ì—… ë°©ì‹:
1. Observer ê´€ì°°ì—ì„œ íŠ¸ë¦¬ê±° ì‹œê·¸ë„ ì¶”ì¶œ
2. ë§¤ì¹­ë˜ëŠ” íŒ¨í„´ ì°¾ê¸° (RAG ê²€ìƒ‰ë¨)
3. ìœ ì‚¬ ì‚¬ë¡€ì—ì„œ í•™ìŠµ (RAG ê²€ìƒ‰ë¨)
4. ìš°ë¦¬ ì‹œì¥ì— ë§ê²Œ ì¡°ì •
5. ê²€ì¦ ê°€ëŠ¥í•œ ê°€ì„¤ ìƒì„±

ì¤‘ìš”: 
- ëª¨ë“  ì£¼ì¥ì— ê·¼ê±° í•„ìš” (íŒ¨í„´/ì‚¬ë¡€ ì¸ìš©)
- ì¶”ì •ì¹˜ëŠ” ëª…í™•íˆ í‘œì‹œ
- Quantifier/Validator í˜‘ì—… ëª…ì‹œ
"""
    
    def _get_hypothesis_generation_prompt(self) -> str:
        """ê°€ì„¤ ìƒì„± í”„ë¡¬í”„íŠ¸"""
        return """# ì„ë¬´: ê¸°íšŒ ê°€ì„¤ ìƒì„±

## Observerì˜ ì‹œì¥ ê´€ì°°
{observer_observation}

## ê²€ìƒ‰ëœ ì •ë³´ (RAG)
{context}

## ì§€ì‹œì‚¬í•­

ìœ„ ì •ë³´ë¥¼ ë°”íƒ•ìœ¼ë¡œ **ê²€ì¦ ê°€ëŠ¥í•œ ê¸°íšŒ ê°€ì„¤**ì„ ìƒì„±í•˜ì„¸ìš”.

êµ¬ì¡°:
1. **íŒ¨í„´ ë§¤ì¹­**: ì–´ë–¤ íŒ¨í„´ì´ ì ìš© ê°€ëŠ¥í•œê°€?
2. **ê¸°íšŒ ë…¼ë¦¬**: 
   - ë¬¸ì œ (Observer ê´€ì°°)
   - í•´ê²° ë°©ì•ˆ (íŒ¨í„´ ì ìš©)
   - ê°€ì¹˜ ì œì•ˆ
3. **ìœ ì‚¬ ì‚¬ë¡€ í•™ìŠµ**: ì„±ê³µ ì‚¬ë¡€ì—ì„œ ë°°ìš¸ ì 
4. **ì‹œì¥ ê·œëª¨ ì¶”ì •** (Quantifierì—ê²Œ ìš”ì²­í•  ë‚´ìš© ëª…ì‹œ)
5. **ë°ì´í„° ê²€ì¦** (Validatorì—ê²Œ í™•ì¸í•  ë‚´ìš© ëª…ì‹œ)
6. **ì‹¤í–‰ ê°€ëŠ¥ì„±**: CSF, ë‚œì´ë„, ë¦¬ìŠ¤í¬

ë°˜ë“œì‹œ:
- íŒ¨í„´/ì‚¬ë¡€ ì¸ìš© (chunk_id ëª…ì‹œ)
- ì¶”ì •ì¹˜ í‘œì‹œ
- ê²€ì¦ í•„ìš” í•­ëª© ëª…í™•íˆ
"""


class ExplorerAgenticRAG(ExplorerRAG):
    """
    Explorer Agentic RAG (ììœ¨ ì‹¤í–‰)
    
    ê°œë…:
    -----
    Agentê°€ ìŠ¤ìŠ¤ë¡œ íŒë‹¨í•˜ë©° Toolì„ ì‚¬ìš©í•©ë‹ˆë‹¤.
    
    Tools:
    ------
    1. search_patterns: íŒ¨í„´ ê²€ìƒ‰
    2. search_cases: ì‚¬ë¡€ ê²€ìƒ‰
    3. get_validation: ê²€ì¦ í”„ë ˆì„ì›Œí¬
    4. ask_quantifier: Quantifierì—ê²Œ ì§ˆë¬¸
    5. ask_validator: Validatorì—ê²Œ ì§ˆë¬¸
    
    ììœ¨ì„±:
    -------
    Explorerê°€ í•„ìš”í•œ Toolì„ ì„ íƒí•˜ì—¬ ì‹¤í–‰
    "Quantifierì—ê²Œ ë­˜ ë¬¼ì–´ë³¼ê¹Œ?" ìŠ¤ìŠ¤ë¡œ íŒë‹¨
    """
    
    def __init__(self):
        super().__init__()
        
        # Agent Tools ì •ì˜
        self.agent_tools = self._initialize_agent_tools()
        
        if self.agent_tools:
            logger.info(f"  â†’ Agentic ëª¨ë“œ: {len(self.agent_tools)}ê°œ ë„êµ¬ í™œì„±í™”")
        else:
            logger.info("  â†’ Agentic ëª¨ë“œ: í–¥í›„ êµ¬í˜„ ì˜ˆì • (LangChain í†µí•©)")
    
    def _initialize_agent_tools(self) -> Optional[List]:
        """
        LangChain Agent Tools ì´ˆê¸°í™”
        
        Returns:
            List of LangChain tools or None if not available
        
        Tools:
            - search_patterns: RAGì—ì„œ íŒ¨í„´ ê²€ìƒ‰
            - search_cases: RAGì—ì„œ ì‚¬ë¡€ ê²€ìƒ‰
            - ask_quantifier: Quantifierì—ê²Œ ì§ˆë¬¸
            - ask_validator: Validatorì—ê²Œ ì§ˆë¬¸
            - generate_hypothesis: ê°€ì„¤ ìƒì„±
        """
        
        try:
            from langchain.tools import Tool
            from langchain.agents import initialize_agent, AgentType
            
            tools = []
            
            # Tool 1: Pattern Search
            tools.append(Tool(
                name="search_patterns",
                func=self._tool_search_patterns,
                description="Search for business patterns in RAG database. Input: pattern description"
            ))
            
            # Tool 2: Case Search
            tools.append(Tool(
                name="search_cases",
                func=self._tool_search_cases,
                description="Search for case studies in RAG database. Input: case description"
            ))
            
            # Tool 3: Ask Quantifier
            tools.append(Tool(
                name="ask_quantifier",
                func=self._tool_ask_quantifier,
                description="Ask Quantifier agent to estimate a value. Input: estimation question"
            ))
            
            # Tool 4: Ask Validator
            tools.append(Tool(
                name="ask_validator",
                func=self._tool_ask_validator,
                description="Ask Validator agent to verify data. Input: validation question"
            ))
            
            # Tool 5: Generate Hypothesis
            tools.append(Tool(
                name="generate_hypothesis",
                func=self._tool_generate_hypothesis,
                description="Generate business hypothesis. Input: hypothesis context"
            ))
            
            return tools
        
        except ImportError:
            logger.debug("  â„¹ï¸  LangChain ë¯¸ì„¤ì¹˜ (pip install langchain)")
            return None
        except Exception as e:
            logger.warning(f"  âš ï¸ Agent Tools ì´ˆê¸°í™” ì‹¤íŒ¨: {e}")
            return None
    
    def _tool_search_patterns(self, query: str) -> str:
        """Tool: RAG íŒ¨í„´ ê²€ìƒ‰"""
        try:
            if hasattr(self, 'pattern_store') and self.pattern_store:
                results = self.pattern_store.similarity_search(query, k=3)
                return "\n".join([r.page_content for r in results[:3]])
        except Exception:
            pass
        return "No patterns found"
    
    def _tool_search_cases(self, query: str) -> str:
        """Tool: RAG ì‚¬ë¡€ ê²€ìƒ‰"""
        try:
            if hasattr(self, 'case_store') and self.case_store:
                results = self.case_store.similarity_search(query, k=3)
                return "\n".join([r.page_content for r in results[:3]])
        except Exception:
            pass
        return "No cases found"
    
    def _tool_ask_quantifier(self, question: str) -> str:
        """Tool: Quantifier í˜‘ì—…"""
        try:
            from umis_rag.agents.quantifier import get_quantifier_rag
            quantifier = get_quantifier_rag()
            result = quantifier.estimate(question)
            if result:
                return f"Estimate: {result.value} {result.unit}"
        except Exception:
            pass
        return "Quantifier unavailable"
    
    def _tool_ask_validator(self, question: str) -> str:
        """Tool: Validator í˜‘ì—…"""
        try:
            from umis_rag.agents.validator import get_validator_rag
            validator = get_validator_rag()
            result = validator.validate(question)
            if result:
                return f"Validation: {result}"
        except Exception:
            pass
        return "Validator unavailable"
    
    def _tool_generate_hypothesis(self, context: str) -> str:
        """Tool: ê°€ì„¤ ìƒì„±"""
        # ê°„ë‹¨í•œ íŒ¨í„´ ê¸°ë°˜ ê°€ì„¤ ìƒì„±
        return f"Hypothesis: Based on {context}, consider market dynamics and trends."
    
    def autonomous_discovery(
        self,
        observer_report: str
    ) -> Dict[str, Any]:
        """
        ì™„ì „ ììœ¨ ê¸°íšŒ ë°œêµ´
        
        ê°œë…:
        -----
        Explorerê°€ Observer ë¦¬í¬íŠ¸ë§Œ ë°›ê³ 
        ìŠ¤ìŠ¤ë¡œ íŒë‹¨í•˜ë©°:
        1. í•„ìš”í•œ íŒ¨í„´ ê²€ìƒ‰
        2. í•„ìš”í•œ ì‚¬ë¡€ ê²€ìƒ‰
        3. Quantifier/Validatorì—ê²Œ ì§ˆë¬¸
        4. ê°€ì„¤ ìƒì„±
        
        í˜„ì¬:
        -----
        ê¸°ë³¸ ì›Œí¬í”Œë¡œìš° êµ¬í˜„
        í–¥í›„ LangChain Agentë¡œ í™•ì¥
        """
        logger.info("[Explorer] ììœ¨ ê¸°íšŒ ë°œêµ´ ì‹œì‘")
        
        # Phase 1-3: íŒ¨í„´ ë° ì‚¬ë¡€ ê²€ìƒ‰ (ê¸°ë³¸ ì›Œí¬í”Œë¡œìš°)
        patterns = self.search_patterns(observer_report, top_k=2)
        
        # ê°€ì¥ ë§¤ì¹­ëœ íŒ¨í„´ìœ¼ë¡œ ì‚¬ë¡€ ê²€ìƒ‰
        best_pattern_id = patterns[0][0].metadata.get("pattern_id")
        cases = self.search_cases(
            observer_report,
            pattern_id=best_pattern_id,
            top_k=3
        )
        
        # Phase 4-6: ê°€ì„¤ ìƒì„±
        hypothesis = self.generate_opportunity_hypothesis(
            observer_observation=observer_report,
            matched_patterns=[p[0] for p in patterns],
            success_cases=[c[0] for c in cases]
        )
        
        return {
            "matched_patterns": patterns,
            "success_cases": cases,
            "hypothesis": hypothesis
        }
    
    # ========================================
    # Strategy Playbook (v7.10.0 Gap #3)
    # ========================================
    
    def generate_strategy_playbook(
        self,
        validated_opportunity: Dict[str, Any],
        market_context: Dict[str, Any],
        quantified_market: Dict[str, Any],
        project_name: str = "default_project"
    ) -> Dict[str, Any]:
        """
        ê²€ì¦ëœ ê¸°íšŒ â†’ ì‹¤í–‰ ê°€ëŠ¥í•œ ì „ëµ Playbook ìƒì„± (v7.10.0)
        
        Q14 (ì–´ë–»ê²Œ ëš«ì–´ì•¼í•˜ëŠ”ë°?): 85% â†’ 95%+
        Q15 (ë­˜ í•´ì•¼í•˜ëŠ”ë°?): 60% â†’ 80%+
        
        Args:
            validated_opportunity: 7-Step ì™„ë£Œëœ ê¸°íšŒ
                {
                    'opportunity_id': 'OPP_XXX',
                    'title': 'êµ¬ë… ëª¨ë¸ í”¼ì•„ë…¸ ì„œë¹„ìŠ¤',
                    'value_proposition': 'ì´ˆê¸° ë¶€ë‹´ ì—†ì´ í”¼ì•„ë…¸ ì‹œì‘',
                    'target_customer': 'í”¼ì•„ë…¸ ì…ë¬¸ì (20-40ëŒ€)',
                    'core_features': [ê¸°ëŠ¥ ë¦¬ìŠ¤íŠ¸],
                    'revenue_model': 'ì›” êµ¬ë…',
                    'unit_economics': {
                        'arpu': 120000,
                        'cac': 180000,
                        'ltv': 2400000,
                        'churn': 0.05
                    }
                }
            
            market_context: Observer êµ¬ì¡° ë¶„ì„
                {
                    'market_structure': {...},
                    'inefficiencies': [...],
                    'competitors': [...]
                }
            
            quantified_market: Quantifier SAM ê³„ì‚°
                {
                    'sam': 1300,  # ì–µì›
                    'target_share': 0.05,
                    'unit_economics': {...}
                }
            
            project_name: í”„ë¡œì íŠ¸ ì´ë¦„ (íŒŒì¼ëª… ìƒì„±ìš©)
        
        Returns:
            {
                'gtm_strategy': {...},
                'product_roadmap': {...},
                'resource_plan': {...},
                'execution_milestones': {...},
                'risk_mitigation': {...},
                'markdown_path': 'strategy_playbook.md',
                'excel_path': 'strategy_playbook.xlsx'
            }
        """
        
        logger.info(f"[Explorer] Strategy Playbook ìƒì„±: {validated_opportunity.get('title')}")
        
        # Step 1: GTM Strategy
        logger.info("  Step 1/7: GTM Strategy")
        gtm = self._design_gtm_strategy(
            validated_opportunity, market_context, quantified_market
        )
        
        # Step 2: Product Roadmap
        logger.info("  Step 2/7: Product Roadmap (RICE)")
        roadmap = self._prioritize_features(
            validated_opportunity, market_context, quantified_market
        )
        
        # Step 3: Resource Plan
        logger.info("  Step 3/7: Resource Plan")
        resources = self._plan_resources(
            quantified_market, validated_opportunity
        )
        
        # Step 4: Execution Milestones
        logger.info("  Step 4/7: Execution Milestones")
        milestones = self._set_milestones(
            roadmap, resources, quantified_market, validated_opportunity
        )
        
        # Step 5: Risk Mitigation
        logger.info("  Step 5/7: Risk Assessment")
        risks = self._assess_and_mitigate_risks(
            validated_opportunity, market_context, quantified_market
        )
        
        # Step 6: Markdown ìƒì„±
        logger.info("  Step 6/7: Markdown ìƒì„±")
        markdown_path = self._generate_playbook_markdown(
            validated_opportunity, gtm, roadmap, resources, 
            milestones, risks, project_name
        )
        
        # Step 7: Excel ìƒì„±
        logger.info("  Step 7/7: Excel ìƒì„±")
        excel_path = self._generate_playbook_excel(
            validated_opportunity, gtm, roadmap, resources,
            milestones, risks, project_name
        )
        
        logger.info(f"  âœ… Strategy Playbook ì™„ë£Œ!")
        logger.info(f"    - Markdown: {markdown_path}")
        logger.info(f"    - Excel: {excel_path}")
        
        return {
            'gtm_strategy': gtm,
            'product_roadmap': roadmap,
            'resource_plan': resources,
            'execution_milestones': milestones,
            'risk_mitigation': risks,
            'markdown_path': markdown_path,
            'excel_path': excel_path
        }
    
    def _design_gtm_strategy(
        self,
        opportunity: Dict,
        market_context: Dict,
        quantified: Dict
    ) -> Dict[str, Any]:
        """GTM (Go-to-Market) ì „ëµ ì„¤ê³„"""
        
        sam = quantified['sam']
        target_share = quantified.get('target_share', 0.05)
        target_revenue = sam * target_share
        
        arpu = opportunity['unit_economics'].get('arpu', 100000)
        cac = opportunity['unit_economics'].get('cac', 150000)
        ltv = opportunity['unit_economics'].get('ltv', 2000000)
        
        # Target customers ê³„ì‚°
        target_customers_annual = int((target_revenue * 100000000) / (arpu * 12))
        
        # Customer Acquisition Channels
        channels = [
            {
                'channel': 'Direct Sales',
                'priority': 1,
                'cac_estimate': cac,
                'rationale': 'ì´ˆê¸° ê³ ê° ë°€ì°©, í”¼ë“œë°± ìˆ˜ì§‘',
                'timeline': 'Month 1-6'
            },
            {
                'channel': 'Digital Marketing',
                'priority': 2,
                'cac_estimate': int(cac * 0.7),
                'rationale': 'ìŠ¤ì¼€ì¼ì—…, ìë™í™” ê°€ëŠ¥',
                'timeline': 'Month 3+'
            }
        ]
        
        # Funnel
        monthly_target = target_customers_annual // 12
        funnel = {
            'awareness': int(monthly_target / 0.03),
            'consideration': int(monthly_target / 0.03 * 0.30),
            'conversion': monthly_target,
            'target_cac': cac
        }
        
        # Distribution
        distribution = {
            'primary_channel': 'Direct (ì˜¨ë¼ì¸)',
            'channel_mix': {'direct': '70%', 'partnership': '30%'},
            'partnerships': []
        }
        
        # Pricing
        competitors = market_context.get('competitors', [])
        competitor_comparison = []
        for comp in competitors[:3]:
            competitor_comparison.append({
                'competitor': comp.get('name', 'Competitor'),
                'price': f'ì›” {arpu * 1.25 / 10000:.0f}ë§Œì›',
                'our_price': f'ì›” {arpu / 10000:.0f}ë§Œì›',
                'differential': '-20%'
            })
        
        pricing = {
            'pricing_model': opportunity.get('revenue_model', 'êµ¬ë…'),
            'price_point': arpu,
            'pricing_strategy': 'Value-based',
            'competitor_comparison': competitor_comparison
        }
        
        # Marketing
        marketing = {
            'positioning': opportunity['value_proposition'],
            'key_message': opportunity['value_proposition'],
            'content_strategy': ['Blog', 'YouTube', 'SNS'],
            'budget_allocation': {
                'digital_ads': '40%',
                'content': '30%',
                'partnership': '20%',
                'ê¸°íƒ€': '10%'
            }
        }
        
        return {
            'customer_acquisition': {
                'target_segment': opportunity['target_customer'],
                'segment_size': target_customers_annual,
                'channels': channels,
                'funnel': funnel
            },
            'distribution': distribution,
            'pricing': pricing,
            'marketing_approach': marketing
        }
    
    def _prioritize_features(
        self,
        opportunity: Dict,
        market_context: Dict,
        quantified: Dict
    ) -> Dict[str, Any]:
        """Product Roadmap ìƒì„± (RICE Framework)"""
        
        features = opportunity.get('core_features', [])
        
        if not features:
            # ê¸°ë³¸ features ì œì•ˆ
            features = [
                {'name': 'ì‚¬ìš©ì ê°€ì…/ì¸ì¦', 'type': 'core', 'complexity': 'simple'},
                {'name': 'í•µì‹¬ ê¸°ëŠ¥ #1', 'type': 'core', 'complexity': 'medium'},
                {'name': 'í•µì‹¬ ê¸°ëŠ¥ #2', 'type': 'core', 'complexity': 'medium'},
                {'name': 'ê²°ì œ ì‹œìŠ¤í…œ', 'type': 'core', 'complexity': 'medium'},
                {'name': 'ëŒ€ì‹œë³´ë“œ', 'type': 'frequent', 'complexity': 'simple'}
            ]
        
        prioritized = []
        sam = quantified['sam']
        target_share = quantified.get('target_share', 0.05)
        arpu = opportunity['unit_economics'].get('arpu', 100000)
        
        total_customers = int((sam * target_share * 100000000) / (arpu * 12))
        monthly_users = total_customers // 12
        
        for feature in features:
            # RICE ê³„ì‚°
            feature_type = feature.get('type', 'core')
            
            # Reach
            if feature_type == 'core':
                reach = monthly_users
            elif feature_type == 'frequent':
                reach = int(monthly_users * 0.70)
            else:
                reach = int(monthly_users * 0.30)
            
            # Impact (ê°„ë‹¨ ë²„ì „)
            if feature_type == 'core':
                impact = 3
            elif feature.get('name') and 'ê²°ì œ' in feature.get('name'):
                impact = 3
            else:
                impact = 2
            
            # Confidence
            confidence = 80  # Default
            if feature.get('validated'):
                confidence = 95
            
            # Effort
            complexity = feature.get('complexity', 'medium')
            effort_map = {'simple': 0.5, 'medium': 1.5, 'complex': 3.0}
            effort = effort_map.get(complexity, 1.5)
            
            # RICE Score
            rice_score = (reach * impact * (confidence / 100)) / effort
            
            prioritized.append({
                'feature': feature.get('name', f'Feature {len(prioritized)+1}'),
                'description': feature.get('description', ''),
                'reach': reach,
                'impact': impact,
                'confidence': confidence,
                'effort': effort,
                'rice_score': round(rice_score, 1),
                'priority': 0
            })
        
        # ì ìˆ˜ìˆœ ì •ë ¬
        prioritized.sort(key=lambda x: x['rice_score'], reverse=True)
        for idx, item in enumerate(prioritized, 1):
            item['priority'] = idx
        
        # MVP / Phase 2 / Phase 3 ë¶„ë¥˜
        mvp = prioritized[:3]
        phase2 = prioritized[3:7] if len(prioritized) > 3 else []
        phase3 = prioritized[7:] if len(prioritized) > 7 else []
        
        return {
            'mvp': {
                'features': mvp,
                'timeline': '3ê°œì›”',
                'total_effort': sum([f['effort'] for f in mvp]),
                'description': 'Must-have í•µì‹¬ ê¸°ëŠ¥'
            },
            'phase_2': {
                'features': phase2,
                'timeline': '6ê°œì›”',
                'total_effort': sum([f['effort'] for f in phase2]),
                'description': 'í™•ì¥ ê¸°ëŠ¥'
            },
            'phase_3': {
                'features': phase3,
                'timeline': '12ê°œì›”',
                'total_effort': sum([f['effort'] for f in phase3]),
                'description': 'ì„±ìˆ™ ê¸°ëŠ¥'
            },
            'all_features': prioritized
        }
    
    def _plan_resources(
        self,
        quantified: Dict,
        opportunity: Dict
    ) -> Dict[str, Any]:
        """Resource Plan ìƒì„±"""
        
        target_revenue = quantified['sam'] * quantified.get('target_share', 0.05)
        
        # Team Structure
        team_3 = [
            {'role': 'CEO/Founder', 'count': 1, 'salary': 0},
            {'role': 'ê°œë°œ', 'count': 2, 'salary': 6000000},
            {'role': 'ë””ìì¸', 'count': 1, 'salary': 5000000},
            {'role': 'ë§ˆì¼€íŒ…', 'count': 1, 'salary': 5500000}
        ]
        
        team_6 = [
            {'role': 'CEO/Founder', 'count': 1, 'salary': 0},
            {'role': 'ê°œë°œ', 'count': 4, 'salary': 6000000},
            {'role': 'ë””ìì¸', 'count': 1, 'salary': 5000000},
            {'role': 'ë§ˆì¼€íŒ…/ì˜ì—…', 'count': 3, 'salary': 5500000},
            {'role': 'CS', 'count': 1, 'salary': 4500000}
        ]
        
        team_12 = [
            {'role': 'Executive', 'count': 2, 'salary': 0},
            {'role': 'ê°œë°œ', 'count': 8, 'salary': 6000000},
            {'role': 'ë§ˆì¼€íŒ…/ì˜ì—…', 'count': 6, 'salary': 5500000},
            {'role': 'CS/ìš´ì˜', 'count': 3, 'salary': 4500000},
            {'role': 'ë°ì´í„°/ë¶„ì„', 'count': 2, 'salary': 6500000}
        ]
        
        # Budget ê³„ì‚°
        def calc_budget(team):
            salary = sum([t['count'] * t['salary'] for t in team])
            opex = salary * 0.50
            return {'salary': salary, 'opex': opex, 'total': salary + opex}
        
        budget_3 = calc_budget(team_3)
        budget_6 = calc_budget(team_6)
        budget_12 = calc_budget(team_12)
        
        # Key Hires
        key_hires = [
            {'role': 'CTO/Tech Lead', 'priority': 1, 'timing': 'Month 1'},
            {'role': 'Product Manager', 'priority': 2, 'timing': 'Month 3'},
            {'role': 'Sales Lead', 'priority': 3, 'timing': 'Month 6'},
            {'role': 'Marketing Lead', 'priority': 4, 'timing': 'Month 6'}
        ]
        
        return {
            'team_structure': {
                'month_3': team_3,
                'month_6': team_6,
                'month_12': team_12
            },
            'budget': {
                'month_3': budget_3,
                'month_6': budget_6,
                'month_12': budget_12,
                'cumulative_burn': {
                    'to_3': budget_3['total'] * 3,
                    'to_6': budget_3['total'] * 3 + budget_6['total'] * 3,
                    'to_12': budget_3['total'] * 3 + budget_6['total'] * 3 + budget_12['total'] * 6
                }
            },
            'key_hires': key_hires
        }
    
    def _set_milestones(
        self,
        roadmap: Dict,
        resources: Dict,
        quantified: Dict,
        opportunity: Dict
    ) -> Dict[str, Any]:
        """3/6/12ê°œì›” Milestone ì„¤ì •"""
        
        sam = quantified['sam']
        target_share = quantified.get('target_share', 0.05)
        target_revenue_annual = sam * target_share
        arpu = opportunity['unit_economics'].get('arpu', 100000)
        
        # Month 3: MVP
        customers_3 = max(100, int((target_revenue_annual * 0.01 * 100000000) / (arpu * 12)))
        mrr_3 = customers_3 * arpu
        
        # Month 6: PMF
        customers_6 = customers_3 * 5
        mrr_6 = customers_6 * arpu
        
        # Month 12: Scale
        customers_12 = customers_6 * 6
        arr_12 = int(target_revenue_annual * 0.30)
        
        return {
            'month_3': {
                'milestone': 'MVP ëŸ°ì¹­',
                'metrics': {
                    'customers': customers_3,
                    'mrr': f'{mrr_3/100000000:.1f}ì–µ',
                    'churn': '< 10%'
                },
                'key_activities': [
                    'MVP ê°œë°œ ì™„ë£Œ',
                    f'Beta í…ŒìŠ¤íŠ¸ ({customers_3//2}ëª…)',
                    f'ì²« {customers_3}ëª… ê³ ê° í™•ë³´'
                ],
                'success_criteria': [
                    'Product-Market Fit ì´ˆê¸° ê²€ì¦',
                    'Churn < 10%',
                    'NPS > 40'
                ]
            },
            'month_6': {
                'milestone': 'PMF ê²€ì¦',
                'metrics': {
                    'customers': customers_6,
                    'mrr': f'{mrr_6/100000000:.1f}ì–µ',
                    'churn': '< 7%'
                },
                'key_activities': [
                    'Phase 2 ê¸°ëŠ¥ ì¶œì‹œ',
                    'íŒŒíŠ¸ë„ˆì‹­ 3ê°œ í™•ë³´',
                    f'{customers_6}ëª… ëŒíŒŒ'
                ],
                'success_criteria': [
                    'PMF í™•ì • (ì¬êµ¬ë§¤ > 60%)',
                    'LTV/CAC > 2.0',
                    'Churn < 7%'
                ]
            },
            'month_12': {
                'milestone': 'ìŠ¤ì¼€ì¼ì—… ì¤€ë¹„',
                'metrics': {
                    'customers': customers_12,
                    'arr': f'{arr_12:.0f}ì–µ',
                    'churn': '< 5%'
                },
                'key_activities': [
                    'Phase 3 ê¸°ëŠ¥ ì¶œì‹œ',
                    'ì‹œë¦¬ì¦ˆ A íˆ¬ì ìœ ì¹˜',
                    'íŒ€ í™•ì¥ (20ëª…)'
                ],
                'success_criteria': [
                    f'ARR {arr_12:.0f}ì–µ ë‹¬ì„±',
                    'Rule of 40 > 40%',
                    'ì‹œì¥ ì ìœ ìœ¨ 1%'
                ]
            }
        }
    
    def _assess_and_mitigate_risks(
        self,
        opportunity: Dict,
        market_context: Dict,
        quantified: Dict
    ) -> Dict[str, Any]:
        """ë¦¬ìŠ¤í¬ í‰ê°€ ë° ëŒ€ì‘"""
        
        risks = []
        
        # Risk 1: ê²½ìŸì‚¬ ê°€ê²© ì¸í•˜
        competitors = market_context.get('competitors', [])
        if len(competitors) >= 3:
            risks.append({
                'risk_id': 'RISK_001',
                'category': 'market',
                'risk': 'ê²½ìŸì‚¬ ê°€ê²© ì¸í•˜',
                'probability': 'high',
                'impact': 'high',
                'severity': 'critical',
                'mitigation': [
                    'ì°¨ë³„í™” ê°•í™” (ì„œë¹„ìŠ¤ í’ˆì§ˆ)',
                    'ì „í™˜ ë¹„ìš© êµ¬ì¶•',
                    'ë¸Œëœë“œ êµ¬ì¶•'
                ],
                'contingency': 'ê°€ê²© 10% ì¶”ê°€ ì¸í•˜ ê°€ëŠ¥'
            })
        
        # Risk 2: Churn ëª©í‘œ ë¯¸ë‹¬
        target_churn = opportunity['unit_economics'].get('churn', 0.05)
        if target_churn <= 0.05:
            risks.append({
                'risk_id': 'RISK_002',
                'category': 'execution',
                'risk': 'Churn Rate ëª©í‘œ ë¯¸ë‹¬ì„±',
                'probability': 'medium',
                'impact': 'high',
                'severity': 'high',
                'mitigation': [
                    'ì˜¨ë³´ë”© ê°•í™”',
                    'ê³ ê° ì„±ê³µ íŒ€',
                    'ì •ê¸° í”¼ë“œë°±'
                ],
                'contingency': 'Churn 10% ì´ˆê³¼ ì‹œ ê¸°ëŠ¥ ê°œì„ '
            })
        
        # Risk 3: Unit Economics
        ltv = opportunity['unit_economics'].get('ltv', 0)
        cac = opportunity['unit_economics'].get('cac', 1)
        ltv_cac = ltv / cac if cac > 0 else 0
        
        if ltv_cac < 3:
            risks.append({
                'risk_id': 'RISK_003',
                'category': 'financial',
                'risk': 'Unit Economics ì•…í™”',
                'probability': 'medium',
                'impact': 'critical',
                'severity': 'critical',
                'mitigation': [
                    'CAC ìµœì í™”',
                    'LTV ì¦ëŒ€ (Churn ê°œì„ )',
                    'ê°€ê²© ì¡°ì • ê²€í† '
                ],
                'contingency': 'Burn rate ê°ì†Œ'
            })
        
        # Critical Assumptions
        assumptions = [
            {
                'assumption_id': 'ASM_001',
                'assumption': f'Churn Rate {target_churn:.0%} ìœ ì§€',
                'basis': 'Validator ë²¤ì¹˜ë§ˆí¬',
                'test_method': 'ì²« 3ê°œì›” Beta ëª¨ë‹ˆí„°ë§',
                'success_criteria': f'Beta Churn < {target_churn * 1.4:.0%}'
            },
            {
                'assumption_id': 'ASM_002',
                'assumption': f'ê°€ê²© ìˆ˜ìš©ì„±',
                'basis': 'ê²½ìŸì‚¬ ëŒ€ë¹„ í• ì¸',
                'test_method': 'Beta ê°€ê²© í…ŒìŠ¤íŠ¸',
                'success_criteria': 'ì „í™˜ìœ¨ > 10%'
            }
        ]
        
        return {
            'key_risks': risks,
            'critical_assumptions': assumptions
        }
    
    def _generate_playbook_markdown(
        self,
        opportunity: Dict,
        gtm: Dict,
        roadmap: Dict,
        resources: Dict,
        milestones: Dict,
        risks: Dict,
        project_name: str
    ) -> str:
        """Markdown íŒŒì¼ ìƒì„±"""
        
        from datetime import datetime
        
        # ì €ì¥ ê²½ë¡œ
        project_dir = Path(f"projects/{project_name}/02_analysis/explorer")
        project_dir.mkdir(parents=True, exist_ok=True)
        
        md_path = project_dir / "strategy_playbook.md"
        
        # í˜„ì¬ ë‚ ì§œ
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Markdown ë‚´ìš© ìƒì„±
        content = f"""# Strategy Playbook: {opportunity.get('title')}

**ìƒì„±ì¼**: {today}
**Agent**: Explorer
**ë²„ì „**: 1.0

---

## Executive Summary

### ê¸°íšŒ ê°œìš”
- **ì œëª©**: {opportunity.get('title')}
- **ê°€ì¹˜ ì œì•ˆ**: {opportunity.get('value_proposition')}
- **íƒ€ê²Ÿ ê³ ê°**: {opportunity.get('target_customer')}
- **SAM**: {quantified.get('sam')}ì–µì›
- **ëª©í‘œ ì ìœ ìœ¨**: {quantified.get('target_share', 0.05):.1%}

### í•µì‹¬ Milestone
- **3ê°œì›”**: {milestones['month_3']['milestone']} - MRR {milestones['month_3']['metrics']['mrr']}
- **6ê°œì›”**: {milestones['month_6']['milestone']} - MRR {milestones['month_6']['metrics']['mrr']}
- **12ê°œì›”**: {milestones['month_12']['milestone']} - ARR {milestones['month_12']['metrics']['arr']}

---

## GTM Strategy

### Customer Acquisition
- **Target Segment**: {gtm['customer_acquisition']['target_segment']}
- **Segment Size**: {gtm['customer_acquisition']['segment_size']:,}ëª…/ë…„

**Acquisition Channels**:
"""
        
        for ch in gtm['customer_acquisition']['channels']:
            content += f"\n{ch['priority']}. **{ch['channel']}** ({ch['timeline']})\n"
            content += f"   - CAC: {ch['cac_estimate']/10000:.0f}ë§Œì›\n"
            content += f"   - Rationale: {ch['rationale']}\n"
        
        content += f"""

### Pricing
- **Model**: {gtm['pricing']['pricing_model']}
- **Price**: {gtm['pricing']['price_point']/10000:.0f}ë§Œì›/ì›”
- **Strategy**: {gtm['pricing']['pricing_strategy']}

---

## Product Roadmap

### MVP (3ê°œì›”)
"""
        
        for feat in roadmap['mvp']['features']:
            content += f"- **{feat['feature']}** (RICE: {feat['rice_score']})\n"
        
        content += f"""

### Phase 2 (6ê°œì›”)
"""
        
        for feat in roadmap['phase_2']['features']:
            content += f"- **{feat['feature']}** (RICE: {feat['rice_score']})\n"
        
        content += f"""

---

## Milestones

### Month 3: {milestones['month_3']['milestone']}
- **Metrics**: {milestones['month_3']['metrics']}
- **Activities**:
"""
        
        for act in milestones['month_3']['key_activities']:
            content += f"  - {act}\n"
        
        content += """

---

## Risk Register

"""
        
        for risk in risks['key_risks']:
            content += f"### {risk['risk_id']}: {risk['risk']}\n"
            content += f"- **Severity**: {risk['severity'].title()}\n"
            content += f"- **Mitigation**:\n"
            for mit in risk['mitigation']:
                content += f"  - {mit}\n"
            content += "\n"
        
        # íŒŒì¼ ì €ì¥
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"    âœ… Markdown ì €ì¥: {md_path}")
        
        return str(md_path)
    
    def _generate_playbook_excel(
        self,
        opportunity: Dict,
        gtm: Dict,
        roadmap: Dict,
        resources: Dict,
        milestones: Dict,
        risks: Dict,
        project_name: str
    ) -> str:
        """Excel íŒŒì¼ ìƒì„± (openpyxl)"""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.warning("    âš ï¸ openpyxl ì—†ìŒ, Excel ìƒì„± ìŠ¤í‚µ")
            return ""
        
        wb = Workbook()
        
        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1['A1'] = 'í•­ëª©'
        ws1['B1'] = 'ë‚´ìš©'
        ws1['A2'] = 'ê¸°íšŒ ì œëª©'
        ws1['B2'] = opportunity.get('title')
        ws1['A3'] = 'ê°€ì¹˜ ì œì•ˆ'
        ws1['B3'] = opportunity.get('value_proposition')
        ws1['A4'] = 'SAM'
        ws1['B4'] = f"{quantified.get('sam')}ì–µì›"
        
        # Sheet 2: GTM Strategy
        ws2 = wb.create_sheet("GTM Strategy")
        headers = ['ì˜ì—­', 'ì „ëµ', 'ì„¸ë¶€ ë‚´ìš©', 'ì˜ˆì‚°']
        for col, h in enumerate(headers, 1):
            ws2.cell(1, col, h)
        
        row = 2
        for ch in gtm['customer_acquisition']['channels']:
            ws2.cell(row, 1, 'ê³ ê° íšë“')
            ws2.cell(row, 2, ch['channel'])
            ws2.cell(row, 3, ch['rationale'])
            ws2.cell(row, 4, f"{ch['cac_estimate']/10000:.0f}ë§Œì›")
            row += 1
        
        # Sheet 3: Product Roadmap
        ws3 = wb.create_sheet("Product Roadmap")
        headers = ['Feature', 'RICE Score', 'Priority', 'Timeline']
        for col, h in enumerate(headers, 1):
            ws3.cell(1, col, h)
        
        row = 2
        for feat in roadmap['all_features']:
            ws3.cell(row, 1, feat['feature'])
            ws3.cell(row, 2, feat['rice_score'])
            ws3.cell(row, 3, feat['priority'])
            
            if feat['priority'] <= 3:
                timeline = 'MVP'
            elif feat['priority'] <= 7:
                timeline = 'Phase 2'
            else:
                timeline = 'Phase 3'
            ws3.cell(row, 4, timeline)
            row += 1
        
        # Sheet 4: Milestones
        ws4 = wb.create_sheet("Milestones")
        headers = ['Milestone', 'íƒ€ì´ë°', 'Metrics', 'Success Criteria']
        for col, h in enumerate(headers, 1):
            ws4.cell(1, col, h)
        
        for idx, (key, timing) in enumerate([('month_3', 'Month 3'), ('month_6', 'Month 6'), ('month_12', 'Month 12')], 2):
            ms = milestones[key]
            ws4.cell(idx, 1, ms['milestone'])
            ws4.cell(idx, 2, timing)
            ws4.cell(idx, 3, str(ms['metrics']))
            ws4.cell(idx, 4, '\n'.join(ms['success_criteria']))
        
        # Sheet 5: Risk Register
        ws5 = wb.create_sheet("Risk Register")
        headers = ['Risk ID', 'Risk', 'Severity', 'Mitigation']
        for col, h in enumerate(headers, 1):
            ws5.cell(1, col, h)
        
        for idx, risk in enumerate(risks['key_risks'], 2):
            ws5.cell(idx, 1, risk['risk_id'])
            ws5.cell(idx, 2, risk['risk'])
            ws5.cell(idx, 3, risk['severity'].title())
            ws5.cell(idx, 4, '\n'.join(risk['mitigation']))
        
        # Header ìŠ¤íƒ€ì¼
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        # ì €ì¥
        project_dir = Path(f"projects/{project_name}/02_analysis/explorer")
        project_dir.mkdir(parents=True, exist_ok=True)
        excel_path = project_dir / "strategy_playbook.xlsx"
        
        wb.save(excel_path)
        
        logger.info(f"    âœ… Excel ì €ì¥: {excel_path}")
        
        return str(excel_path)


# í¸ì˜ í•¨ìˆ˜
def create_explorer_agent() -> ExplorerRAG:
    """Explorer RAG ì—ì´ì „íŠ¸ ìƒì„± (Factory)"""
    return ExplorerRAG()


def create_explorer_agentic() -> ExplorerAgenticRAG:
    """Explorer Agentic RAG ìƒì„± (í–¥í›„ ììœ¨ ì‹¤í–‰)"""
    return ExplorerAgenticRAG()

