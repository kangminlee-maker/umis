"""
Cost Structure Sheet Builder
비용 구조 시트

Sheet 4: Cost_Structure
- COGS (Cost of Goods Sold)
- OPEX (Operating Expenses): S&M, R&D, G&A
- Total Costs
- Year 0 ~ Year 5
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class CostBuilder:
    """
    Cost Structure 시트 빌더
    
    기능:
      - COGS 계산 (Revenue × (1 - Gross Margin))
      - OPEX 계산 (S&M, R&D, G&A)
      - 총 비용 계산
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, years: int = 5) -> None:
        """
        Cost Structure 시트 생성
        
        Args:
            years: 예측 년수 (기본 5년)
        """
        
        ws = self.wb.create_sheet("Cost_Structure")
        
        # === 1. 제목 ===
        ws['A1'] = "Cost Structure"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f"Year 0 ~ Year {years} 비용 구조 (매출 % 기준)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:H2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Cost Item'] + [f'Year {y}' for y in range(years + 1)] + ['% of Rev']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. Revenue (참조) ===
        row += 1
        ws.cell(row=row, column=1).value = "Revenue"
        ws.cell(row=row, column=1).font = Font(size=10, italic=True, color="666666")
        
        revenue_row = row  # Revenue 행 번호 저장
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=Revenue_Y{year}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(italic=True, color="666666")
        
        # === 4. COGS (원가) ===
        row += 1
        ws.cell(row=row, column=1).value = "COGS"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        cogs_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col (B=66, C=67, ...)
            
            # COGS = Revenue × (1 - Gross Margin) - 수정: revenue_row 사용
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*(1-GrossMarginTarget)'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # % of Revenue
        ws.cell(row=row, column=years + 3).value = "=(1-GrossMarginTarget)"
        ws.cell(row=row, column=years + 3).number_format = '0.0%'
        
        # Named Range (Year별 COGS)
        for year in range(years + 1):
            col_letter = chr(65 + 2 + year)
            self.fe.define_named_range(f'COGS_Y{year}', 'Cost_Structure', f'{col_letter}{row}')
        
        # === 5. Gross Profit (매출총이익) ===
        row += 1
        ws.cell(row=row, column=1).value = "Gross Profit"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        gross_profit_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Gross Profit = Revenue - COGS
            ws.cell(row=row, column=col).value = f'={col_letter}{row-2}-{col_letter}{row-1}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # % of Revenue
        ws.cell(row=row, column=years + 3).value = "=GrossMarginTarget"
        ws.cell(row=row, column=years + 3).number_format = '0.0%'
        
        # === 6. OPEX (운영비) ===
        row += 2
        ws.cell(row=row, column=1).value = "Operating Expenses (OPEX)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:H{row}')
        
        opex_items = [
            {'name': 'S&M', 'percent_name': 'SMPercent'},
            {'name': 'R&D', 'percent_name': 'RDPercent'},
            {'name': 'G&A', 'percent_name': 'GAPercent'}
        ]
        
        opex_start_row = row + 1
        
        for opex in opex_items:
            row += 1
            ws.cell(row=row, column=1).value = opex['name']
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # Year 0-5
            for year in range(years + 1):
                col = 2 + year
                col_letter = chr(64 + col)  # 수정: 64 + col
                
                # OPEX = Revenue × OPEX % - 수정: revenue_row 사용
                ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*{opex["percent_name"]}'
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # % of Revenue
            ws.cell(row=row, column=years + 3).value = f'={opex["percent_name"]}'
            ws.cell(row=row, column=years + 3).number_format = '0.0%'
        
        opex_end_row = row
        
        # === 7. Total OPEX ===
        row += 1
        ws.cell(row=row, column=1).value = "Total OPEX"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Total OPEX = SUM(S&M, R&D, G&A)
            ws.cell(row=row, column=col).value = f'=SUM({col_letter}{opex_start_row}:{col_letter}{opex_end_row})'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            # Named Range
            self.fe.define_named_range(f'OPEX_Y{year}', 'Cost_Structure', f'{col_letter}{row}')
        
        # % of Revenue
        ws.cell(row=row, column=years + 3).value = "=SMPercent+RDPercent+GAPercent"
        ws.cell(row=row, column=years + 3).number_format = '0.0%'
        
        # === 8. Total Costs (COGS + OPEX) ===
        row += 1
        ws.cell(row=row, column=1).value = "Total Costs"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Total = COGS + OPEX
            ws.cell(row=row, column=col).value = f'={col_letter}{cogs_row}+{col_letter}{row-1}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            
            # Named Range
            self.fe.define_named_range(f'TotalCosts_Y{year}', 'Cost_Structure', f'{col_letter}{row}')
        
        # === 9. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• COGS = Revenue × (1 - Gross Margin) 자동 계산"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• OPEX는 매출 대비 % 기준 (Assumptions 시트에서 조정)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Total Costs = COGS + OPEX"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        print(f"   ✅ Cost Structure 시트 생성 완료")
        print(f"      - COGS, OPEX (S&M, R&D, G&A)")
        print(f"      - Named Range: COGS_Y0~Y{years}, OPEX_Y0~Y{years}, TotalCosts_Y0~Y{years}")


# 테스트는 별도 스크립트에서

