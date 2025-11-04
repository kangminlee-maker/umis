"""
Break-even Analysis Sheet Builder
손익분기 분석 시트

Sheet 9: BreakEven
- 손익분기 매출
- 손익분기 달성 시점
- 필요 고객 수 (Unit Economics 연계)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class BreakEvenBuilder:
    """
    Break-even Analysis 시트 빌더
    
    기능:
      - 손익분기 매출 계산
      - 달성 시점 예측
      - 필요 고객 수
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """Break-even Analysis 시트 생성"""
        
        ws = self.wb.create_sheet("BreakEven")
        
        # === 1. 제목 ===
        ws['A1'] = "Break-even Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "손익분기 매출 및 달성 시점 분석"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        
        # === 2. 고정비 vs 변동비 (간단화) ===
        row = 4
        ws.cell(row=row, column=1).value = "비용 구조 (간단화)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Total Fixed Costs (연간)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Fixed Costs = OPEX (R&D + G&A)
        ws.cell(row=row, column=2).value = "=BaseRevenue*(RDPercent+GAPercent)"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=4).value = "R&D + G&A (고정비 가정)"
        
        fixed_row = row
        
        row += 1
        ws.cell(row=row, column=1).value = "Variable Cost Ratio"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Variable = COGS + S&M
        ws.cell(row=row, column=2).value = "=(1-GrossMarginTarget)+SMPercent"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=4).value = "COGS + S&M (변동비 가정)"
        
        var_ratio_row = row
        
        # === 3. 손익분기 매출 ===
        row += 2
        ws.cell(row=row, column=1).value = "손익분기 분석"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Break-even Revenue"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        
        # BEP = Fixed Costs / (1 - Variable Ratio)
        ws.cell(row=row, column=2).value = f"=B{fixed_row}/(1-B{var_ratio_row})"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=12, bold=True)
        ws.cell(row=row, column=3).value = "원"
        
        bep_row = row
        
        # === 4. 달성 시점 ===
        row += 2
        ws.cell(row=row, column=1).value = "달성 시점 예측"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        # Year별 매출과 BEP 비교
        for year in range(6):
            row += 1
            ws.cell(row=row, column=1).value = f"Year {year}"
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # Revenue
            ws.cell(row=row, column=2).value = f"=Revenue_Y{year}"
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # vs BEP
            ws.cell(row=row, column=3).value = f"=B{row}-B${bep_row}"
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            # Status
            ws.cell(row=row, column=4).value = f'=IF(B{row}>=B${bep_row}, "✅ 손익분기 달성", "진행 중")'
            ws.cell(row=row, column=4).font = Font(size=9)
        
        # === 5. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• BEP = Fixed Costs / (1 - Variable Cost Ratio)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 매출이 BEP 이상이면 손익분기 달성 (이익 발생)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 간단화: R&D+G&A=고정비, COGS+S&M=변동비로 가정"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ Break-even 시트 생성 완료")
        print(f"      - 손익분기 매출 계산")
        print(f"      - Year별 달성 여부")


# 테스트는 별도 스크립트에서

