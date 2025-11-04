"""
Financial Projection Dashboard Builder
재무 예측 요약 대시보드

Sheet 1: Dashboard
- 핵심 재무 지표 요약
- Year 5 Big Numbers
- 성장 추이
- 권장사항
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class FPDashboardBuilder:
    """
    Financial Projection Dashboard 시트 빌더
    
    기능:
      - 핵심 지표 요약
      - Year 5 Big Numbers
      - 시나리오 비교
      - 다음 액션
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, market_name: str = "Target Market") -> None:
        """
        Dashboard 시트 생성
        
        Args:
            market_name: 시장/비즈니스 이름
        """
        
        ws = self.wb.create_sheet("Dashboard", 0)  # 첫 번째 시트
        
        # === 1. 대시보드 제목 ===
        ws['A1'] = "Financial Projection Dashboard"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 35
        
        ws['A2'] = market_name
        ws['A2'].font = Font(size=12, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25
        
        # === 2. Year 5 핵심 지표 ===
        row = 4
        ws.cell(row=row, column=1).value = "📊 Year 5 핵심 지표"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # Revenue Year 5
        row += 1
        ws.cell(row=row, column=1).value = "Revenue (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=Revenue_Y5"
        ws.cell(row=row, column=2).number_format = '₩#,##0'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        
        # Net Income Year 5
        row += 1
        ws.cell(row=row, column=1).value = "Net Income (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=NetIncome_Y5"
        ws.cell(row=row, column=2).number_format = '₩#,##0'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        
        # CAGR
        row += 1
        ws.cell(row=row, column=1).value = "CAGR (Year 0-5)"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=((Revenue_Y5/Revenue_Y0)^(1/5))-1"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        
        # === 3. 성장 추이 (Year 0 → Year 5) ===
        row += 2
        ws.cell(row=row, column=1).value = "📈 성장 추이"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        # 헤더
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        ws.cell(row=row, column=1).value = "Metric"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "Year 0"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "Year 5"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        
        ws.cell(row=row, column=4).value = "Growth"
        ws.cell(row=row, column=4).font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        
        # Revenue
        row += 1
        ws.cell(row=row, column=1).value = "Revenue"
        ws.cell(row=row, column=2).value = "=Revenue_Y0"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "=Revenue_Y5"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).value = "=C{}/B{}-1".format(row, row)
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        # Net Income
        row += 1
        ws.cell(row=row, column=1).value = "Net Income"
        ws.cell(row=row, column=2).value = "=NetIncome_Y0"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "=NetIncome_Y5"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).value = "=IFERROR(C{}/B{}-1, 0)".format(row, row)
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        # === 4. 시나리오 비교 (Year 5) ===
        row += 2
        ws.cell(row=row, column=1).value = "🎯 시나리오 비교 (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "시나리오 참고:"
        ws.cell(row=row, column=2).value = "FP_Scenarios 시트에서 Bear/Base/Bull 비교"
        ws.cell(row=row, column=2).font = Font(size=9, italic=True)
        ws.merge_cells(f'B{row}:E{row}')
        
        # === 5. 다음 액션 ===
        row += 2
        ws.cell(row=row, column=1).value = "📋 다음 액션"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        actions = [
            "1. Assumptions 시트에서 성장률, Margin 조정",
            "2. Revenue_Buildup에서 세그먼트별 성장률 세밀 조정",
            "3. PL_3Year / PL_5Year에서 손익 추이 확인",
            "4. CashFlow에서 현금 소진 시점 확인 (Ending Cash < 0?)",
            "5. FP_Scenarios에서 Bear Case 확인 (리스크 관리)",
            "6. BreakEven에서 손익분기 달성 시점 확인"
        ]
        
        for action in actions:
            row += 1
            ws.cell(row=row, column=1).value = action
            ws.cell(row=row, column=1).font = Font(size=9)
            ws.merge_cells(f'A{row}:E{row}')
        
        # === 6. 시트 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "📊 상세 분석 시트"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True, color="666666")
        
        sheets = [
            "• Assumptions: 핵심 가정 입력",
            "• Revenue_Buildup: 세그먼트별 매출",
            "• Cost_Structure: COGS + OPEX",
            "• PL_3Year / PL_5Year: 손익계산서",
            "• CashFlow: 현금흐름표",
            "• Key_Metrics: 성장률, Margin 추이",
            "• FP_Scenarios: Bear/Base/Bull 비교",
            "• BreakEven: 손익분기 분석",
            "• DCF_Valuation: 기업 가치 평가"
        ]
        
        for sheet in sheets:
            row += 1
            ws.cell(row=row, column=1).value = sheet
            ws.cell(row=row, column=1).font = Font(size=9, color="666666")
            ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ Dashboard 시트 생성 완료")
        print(f"      - Year 5 Big Numbers")
        print(f"      - 성장 추이")
        print(f"      - 다음 액션 가이드")


# 테스트는 별도 스크립트에서

