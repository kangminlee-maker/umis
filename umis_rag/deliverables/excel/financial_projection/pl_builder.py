"""
P&L (Profit & Loss) Forecast Sheet Builder
손익계산서 예측 시트

P&L 구조:
- Revenue
- COGS (Cost of Goods Sold)
- Gross Profit
- OPEX (S&M, R&D, G&A)
- EBITDA
- D&A (Depreciation & Amortization)
- EBIT
- Interest
- EBT (Earnings Before Tax)
- Tax
- Net Income
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class PLBuilder:
    """
    P&L Forecast 시트 빌더
    
    기능:
      - 손익계산서 자동 생성
      - Revenue → Net Income 계산
      - Margin % 자동 계산
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        sheet_name: str,
        years: int = 3,
        start_year: int = 0,
        define_named_ranges: bool = True
    ) -> None:
        """
        P&L Forecast 시트 생성
        
        Args:
            sheet_name: 시트 이름 ('PL_3Year' 또는 'PL_5Year')
            years: 표시할 년수 (3 또는 5)
            start_year: 시작 년도 (0부터)
            define_named_ranges: Named Range 정의 여부 (중복 방지)
        """
        
        ws = self.wb.create_sheet(sheet_name)
        
        # === 1. 제목 ===
        ws['A1'] = f"P&L Forecast ({years} Years)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A1:{chr(65 + years + 1)}1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "손익계산서 예측 (단위: 원)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells(f'A2:{chr(65 + years + 1)}2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        for i in range(years + 1):
            col_letter = chr(66 + i)  # B, C, D, ...
            ws.column_dimensions[col_letter].width = 18
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        # Item, Year 0, Year 1, ..., Year N
        headers = ['Item'] + [f'Year {start_year + y}' for y in range(years + 1)]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. Revenue ===
        row += 1
        ws.cell(row=row, column=1).value = "Revenue"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        revenue_row = row
        
        for year in range(years + 1):
            col = 2 + year
            actual_year = start_year + year
            ws.cell(row=row, column=col).value = f'=Revenue_Y{actual_year}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        # === 4. COGS ===
        row += 1
        ws.cell(row=row, column=1).value = "COGS (Cost of Goods Sold)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        cogs_row = row
        
        for year in range(years + 1):
            col = 2 + year
            actual_year = start_year + year
            ws.cell(row=row, column=col).value = f'=COGS_Y{actual_year}'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 5. Gross Profit ===
        row += 1
        ws.cell(row=row, column=1).value = "Gross Profit"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        gross_profit_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Gross Profit = Revenue - COGS
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}-{col_letter}{cogs_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # === 6. Gross Margin % ===
        row += 1
        ws.cell(row=row, column=1).value = "  Gross Margin %"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Margin % = Gross Profit / Revenue
            ws.cell(row=row, column=col).value = f'={col_letter}{gross_profit_row}/{col_letter}{revenue_row}'
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True)
        
        # === 7. Operating Expenses ===
        row += 1
        ws.cell(row=row, column=1).value = "Operating Expenses"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "  S&M"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        sm_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*SMPercent'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        row += 1
        ws.cell(row=row, column=1).value = "  R&D"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        rd_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*RDPercent'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        row += 1
        ws.cell(row=row, column=1).value = "  G&A"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        ga_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*GAPercent'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 8. Total OPEX ===
        row += 1
        ws.cell(row=row, column=1).value = "Total Operating Expenses"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        opex_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Total OPEX = S&M + R&D + G&A
            ws.cell(row=row, column=col).value = f'={col_letter}{sm_row}+{col_letter}{rd_row}+{col_letter}{ga_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # === 9. EBITDA ===
        row += 1
        ws.cell(row=row, column=1).value = "EBITDA"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        ebitda_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # EBITDA = Gross Profit - OPEX
            ws.cell(row=row, column=col).value = f'={col_letter}{gross_profit_row}-{col_letter}{opex_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            # Named Range (중복 방지)
            if define_named_ranges:
                actual_year = start_year + year
                self.fe.define_named_range(f'EBITDA_Y{actual_year}', sheet_name, f'{col_letter}{row}')
        
        # EBITDA Margin %
        row += 1
        ws.cell(row=row, column=1).value = "  EBITDA Margin %"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{ebitda_row}/{col_letter}{revenue_row}'
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True)
        
        # === 10. D&A (감가상각) ===
        row += 1
        ws.cell(row=row, column=1).value = "D&A (Depreciation & Amortization)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        da_row = row
        
        # 간단화: D&A = Revenue × 2% (가정)
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{revenue_row}*0.02'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 11. EBIT ===
        row += 1
        ws.cell(row=row, column=1).value = "EBIT (Operating Income)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        ebit_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # EBIT = EBITDA - D&A
            ws.cell(row=row, column=col).value = f'={col_letter}{ebitda_row}-{col_letter}{da_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # === 12. Interest (이자 비용) ===
        row += 1
        ws.cell(row=row, column=1).value = "Interest Expense"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        interest_row = row
        
        # 간단화: 이자 = 0 (무차입 가정)
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = 0
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 13. EBT (Earnings Before Tax) ===
        row += 1
        ws.cell(row=row, column=1).value = "EBT (Earnings Before Tax)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ebt_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # EBT = EBIT - Interest
            ws.cell(row=row, column=col).value = f'={col_letter}{ebit_row}-{col_letter}{interest_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 14. Tax (법인세) ===
        row += 1
        ws.cell(row=row, column=1).value = "Tax"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        tax_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Tax = EBT × Tax Rate (EBT > 0일 때만)
            ws.cell(row=row, column=col).value = f'=IF({col_letter}{ebt_row}>0, {col_letter}{ebt_row}*TaxRate, 0)'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 15. Net Income (순이익) ===
        row += 1
        ws.cell(row=row, column=1).value = "Net Income"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
        net_income_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            
            # Net Income = EBT - Tax
            ws.cell(row=row, column=col).value = f'={col_letter}{ebt_row}-{col_letter}{tax_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(size=11, bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            
            # Named Range (중복 방지)
            if define_named_ranges:
                actual_year = start_year + year
                self.fe.define_named_range(f'NetIncome_Y{actual_year}', sheet_name, f'{col_letter}{row}')
        
        # Net Margin %
        row += 1
        ws.cell(row=row, column=1).value = "  Net Margin %"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            ws.cell(row=row, column=col).value = f'={col_letter}{net_income_row}/{col_letter}{revenue_row}'
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True)
        
        # === 16. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Gross Profit = Revenue - COGS (매출총이익)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• EBITDA = Gross Profit - OPEX (영업이익, 감가상각 전)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Net Income = 세후 순이익 (최종 수익)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        print(f"   ✅ {sheet_name} 시트 생성 완료")
        print(f"      - {years+1}개년 P&L (Year {start_year} ~ Year {start_year + years})")
        print(f"      - Revenue → Net Income 계산")


# 테스트는 별도 스크립트에서

