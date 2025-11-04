"""
Financial Projection Model Generator (완성 버전)
재무 예측 모델 Excel 자동 생성

버전: Batch 6 완성 (10개 시트)
"""

from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

from openpyxl import Workbook

from ..formula_engine import FormulaEngine
from .fp_assumptions_builder import FPAssumptionsBuilder
from .revenue_builder import RevenueBuilder
from .cost_builder import CostBuilder
from .pl_builder import PLBuilder
from .cashflow_builder import CashFlowBuilder
from .metrics_builder import MetricsBuilder
from .fp_scenarios_builder import FPScenariosBuilder
from .breakeven_builder import BreakEvenBuilder
from .dcf_builder import DCFBuilder
from .fp_dashboard_builder import FPDashboardBuilder


class FinancialProjectionGenerator:
    """
    Financial Projection Excel 자동 생성기 (완성)
    
    생성 시트 (10개):
      1. Dashboard (요약)
      2. Assumptions
      3. Revenue_Buildup
      4. Cost_Structure
      5. PL_3Year
      6. PL_5Year
      7. CashFlow
      8. Key_Metrics
      9. FP_Scenarios
      10. BreakEven
      11. DCF_Valuation (선택)
    """
    
    def __init__(self):
        """초기화"""
        self.formula_engine: Optional[FormulaEngine] = None
    
    def generate(
        self,
        market_name: str,
        assumptions_data: Dict,
        segments: List[Dict],
        years: int = 5,
        output_dir: Path = Path('.')
    ) -> Path:
        """
        Financial Projection Workbook 생성 (Batch 4)
        
        Args:
            market_name: 시장/비즈니스 이름
            assumptions_data: 가정 데이터
                {
                    'base_revenue_y0': 1250_0000_0000,
                    'growth_rate_yoy': 0.28,
                    'gross_margin': 0.70,
                    'ebitda_margin': 0.15,
                    'net_margin': 0.10,
                    'sm_percent': 0.30,
                    'rd_percent': 0.15,
                    'ga_percent': 0.10,
                    'tax_rate': 0.25,
                    'discount_rate': 0.12
                }
            segments: 세그먼트 목록
                [
                    {'name': 'B2C', 'y0_revenue': 800_0000_0000, 'growth': 0.15},
                    {'name': 'B2B', 'y0_revenue': 300_0000_0000, 'growth': 0.35},
                    {'name': 'B2G', 'y0_revenue': 150_0000_0000, 'growth': 0.45}
                ]
            years: 예측 년수 (기본 5년)
            output_dir: 출력 디렉토리
        
        Returns:
            생성된 Excel 파일 경로
        """
        
        print(f"🚀 Financial Projection Model 생성 시작")
        print(f"   시장: {market_name}")
        print(f"   버전: 완성 (11개 시트)")
        print(f"   예측 기간: {years}년")
        
        # 1. 워크북 초기화
        wb = Workbook()
        self.formula_engine = FormulaEngine(wb)
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 2. Sheet 2: Assumptions
        print(f"   2/11 Assumptions...")
        assumptions_builder = FPAssumptionsBuilder(wb, self.formula_engine)
        assumptions_builder.create_sheet(assumptions_data)
        
        # 3. Sheet 3: Revenue Build-up
        print(f"   3/11 Revenue Build-up...")
        revenue_builder = RevenueBuilder(wb, self.formula_engine)
        revenue_builder.create_sheet(segments, years)
        
        # 4. Sheet 4: Cost Structure
        print(f"   4/11 Cost Structure...")
        cost_builder = CostBuilder(wb, self.formula_engine)
        cost_builder.create_sheet(years)
        
        # 5. Sheet 5: P&L 3 Year
        print(f"   5/11 P&L 3 Year...")
        pl_3year_builder = PLBuilder(wb, self.formula_engine)
        pl_3year_builder.create_sheet('PL_3Year', years=3, start_year=0, define_named_ranges=False)
        
        # 6. Sheet 6: P&L 5 Year (Named Range 정의)
        print(f"   6/11 P&L 5 Year...")
        pl_5year_builder = PLBuilder(wb, self.formula_engine)
        pl_5year_builder.create_sheet('PL_5Year', years=5, start_year=0, define_named_ranges=True)
        
        # 7. Sheet 7: Cash Flow
        print(f"   7/11 Cash Flow...")
        cashflow_builder = CashFlowBuilder(wb, self.formula_engine)
        cashflow_builder.create_sheet(years)
        
        # 8. Sheet 8: Key Metrics
        print(f"   8/11 Key Metrics...")
        metrics_builder = MetricsBuilder(wb, self.formula_engine)
        metrics_builder.create_sheet(years, 'PL_5Year')
        
        # 9. Sheet 9: Scenarios (Batch 6)
        print(f"   9/11 FP Scenarios...")
        scenarios_builder = FPScenariosBuilder(wb, self.formula_engine)
        scenarios_builder.create_sheet()
        
        # 10. Sheet 10: Break-even (Batch 6)
        print(f"   10/11 Break-even...")
        breakeven_builder = BreakEvenBuilder(wb, self.formula_engine)
        breakeven_builder.create_sheet()
        
        # 11. Sheet 11: DCF Valuation (Batch 6)
        print(f"   11/11 DCF Valuation...")
        dcf_builder = DCFBuilder(wb, self.formula_engine)
        dcf_builder.create_sheet(years)
        
        # 12. Sheet 1: Dashboard (Batch 6, 맨 앞으로)
        print(f"   1/11 Dashboard...")
        dashboard_builder = FPDashboardBuilder(wb, self.formula_engine)
        dashboard_builder.create_sheet(market_name)
        
        # 13. 강제 재계산 설정
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        # 14. 저장
        filename = f"financial_projection_{market_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = output_dir / filename
        
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        print(f"\n✅ Excel 생성 완료: {filepath}")
        print(f"📊 시트: {len(wb.sheetnames)}개")
        print(f"📋 Named Range: {len(self.formula_engine.named_ranges)}개")
        print(f"🎉 Financial Projection Model 완성!")
        
        return filepath


# 테스트는 별도 스크립트에서
# python scripts/test_financial_projection_batch4.py

