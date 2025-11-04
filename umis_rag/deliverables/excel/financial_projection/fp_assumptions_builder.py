"""
Financial Projection Assumptions Sheet Builder
재무 예측 가정 시트

Sheet 2: Assumptions
- 성장률 (YoY, CAGR)
- Gross Margin, EBITDA Margin, Net Margin
- OPEX 비율 (S&M, R&D, G&A)
- Tax Rate, Discount Rate (DCF용)
- 세그먼트별 성장률
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class FPAssumptionsBuilder:
    """
    Financial Projection Assumptions 시트 빌더
    
    기능:
      - 재무 예측 핵심 가정
      - 성장률, Margin, 비용율
      - Named Range 정의
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, assumptions_data: dict = None) -> None:
        """
        Assumptions 시트 생성
        
        Args:
            assumptions_data: 가정 데이터
                {
                    'base_revenue_y0': 1250_0000_0000,  # 현재 매출 (125억)
                    'growth_rate_yoy': 0.28,  # YoY 성장률 28%
                    'gross_margin': 0.70,  # Gross Margin 70%
                    'ebitda_margin': 0.15,  # EBITDA Margin 15%
                    'net_margin': 0.10,  # Net Margin 10%
                    'sm_percent': 0.30,  # S&M 비율 30%
                    'rd_percent': 0.15,  # R&D 비율 15%
                    'ga_percent': 0.10,  # G&A 비율 10%
                    'tax_rate': 0.25,  # 법인세율 25%
                    'discount_rate': 0.12  # 할인율 12% (DCF용)
                }
        """
        
        # 기본값
        if assumptions_data is None:
            assumptions_data = {
                'base_revenue_y0': 1000_0000_0000,  # 100억
                'growth_rate_yoy': 0.25,  # 25%
                'gross_margin': 0.60,
                'ebitda_margin': 0.12,
                'net_margin': 0.08,
                'sm_percent': 0.25,
                'rd_percent': 0.12,
                'ga_percent': 0.08,
                'tax_rate': 0.25,
                'discount_rate': 0.10
            }
        
        ws = self.wb.create_sheet("Assumptions")
        
        # === 1. 제목 ===
        ws['A1'] = "Financial Projection Assumptions"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "재무 예측의 핵심 가정 (노란색 셀만 수정)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        
        # === 2. 기준 매출 ===
        row = 4
        ws.cell(row=row, column=1).value = "1. 기준 매출 (Year 0)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        input_fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1).value = "Base Revenue (Year 0)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = assumptions_data['base_revenue_y0']
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "현재 연간 매출 (기준점)"
        
        # Named Range
        self.fe.define_named_range('BaseRevenue', 'Assumptions', f'B{row}')
        
        # === 3. 성장률 ===
        row += 2
        ws.cell(row=row, column=1).value = "2. 성장률 (Growth Rates)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "YoY Growth Rate (전체)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = assumptions_data['growth_rate_yoy']
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "%"
        ws.cell(row=row, column=4).value = "Year-over-Year 평균 성장률"
        
        # Named Range
        self.fe.define_named_range('GrowthRateYoY', 'Assumptions', f'B{row}')
        
        # === 4. Margin ===
        row += 2
        ws.cell(row=row, column=1).value = "3. Margin (수익률)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        margins = [
            {
                'name': 'GrossMarginTarget',
                'label': 'Gross Margin (목표)',
                'value': assumptions_data['gross_margin'],
                'description': '(Revenue - COGS) / Revenue'
            },
            {
                'name': 'EBITDAMargin',
                'label': 'EBITDA Margin (목표)',
                'value': assumptions_data['ebitda_margin'],
                'description': 'EBITDA / Revenue'
            },
            {
                'name': 'NetMargin',
                'label': 'Net Margin (목표)',
                'value': assumptions_data['net_margin'],
                'description': 'Net Income / Revenue'
            }
        ]
        
        for margin in margins:
            row += 1
            ws.cell(row=row, column=1).value = margin['label']
            ws.cell(row=row, column=1).font = Font(size=10)
            ws.cell(row=row, column=2).value = margin['value']
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=3).value = "%"
            ws.cell(row=row, column=4).value = margin['description']
            
            # Named Range
            self.fe.define_named_range(margin['name'], 'Assumptions', f'B{row}')
        
        # === 5. OPEX 비율 ===
        row += 2
        ws.cell(row=row, column=1).value = "4. OPEX 비율 (% of Revenue)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        opex_items = [
            {
                'name': 'SMPercent',
                'label': 'S&M (Sales & Marketing)',
                'value': assumptions_data['sm_percent'],
                'description': '영업 및 마케팅 비용 / 매출'
            },
            {
                'name': 'RDPercent',
                'label': 'R&D (Research & Development)',
                'value': assumptions_data['rd_percent'],
                'description': '연구개발 비용 / 매출'
            },
            {
                'name': 'GAPercent',
                'label': 'G&A (General & Administrative)',
                'value': assumptions_data['ga_percent'],
                'description': '일반관리 비용 / 매출'
            }
        ]
        
        for opex in opex_items:
            row += 1
            ws.cell(row=row, column=1).value = opex['label']
            ws.cell(row=row, column=1).font = Font(size=10)
            ws.cell(row=row, column=2).value = opex['value']
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=3).value = "%"
            ws.cell(row=row, column=4).value = opex['description']
            
            # Named Range
            self.fe.define_named_range(opex['name'], 'Assumptions', f'B{row}')
        
        # === 6. 기타 가정 ===
        row += 2
        ws.cell(row=row, column=1).value = "5. 기타 가정"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        other_items = [
            {
                'name': 'TaxRate',
                'label': 'Tax Rate (법인세율)',
                'value': assumptions_data['tax_rate'],
                'description': '법인세 (일반 25%)'
            },
            {
                'name': 'DiscountRate',
                'label': 'Discount Rate (할인율)',
                'value': assumptions_data['discount_rate'],
                'description': 'DCF 현가 계산용 (WACC)'
            }
        ]
        
        for item in other_items:
            row += 1
            ws.cell(row=row, column=1).value = item['label']
            ws.cell(row=row, column=1).font = Font(size=10)
            ws.cell(row=row, column=2).value = item['value']
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).number_format = '0.0%'
            ws.cell(row=row, column=3).value = "%"
            ws.cell(row=row, column=4).value = item['description']
            
            # Named Range
            self.fe.define_named_range(item['name'], 'Assumptions', f'B{row}')
        
        # === 7. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 입력 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        guides = [
            "• YoY Growth Rate: 전년 대비 성장률 (일정하다고 가정)",
            "• Gross Margin: 원가를 제외한 수익률",
            "• EBITDA Margin: 영업이익률 (감가상각 전)",
            "• Net Margin: 세후 순이익률",
            "• OPEX %: 각 비용 항목의 매출 대비 비율",
            "• Discount Rate: 일반적으로 WACC 또는 요구 수익률 (10-15%)"
        ]
        
        for guide in guides:
            row += 1
            ws.cell(row=row, column=1).value = guide
            ws.cell(row=row, column=1).font = Font(size=9)
            ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ Assumptions 시트 생성 완료")
        print(f"      - 10개 Named Range 정의")


# 테스트는 별도 스크립트에서

