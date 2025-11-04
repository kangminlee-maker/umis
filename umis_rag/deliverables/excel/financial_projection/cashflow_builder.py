"""
Cash Flow Forecast Sheet Builder
현금흐름표 예측 시트

Cash Flow 구조:
- Operating Cash Flow
- Investment Cash Flow (CAPEX)
- Financing Cash Flow
- Net Cash Flow
- Ending Cash Balance
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class CashFlowBuilder:
    """
    Cash Flow Forecast 시트 빌더
    
    기능:
      - 운영 현금흐름 (EBITDA 기반)
      - 투자 현금흐름 (CAPEX)
      - 재무 현금흐름
      - 현금 잔액 추적
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, years: int = 5) -> None:
        """
        Cash Flow Forecast 시트 생성
        
        Args:
            years: 예측 년수 (기본 5년)
        """
        
        ws = self.wb.create_sheet("CashFlow")
        
        # === 1. 제목 ===
        ws['A1'] = "Cash Flow Forecast"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A1:{chr(65 + years + 1)}1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "현금흐름표 예측 (단위: 원)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells(f'A2:{chr(65 + years + 1)}2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 35
        for i in range(years + 1):
            col_letter = chr(66 + i)
            ws.column_dimensions[col_letter].width = 18
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Item'] + [f'Year {y}' for y in range(years + 1)]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. Operating Cash Flow ===
        row += 1
        ws.cell(row=row, column=1).value = "Operating Cash Flow"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # EBITDA (간단화: Operating CF = EBITDA)
        row += 1
        ws.cell(row=row, column=1).value = "  EBITDA"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        ebitda_row = row
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=EBITDA_Y{year}'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # 간단화: Working Capital 변화 생략
        
        # Total Operating CF
        row += 1
        ws.cell(row=row, column=1).value = "Total Operating CF"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        operating_cf_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            ws.cell(row=row, column=col).value = f'={col_letter}{ebitda_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # === 4. Investment Cash Flow ===
        row += 2
        ws.cell(row=row, column=1).value = "Investment Cash Flow"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # CAPEX (자본적 지출)
        row += 1
        ws.cell(row=row, column=1).value = "  CAPEX"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        capex_row = row
        
        # 간단화: CAPEX = Revenue × 5%
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=-Revenue_Y{year}*0.05'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # Total Investment CF
        row += 1
        ws.cell(row=row, column=1).value = "Total Investment CF"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        investment_cf_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            ws.cell(row=row, column=col).value = f'={col_letter}{capex_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # === 5. Financing Cash Flow ===
        row += 2
        ws.cell(row=row, column=1).value = "Financing Cash Flow"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # 간단화: 차입/상환 없음
        row += 1
        ws.cell(row=row, column=1).value = "  Debt/Equity Issuance"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        financing_row = row
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = 0
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 6. Net Cash Flow ===
        row += 2
        ws.cell(row=row, column=1).value = "Net Cash Flow"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        
        net_cf_row = row
        
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            
            # Net CF = Operating CF + Investment CF + Financing CF
            ws.cell(row=row, column=col).value = (
                f'={col_letter}{operating_cf_row}+{col_letter}{investment_cf_row}+{col_letter}{financing_row}'
            )
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        
        # === 7. Ending Cash Balance ===
        row += 1
        ws.cell(row=row, column=1).value = "Ending Cash Balance"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        # 초기 현금 (Year 0)
        col = 2
        ws.cell(row=row, column=col).value = 100_0000_0000  # 10억 초기 현금
        ws.cell(row=row, column=col).number_format = '#,##0'
        ws.cell(row=row, column=col).fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        
        # Year 1-5
        for year in range(1, years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            prev_col_letter = chr(65 + col - 1)
            
            # Ending Cash = Previous Cash + Net CF
            ws.cell(row=row, column=col).value = f'={prev_col_letter}{row}+{col_letter}{net_cf_row}'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # === 8. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Operating CF = EBITDA (간단화, Working Capital 변화 생략)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Investment CF = CAPEX (매출의 5% 가정)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Ending Cash < 0이면 추가 자금 조달 필요"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        print(f"   ✅ Cash Flow 시트 생성 완료")
        print(f"      - Operating CF, Investment CF, Financing CF")
        print(f"      - Net Cash Flow, Ending Cash Balance")


# 테스트는 별도 스크립트에서

