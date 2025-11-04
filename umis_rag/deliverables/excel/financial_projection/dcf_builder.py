"""
DCF Valuation Sheet Builder  
DCF 기업 가치 평가 시트 (간단화)

Sheet 10: DCF_Valuation
- Free Cash Flow 현가 계산
- Terminal Value
- Enterprise Value
- Equity Value (간단화)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class DCFBuilder:
    """
    DCF Valuation 시트 빌더 (간단화 버전)
    
    기능:
      - Free Cash Flow 현가 계산
      - Terminal Value (영구 성장률 방식)
      - Enterprise Value
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, years: int = 5) -> None:
        """
        DCF Valuation 시트 생성 (간단화)
        
        Args:
            years: 예측 년수
        """
        
        ws = self.wb.create_sheet("DCF_Valuation")
        
        # === 1. 제목 ===
        ws['A1'] = "DCF Valuation (간단화)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "Discounted Cash Flow 기업 가치 평가"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        
        # === 2. 가정 ===
        row = 4
        ws.cell(row=row, column=1).value = "DCF 가정"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "Discount Rate (WACC)"
        ws.cell(row=row, column=2).value = "=DiscountRate"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=4).value = "Assumptions 시트에서 참조"
        
        discount_row = row
        
        row += 1
        ws.cell(row=row, column=1).value = "Terminal Growth Rate"
        ws.cell(row=row, column=2).value = 0.03  # 3% 영구 성장
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        ws.cell(row=row, column=4).value = "영구 성장률 (보수적 3%)"
        
        terminal_growth_row = row
        
        # === 3. FCF 현가 계산 ===
        row += 2
        ws.cell(row=row, column=1).value = "Free Cash Flow 현가"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        # 간단화: FCF ≈ EBITDA (Working Capital, CAPEX 무시)
        fcf_pv_ranges = []  # Named Ranges for PV of each year
        fcf_rows_start = row + 1
        
        for year in range(1, years + 1):
            row += 1
            ws.cell(row=row, column=1).value = f"Year {year} FCF"
            
            # FCF = EBITDA
            ws.cell(row=row, column=2).value = f"=EBITDA_Y{year}"
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # Present Value = FCF / (1 + Discount)^Year
            ws.cell(row=row, column=3).value = f"=B{row}/((1+B${discount_row})^{year})"
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            # Named Range for each year's PV
            nr_name = f'DCF_PV_Y{year}'
            self.fe.define_named_range(nr_name, 'DCF_Valuation', f'C{row}')
            fcf_pv_ranges.append(nr_name)
        
        # === 4. PV 합계 ===
        row += 1
        pv_sum_row = row
        
        ws.cell(row=row, column=1).value = "PV of FCF (Year 1-5)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # SUM using Named Ranges
        ws.cell(row=row, column=3).value = f"=SUM({','.join(fcf_pv_ranges)})"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # === 5. Terminal Value ===
        row += 2
        ws.cell(row=row, column=1).value = "Terminal Value"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = f"Year {years} FCF"
        ws.cell(row=row, column=2).value = f"=EBITDA_Y{years}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        year5_fcf_row = row
        
        row += 1
        ws.cell(row=row, column=1).value = "Terminal Value"
        
        # TV = FCF × (1 + g) / (WACC - g)
        ws.cell(row=row, column=2).value = (
            f"=B{year5_fcf_row}*(1+B${terminal_growth_row})/(B${discount_row}-B${terminal_growth_row})"
        )
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        tv_row = row
        
        row += 1
        ws.cell(row=row, column=1).value = "PV of Terminal Value"
        
        # PV of TV = TV / (1 + Discount)^5
        ws.cell(row=row, column=2).value = f"=B{tv_row}/((1+B${discount_row})^{years})"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        pv_tv_row = row
        
        # === 6. Enterprise Value ===
        row += 2
        ws.cell(row=row, column=1).value = "Enterprise Value"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
        # EV = PV of FCF + PV of TV
        ws.cell(row=row, column=2).value = f"=C{pv_sum_row}+B{pv_tv_row}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.cell(row=row, column=3).value = "원"
        
        # === 7. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• 간단화: FCF ≈ EBITDA (CAPEX, Working Capital 무시)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Terminal Value = 영구 현금흐름의 현재 가치"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Enterprise Value = 기업의 총 가치 (부채 제외 전)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ DCF Valuation 시트 생성 완료")
        print(f"      - Free Cash Flow 현가")
        print(f"      - Terminal Value")
        print(f"      - Enterprise Value")


# 테스트는 별도 스크립트에서

