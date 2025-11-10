"""
Revenue Build-up Sheet Builder
세그먼트별 매출 구축 시트

Sheet 3: Revenue_Buildup
- Year 0 ~ Year 5 매출
- 세그먼트별 (B2C, B2B, B2G, Global 등)
- 성장률 적용
- 총 매출 계산
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles
from ..builder_contract import BuilderContract, ValidationStatus


class RevenueBuilder:
    """
    Revenue Build-up 시트 빌더
    
    기능:
      - 세그먼트별 매출 예측
      - 연도별 성장률 적용
      - 총 매출 자동 계산
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        segments: list = None,
        years: int = 5
    ) -> BuilderContract:
        """
        Revenue Build-up 시트 생성
        
        Args:
            segments: 세그먼트 목록
                [
                    {'name': 'B2C', 'y0_revenue': 80_0000_0000, 'growth': 0.15},
                    {'name': 'B2B', 'y0_revenue': 30_0000_0000, 'growth': 0.35},
                    {'name': 'B2G', 'y0_revenue': 15_0000_0000, 'growth': 0.45},
                ]
            years: 예측 년수 (기본 5년)
        
        Returns:
            BuilderContract: 생성한 Named Range 목록 포함
        """
        
        # Contract 생성
        contract = BuilderContract(sheet_name='Revenue_Buildup')
        
        # FormulaEngine에 Contract 연결 (Named Range 자동 등록)
        self.fe.set_contract(contract)
        
        # 기본 세그먼트
        if segments is None:
            segments = [
                {'name': 'B2C (개인)', 'y0_revenue': 700_0000_0000, 'growth': 0.10},
                {'name': 'B2B (기업)', 'y0_revenue': 200_0000_0000, 'growth': 0.30},
                {'name': 'B2G (정부)', 'y0_revenue': 100_0000_0000, 'growth': 0.40},
            ]
        
        ws = self.wb.create_sheet("Revenue_Buildup")
        
        # === 1. 제목 ===
        ws['A1'] = "Revenue Build-up (세그먼트별)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f"Year 0 ~ Year {years} 매출 예측 (세그먼트별 성장률)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:H2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        # 헤더 (Segment, Year 0 ~ Year 5, Growth)
        year_headers = ['Segment'] + [f'Year {y}' for y in range(years + 1)] + ['Growth %']
        
        for col_idx, header in enumerate(year_headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 세그먼트별 매출 ===
        input_fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        
        segment_rows = []
        segment_year0_ranges = []  # 각 세그먼트 Year 0 Named Range
        
        for idx, seg in enumerate(segments, start=1):
            row += 1
            segment_rows.append(row)
            
            # A: Segment 이름
            ws.cell(row=row, column=1).value = seg['name']
            ws.cell(row=row, column=1).font = Font(size=10, bold=True)
            
            # B: Year 0 (입력)
            ws.cell(row=row, column=2).value = seg['y0_revenue']
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # C-G: Year 1-5 (성장률 적용)
            for year in range(1, years + 1):
                col = 2 + year  # C=3, D=4, ...
                prev_col_letter = chr(64 + col - 1)  # B, C, D, ...
                
                # 세그먼트별 성장률 사용
                # H 컬럼 (col 8) = years(5) + 3
                growth_cell = f'${chr(64 + years + 3)}${row}'  # 수정: 64 + years + 3 = H
                ws.cell(row=row, column=col).value = f'={prev_col_letter}{row}*(1+{growth_cell})'
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # H: Growth % (입력) - col 8
            ws.cell(row=row, column=years + 3).value = seg['growth']  # 수정: years + 3
            ws.cell(row=row, column=years + 3).fill = input_fill
            ws.cell(row=row, column=years + 3).number_format = '0.0%'
            
            # 각 세그먼트 Year 0에 Named Range 정의
            seg_range_name = f'Rev_Segment{idx}_Y0'
            self.fe.define_named_range(seg_range_name, 'Revenue_Buildup', f'B{row}')
            segment_year0_ranges.append(seg_range_name)
        
        # === 4. 총 매출 (Total Revenue) ===
        row += 1
        total_revenue_row = row
        ws.cell(row=row, column=1).value = "Total Revenue"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        
        # 각 세그먼트의 Year별 Named Range 생성
        segment_year_ranges = {}  # {year: [range_names]}
        
        # Year 0은 이미 생성됨 (segment_year0_ranges 사용)
        segment_year_ranges[0] = segment_year0_ranges
        
        # Year 1-5 Named Range 생성
        for year in range(1, years + 1):
            segment_year_ranges[year] = []
            for idx, seg_row in enumerate(segment_rows, start=1):
                col = 2 + year
                col_letter = chr(64 + col)
                nr_name = f'Rev_Segment{idx}_Y{year}'
                self.fe.define_named_range(nr_name, 'Revenue_Buildup', f'{col_letter}{seg_row}')
                segment_year_ranges[year].append(nr_name)
        
        # Year 0 ~ Year 5 합계 (Named Range 기반)
        for year in range(years + 1):
            col = 2 + year
            col_letter = chr(64 + col)
            
            # 모든 년도 Named Range 기반 SUM
            ranges_str = ','.join(segment_year_ranges[year])
            sum_formula = f"=SUM({ranges_str})"
            ws.cell(row=row, column=col).value = sum_formula
            
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(size=11, bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            
            # Named Range (Year별 총 매출)
            self.fe.define_named_range(f'Revenue_Y{year}', 'Revenue_Buildup', f'{col_letter}{row}')
        
        # === 5. YoY 성장률 (계산) ===
        row += 1
        ws.cell(row=row, column=1).value = "YoY Growth %"
        ws.cell(row=row, column=1).font = Font(size=10, italic=True)
        
        # Year 1-5 성장률
        for year in range(1, years + 1):
            col = 2 + year
            col_letter = chr(64 + col)  # 수정: 64 + col
            prev_col_letter = chr(64 + col - 1)  # 수정: 64 + col - 1
            
            ws.cell(row=row, column=col).value = f'=({col_letter}{row-1}-{prev_col_letter}{row-1})/{prev_col_letter}{row-1}'
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True)
        
        # === 6. 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• 세그먼트별 성장률을 다르게 설정 가능 (마지막 컬럼 수정)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 총 매출은 자동 계산 (세그먼트 합계)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• YoY %는 전년 대비 성장률 (자동 계산)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:H{row}')
        
        # 메타데이터 추가
        contract.add_metadata('num_segments', len(segments))
        contract.add_metadata('years', years)
        contract.add_metadata('total_revenue_row', total_revenue_row)
        
        # === Inline Validation (v7.2.0) ===
        self._validate_revenue_sheet(contract, segments, years)
        
        print(f"   ✅ Revenue Build-up 시트 생성 완료")
        print(f"      - {len(segments)}개 세그먼트")
        print(f"      - {years+1}개년 매출 (Year 0 ~ Year {years})")
        print(f"      - Named Range: Revenue_Y0 ~ Revenue_Y{years}")
        print(f"      - BuilderContract: {len(contract.list_named_ranges())} named ranges")
        
        # Validation 결과 출력
        if contract.validation_results:
            print(f"      - Validations: {len(contract.validation_results)} checks")
            if contract.has_failures():
                print(f"        ❌ {sum(1 for r in contract.validation_results if r.status == ValidationStatus.FAILED)} failed")
            if contract.has_warnings():
                print(f"        ⚠️  {sum(1 for r in contract.validation_results if r.status == ValidationStatus.WARNING)} warnings")
        
        return contract
    
    def _validate_revenue_sheet(
        self,
        contract: BuilderContract,
        segments: list,
        years: int
    ) -> None:
        """
        Revenue 시트 Inline Validation
        
        Args:
            contract: BuilderContract
            segments: 세그먼트 목록
            years: 년수
        """
        
        # 1. 세그먼트 수 검증
        if len(segments) >= 1:
            contract.add_validation(
                'segment_count',
                ValidationStatus.PASSED,
                f'{len(segments)} segments provided'
            )
        else:
            contract.add_validation(
                'segment_count',
                ValidationStatus.FAILED,
                'No segments provided'
            )
        
        # 2. Years 검증
        if years >= 1:
            contract.add_validation(
                'years_count',
                ValidationStatus.PASSED,
                f'{years} years projection'
            )
        else:
            contract.add_validation(
                'years_count',
                ValidationStatus.FAILED,
                'Years must be >= 1'
            )
        
        # 3. Named Range 개수 검증
        expected_ranges = len(segments) * (years + 1) + (years + 1)  # 세그먼트별 + Total
        actual_ranges = len(contract.list_named_ranges())
        
        if actual_ranges == expected_ranges:
            contract.add_validation(
                'named_range_count',
                ValidationStatus.PASSED,
                f'{actual_ranges} named ranges (expected: {expected_ranges})'
            )
        else:
            contract.add_validation(
                'named_range_count',
                ValidationStatus.WARNING,
                f'{actual_ranges} named ranges (expected: {expected_ranges})',
                {'expected': expected_ranges, 'actual': actual_ranges}
            )
        
        # 4. Revenue_Y0 ~ Revenue_Y{years} 존재 검증
        missing_ranges = []
        for year in range(years + 1):
            range_name = f'Revenue_Y{year}'
            if not contract.has_named_range(range_name):
                missing_ranges.append(range_name)
        
        if not missing_ranges:
            contract.add_validation(
                'revenue_year_ranges',
                ValidationStatus.PASSED,
                f'All Revenue_Y0~Y{years} defined'
            )
        else:
            contract.add_validation(
                'revenue_year_ranges',
                ValidationStatus.FAILED,
                f'Missing ranges: {", ".join(missing_ranges)}',
                {'missing': missing_ranges}
            )


# 테스트는 별도 스크립트에서

