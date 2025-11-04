"""
Financial Projection Scenarios Sheet Builder
재무 시나리오 분석 시트

Sheet 8: Scenarios
- Base Case (기본)
- Bull Case (낙관적, +30%)
- Bear Case (보수적, -20%)
- 시나리오별 P&L 요약
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class FPScenariosBuilder:
    """
    Financial Projection Scenarios 시트 빌더
    
    기능:
      - 3가지 시나리오 (Bear/Base/Bull)
      - 시나리오별 매출, EBITDA, Net Income
      - 비교 분석
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """Scenarios 시트 생성"""
        
        ws = self.wb.create_sheet("FP_Scenarios")
        
        # === 1. 제목 ===
        ws['A1'] = "Financial Scenarios Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "Bear/Base/Bull 3가지 시나리오 비교"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Metric (Year 5)', 'Bear Case', 'Base Case', 'Bull Case', 'Range']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 성장률 조정 ===
        row += 1
        ws.cell(row=row, column=1).value = "Growth Rate Adjustment"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "YoY Growth Rate"
        
        # Bear: -20%
        ws.cell(row=row, column=2).value = "=GrowthRateYoY*0.8"
        ws.cell(row=row, column=2).number_format = '0.0%'
        
        # Base
        ws.cell(row=row, column=3).value = "=GrowthRateYoY"
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Bull: +30%
        ws.cell(row=row, column=4).value = "=GrowthRateYoY*1.3"
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        # Range
        ws.cell(row=row, column=5).value = "=D{}-B{}".format(row, row)
        ws.cell(row=row, column=5).number_format = '0.0%'
        
        # === 4. Year 5 재무 지표 ===
        row += 2
        ws.cell(row=row, column=1).value = "Year 5 Financial Results"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # Revenue
        row += 1
        ws.cell(row=row, column=1).value = "Revenue (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # Bear: 20% 낮은 성장률로 5년 계산
        ws.cell(row=row, column=2).value = "=BaseRevenue*(1+GrowthRateYoY*0.8)^5"
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Base
        ws.cell(row=row, column=3).value = "=Revenue_Y5"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Bull: 30% 높은 성장률
        ws.cell(row=row, column=4).value = "=BaseRevenue*(1+GrowthRateYoY*1.3)^5"
        ws.cell(row=row, column=4).number_format = '#,##0'
        
        # Range
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # EBITDA
        row += 1
        ws.cell(row=row, column=1).value = "EBITDA (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Bear
        ws.cell(row=row, column=2).value = f"=B{row-1}*EBITDAMargin*0.9"  # EBITDA도 10% 낮춤
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Base
        ws.cell(row=row, column=3).value = "=EBITDA_Y5"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Bull
        ws.cell(row=row, column=4).value = f"=D{row-1}*EBITDAMargin*1.1"  # EBITDA 10% 높임
        ws.cell(row=row, column=4).number_format = '#,##0'
        
        # Range
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # Net Income
        row += 1
        ws.cell(row=row, column=1).value = "Net Income (Year 5)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # Bear
        ws.cell(row=row, column=2).value = f"=B{row-2}*NetMargin*0.8"  # Net Margin 20% 낮춤
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Base
        ws.cell(row=row, column=3).value = "=NetIncome_Y5"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Bull
        ws.cell(row=row, column=4).value = f"=D{row-2}*NetMargin*1.2"  # Net Margin 20% 높임
        ws.cell(row=row, column=4).number_format = '#,##0'
        
        # Range
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # === 5. 시나리오 설명 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 시나리오 설명"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Bear Case: 성장률 -20%, EBITDA Margin -10%, Net Margin -20%"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Base Case: 현재 가정 유지 (가장 현실적 시나리오)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Bull Case: 성장률 +30%, EBITDA Margin +10%, Net Margin +20%"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ FP Scenarios 시트 생성 완료")
        print(f"      - 3가지 시나리오 (Bear/Base/Bull)")


# 테스트는 별도 스크립트에서

