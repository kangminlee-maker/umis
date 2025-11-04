"""
Key Metrics Sheet Builder
핵심 재무 비율 시트

Sheet 7: Key_Metrics
- Revenue Growth (YoY, CAGR)
- Margin Metrics (Gross, EBITDA, Net)
- Efficiency Metrics
- Trend Analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class MetricsBuilder:
    """
    Key Metrics 시트 빌더
    
    기능:
      - 핵심 재무 비율 계산
      - 성장률, Margin 추이
      - 효율성 지표
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, years: int = 5, pl_sheet: str = 'PL_5Year') -> None:
        """
        Key Metrics 시트 생성
        
        Args:
            years: 예측 년수
            pl_sheet: P&L 시트 이름 (참조용)
        """
        
        ws = self.wb.create_sheet("Key_Metrics")
        
        # === 1. 제목 ===
        ws['A1'] = "Key Financial Metrics"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A1:{chr(65 + years + 1)}1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "핵심 재무 비율 및 성장 지표"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells(f'A2:{chr(65 + years + 1)}2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        for i in range(years + 1):
            col_letter = chr(66 + i)
            ws.column_dimensions[col_letter].width = 15
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Metric'] + [f'Year {y}' for y in range(years + 1)]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 성장 지표 ===
        row += 1
        ws.cell(row=row, column=1).value = "1. 성장 지표 (Growth Metrics)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # Revenue
        row += 1
        ws.cell(row=row, column=1).value = "Revenue"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        revenue_metric_row = row
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=Revenue_Y{year}'
            ws.cell(row=row, column=col).number_format = '#,##0'
        
        # YoY Growth %
        row += 1
        ws.cell(row=row, column=1).value = "  YoY Growth %"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)
        
        for year in range(1, years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            prev_col_letter = chr(65 + col - 1)
            
            # YoY = (This Year - Last Year) / Last Year
            ws.cell(row=row, column=col).value = (
                f'=({col_letter}{revenue_metric_row}-{prev_col_letter}{revenue_metric_row})/{prev_col_letter}{revenue_metric_row}'
            )
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True)
        
        # CAGR (Year 0 → Year N)
        row += 1
        ws.cell(row=row, column=1).value = "  CAGR (Cumulative)"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)
        
        for year in range(1, years + 1):
            col = 2 + year
            col_letter = chr(65 + col)
            
            # CAGR = (End / Start)^(1/Years) - 1
            ws.cell(row=row, column=col).value = (
                f'=(({col_letter}{revenue_metric_row}/B{revenue_metric_row})^(1/{year}))-1'
            )
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(italic=True, bold=True)
        
        # === 4. Margin 지표 ===
        row += 2
        ws.cell(row=row, column=1).value = "2. Margin 지표 (Profitability)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # Gross Margin %
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin %"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        for year in range(years + 1):
            col = 2 + year
            # P&L 시트에서 참조 (간단화: 고정 Margin)
            ws.cell(row=row, column=col).value = "=GrossMarginTarget"
            ws.cell(row=row, column=col).number_format = '0.0%'
        
        # EBITDA Margin %
        row += 1
        ws.cell(row=row, column=1).value = "EBITDA Margin %"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=EBITDA_Y{year}/Revenue_Y{year}'
            ws.cell(row=row, column=col).number_format = '0.0%'
        
        # Net Margin %
        row += 1
        ws.cell(row=row, column=1).value = "Net Margin %"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = f'=NetIncome_Y{year}/Revenue_Y{year}'
            ws.cell(row=row, column=col).number_format = '0.0%'
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        # === 5. 효율성 지표 ===
        row += 2
        ws.cell(row=row, column=1).value = "3. 효율성 지표 (Efficiency)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        # Revenue per Employee (간단화: 생략 또는 입력 필요)
        # OPEX % of Revenue
        row += 1
        ws.cell(row=row, column=1).value = "OPEX % of Revenue"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        for year in range(years + 1):
            col = 2 + year
            ws.cell(row=row, column=col).value = "=SMPercent+RDPercent+GAPercent"
            ws.cell(row=row, column=col).number_format = '0.0%'
        
        # === 6. 핵심 요약 ===
        row += 2
        ws.cell(row=row, column=1).value = "📊 핵심 요약"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:{chr(65 + years + 1)}{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = f"Year {years} CAGR:"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Year 5 CAGR 참조
        ws.cell(row=row, column=2).value = f'=C{revenue_metric_row + 2}'  # CAGR row
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = f"Year {years} Net Margin:"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Year 5 Net Margin 참조
        last_col_letter = chr(65 + 2 + years)
        ws.cell(row=row, column=2).value = f'={last_col_letter}{revenue_metric_row + 7}'  # Net Margin row
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        print(f"   ✅ Key Metrics 시트 생성 완료")
        print(f"      - 성장 지표 (Revenue, YoY, CAGR)")
        print(f"      - Margin 지표 (Gross, EBITDA, Net)")
        print(f"      - 효율성 지표")


# 테스트는 별도 스크립트에서

