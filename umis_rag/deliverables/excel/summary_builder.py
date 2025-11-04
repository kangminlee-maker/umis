"""
Summary Dashboard Builder
핵심 지표 및 분석 결과 요약 대시보드

구조:
  - 시장 개요
  - SAM 4가지 방법 비교
  - Convergence 상태
  - 시나리오 분석
  - 검증 상태
"""

from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

from .formula_engine import FormulaEngine, ExcelStyles


class SummaryBuilder:
    """
    Summary Dashboard 시트 빌더
    
    기능:
      - 핵심 지표 요약
      - SAM 방법별 비교
      - Convergence 상태
      - 시나리오 범위
      - 검증 완료율
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        market_name: Optional[str] = "Target Market"
    ) -> None:
        """
        Summary 시트 생성
        
        Args:
            market_name: 시장 이름
        
        구조:
          Section 1: 시장 개요
          Section 2: SAM 계산 결과
          Section 3: Convergence 분석
          Section 4: 시나리오 분석
          Section 5: 검증 상태
        """
        
        ws = self.wb.create_sheet("Summary", 0)  # 첫 번째 시트로
        
        # === 1. 대시보드 제목 ===
        ws['A1'] = "Market Sizing Summary Dashboard"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = market_name
        ws['A2'].font = Font(size=12, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        
        # === 2. Section 1: 핵심 지표 ===
        row = 4
        ws.cell(row=row, column=1).value = "1. Key Metrics"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        # TAM
        ws.cell(row=row, column=1).value = "Total Addressable Market (TAM):"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=TAM"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        
        row += 1
        # SAM (Average) - Convergence에서 가져옴
        ws.cell(row=row, column=1).value = "Serviceable Addressable Market (SAM):"
        ws.cell(row=row, column=1).font = Font(size=10)
        # Convergence의 평균은 B8 (C16이 아님)
        ws.cell(row=row, column=2).value = "=Convergence_Analysis!B8"  # 수정: B8 (평균 SAM)
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(size=10, bold=True, color="0070C0")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # === 3. Section 2: SAM 방법별 결과 ===
        row += 2
        ws.cell(row=row, column=1).value = "2. SAM Calculation Results"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        # 헤더
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        ws.cell(row=row, column=1).value = "Method"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "SAM"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "% of Avg"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        
        # 4가지 방법
        sam_methods = [
            ('Method 1: Top-Down', 'SAM'),
            ('Method 2: Bottom-Up', 'SAM_Method2'),
            ('Method 3: Proxy', 'SAM_Method3'),
            ('Method 4: Competitor Revenue', 'SAM_Method4')
        ]
        
        avg_sam_cell = f'B{row-2}'  # Section 1의 Average SAM 셀
        
        for method_name, sam_range in sam_methods:
            row += 1
            ws.cell(row=row, column=1).value = method_name
            ws.cell(row=row, column=1).font = Font(size=10)
            
            ws.cell(row=row, column=2).value = f"={sam_range}"
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            ws.cell(row=row, column=3).value = f"=B{row}/{avg_sam_cell}*100"
            ws.cell(row=row, column=3).number_format = '0.0"%"'
        
        # === 4. Section 3: Convergence 상태 ===
        row += 2
        ws.cell(row=row, column=1).value = "3. Convergence Analysis"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        # Max/Min Ratio
        ws.cell(row=row, column=1).value = "Max/Min Ratio:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Convergence_Analysis!B11"  # 수정: B11 (Max/Min)
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        
        row += 1
        # Convergence Status
        ws.cell(row=row, column=1).value = "Convergence Status:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Convergence_Analysis!B12"  # 수정: B12 (상태)
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        
        # 조건부 서식 (통과 = 녹색, 재검토 = 빨간색)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        pass_rule = FormulaRule(
            formula=[f'ISNUMBER(SEARCH("통과", B{row}))'],
            stopIfTrue=True,
            fill=green_fill,
            font=green_font
        )
        ws.conditional_formatting.add(f'B{row}', pass_rule)
        
        red_fill_2 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        fail_rule = FormulaRule(
            formula=[f'ISNUMBER(SEARCH("재검토", B{row}))'],
            stopIfTrue=True,
            fill=red_fill_2,
            font=red_font
        )
        ws.conditional_formatting.add(f'B{row}', fail_rule)
        
        row += 1
        # Standard Deviation
        ws.cell(row=row, column=1).value = "Standard Deviation:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Convergence_Analysis!C17"
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        row += 1
        # Coefficient of Variation
        ws.cell(row=row, column=1).value = "Coefficient of Variation (CV%):"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Convergence_Analysis!C18"
        ws.cell(row=row, column=2).number_format = '0.0"%"'
        
        # === 5. Section 4: 시나리오 분석 ===
        row += 2
        ws.cell(row=row, column=1).value = "4. Scenario Analysis"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        # 헤더
        ws.cell(row=row, column=1).value = "Scenario"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "Average SAM"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        # Best Case
        row += 1
        ws.cell(row=row, column=1).value = "Best Case (+15%)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=AvgSAM_Best"  # 수정: Named Range 사용
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Base Case
        row += 1
        ws.cell(row=row, column=1).value = "Base Case (Current)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=AvgSAM_Base"  # 수정: Named Range 사용
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        # Worst Case
        row += 1
        ws.cell(row=row, column=1).value = "Worst Case (-15%)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=AvgSAM_Worst"  # 수정: Named Range 사용
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Range
        row += 1
        ws.cell(row=row, column=1).value = "Range (Best - Worst)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = f"=B{row-3}-B{row-1}"  # Best - Worst
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        # === 6. Section 5: 검증 상태 ===
        row += 2
        ws.cell(row=row, column=1).value = "5. Validation Status"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        # Total Items (Validation_Log에서 참조)
        ws.cell(row=row, column=1).value = "Total Validation Items:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Validation_Log!B19"
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        
        row += 1
        # Validated
        ws.cell(row=row, column=1).value = "Validated:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Validation_Log!B20"
        ws.cell(row=row, column=2).font = Font(size=10, bold=True, color="006100")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        row += 1
        # Pending
        ws.cell(row=row, column=1).value = "Pending:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Validation_Log!B21"
        ws.cell(row=row, column=2).font = Font(size=10, bold=True, color="9C6500")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        row += 1
        # Completion Rate
        ws.cell(row=row, column=1).value = "Completion Rate:"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = "=Validation_Log!B22"
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        
        # === 7. 푸터 ===
        row += 3
        ws.cell(row=row, column=1).value = "📊 Detailed Analysis"
        ws.cell(row=row, column=1).font = Font(size=10, italic=True, color="666666")
        
        row += 1
        ws.cell(row=row, column=1).value = "• Assumptions: 가정 및 추정치 상세"
        ws.cell(row=row, column=1).font = Font(size=9, color="666666")
        
        row += 1
        ws.cell(row=row, column=1).value = "• Methods 1-4: 각 계산 방법 상세"
        ws.cell(row=row, column=1).font = Font(size=9, color="666666")
        
        row += 1
        ws.cell(row=row, column=1).value = "• Convergence Analysis: 수렴 분석 상세"
        ws.cell(row=row, column=1).font = Font(size=9, color="666666")
        
        row += 1
        ws.cell(row=row, column=1).value = "• Scenarios: 시나리오 분석 상세"
        ws.cell(row=row, column=1).font = Font(size=9, color="666666")
        
        row += 1
        ws.cell(row=row, column=1).value = "• Validation Log: 검증 이력 상세"
        ws.cell(row=row, column=1).font = Font(size=9, color="666666")
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        print(f"   ✅ Summary 시트 생성 완료 (Dashboard)")
        print(f"      - 5개 Section (핵심 지표, SAM, Convergence, 시나리오, 검증)")


# 테스트는 별도 스크립트에서
# python scripts/test_excel_generation.py

