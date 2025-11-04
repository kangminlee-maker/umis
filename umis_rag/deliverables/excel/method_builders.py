"""
Method Builders (4가지 SAM 계산 방법)

1. Top-Down: TAM에서 필터 적용
2. Bottom-Up: 고객 단위에서 합산
3. Proxy: 유사 시장 활용
4. Competitor Revenue: 경쟁사 매출 역산

피드백 반영:
  - Named Range SAM 2단계 정의
  - 절대참조 사용
"""

from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

from .formula_engine import FormulaEngine, ExcelStyles


class Method1TopDownBuilder:
    """
    Method 1: Top-Down Approach
    TAM → 필터들 순차 적용 → SAM
    
    피드백 반영:
      - SAM을 셀에 직접 쓰지 않고 Named Range로 정의
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        tam: Dict,
        narrowing_steps: List[Dict]
    ) -> None:
        """
        Method 1 시트 생성
        
        Args:
            tam: TAM 정의
                {
                    'value': 1000000000000,
                    'definition': '전체 악기 시장',
                    'source': 'ASM_001'
                }
            
            narrowing_steps: 필터 단계들
                [
                    {
                        'dimension': '지역',
                        'ratio_source': 'ASM_002',
                        'description': '한국 비중 15%'
                    },
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("Method_1_TopDown")
        
        # 제목
        ws['A1'] = "Method 1: Top-Down Approach"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # TAM 시작
        ws['A3'] = "TAM"
        ws['A4'] = tam.get('definition', '')
        ws['A5'] = self.fe.create_assumption_ref(tam.get('source', 'TAM_VALUE'))
        
        # TAM 셀 서식
        ws['A5'].fill = PatternFill(
            start_color=ExcelStyles.CALC_FILL,
            end_color=ExcelStyles.CALC_FILL,
            fill_type="solid"
        )
        ws['A5'].number_format = '#,##0'
        
        # 코멘트
        ws['A5'].comment = Comment(
            f"출처: {tam.get('source')}\n{tam.get('definition')}",
            "Bill"
        )
        
        # Narrowing Steps (필터들)
        col = ord('B')
        prev_result = "A5"
        
        for step in narrowing_steps:
            col_letter = chr(col)
            
            # 차원 이름
            ws[f'{col_letter}3'] = step.get('dimension')
            ws[f'{col_letter}3'].font = Font(bold=True)
            ws[f'{col_letter}3'].alignment = Alignment(horizontal="center")
            
            # 설명
            ws[f'{col_letter}4'] = step.get('description')
            
            # 비율 (Named Range 참조)
            ratio_source = step.get('ratio_source')
            ws[f'{col_letter}5'] = self.fe.create_assumption_ref(ratio_source)
            
            # 계산: 이전 결과 × 비율
            ws[f'{col_letter}6'] = f"={prev_result}*{col_letter}5"
            
            # 계산 결과 셀 서식
            ws[f'{col_letter}6'].fill = PatternFill(
                start_color=ExcelStyles.CALC_FILL,
                end_color=ExcelStyles.CALC_FILL,
                fill_type="solid"
            )
            ws[f'{col_letter}6'].number_format = '#,##0'
            
            # 코멘트
            ws[f'{col_letter}6'].comment = Comment(
                f"{step.get('description')}\n출처: {ratio_source}",
                "Bill"
            )
            
            prev_result = f"{col_letter}6"
            col += 1
        
        # 최종 SAM (피드백 반영!)
        final_col = chr(col - 1)
        final_cell = f"{final_col}6"
        
        # Step 1: 실제 셀에 결과 표시
        # (이미 위에서 계산됨)
        
        # Step 2: 그 셀을 가리키는 Named Range 'SAM' 정의
        self.fe.define_named_range(
            name='SAM',
            sheet='Method_1_TopDown',
            cell=final_cell,
            scope='workbook'  # Workbook-scope
        )
        
        # SAM 셀 강조
        ws[final_cell].fill = PatternFill(
            start_color=ExcelStyles.RESULT_FILL,
            end_color=ExcelStyles.RESULT_FILL,
            fill_type="solid"
        )
        ws[final_cell].font = Font(bold=True, size=12)
        
        # SAM 레이블
        ws[f'{final_col}2'] = "= SAM"
        ws[f'{final_col}2'].font = Font(bold=True, size=12)
        ws[f'{final_col}2'].alignment = Alignment(horizontal="center")
        
        # 검증
        ws['A10'] = "검증"
        ws['A10'].font = Font(bold=True)
        
        ws['A11'] = "TAM > SAM?"
        ws['B11'] = f'=IF(A5>{final_cell}, "✅", "❌")'
        
        ws['A12'] = "SAM > 0?"
        ws['B12'] = f'=IF({final_cell}>0, "✅", "❌")'
        
        print(f"   ✅ Method 1: Top-Down (SAM Named Range 정의)")


class Method2BottomUpBuilder:
    """Method 2: Bottom-Up Approach"""
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, segments: List[Dict]) -> None:
        """
        Method 2 시트 생성 (v7.2.0 - Narrowing 추가)
        
        Args:
            segments: 세그먼트 목록
                [
                    {
                        'name': '개인 구독',
                        'total_population': 'ASM_010',  # 전체 모집단
                        'narrowing_filters': [           # Narrowing 단계
                            {'name': '지역 (한국)', 'filter_id': 'FILTER_KOREA'},
                            {'name': '피아노 관심', 'filter_id': 'FILTER_PIANO'}
                        ],
                        'purchase_rate': 'PURCHASE_RATE_EST',  # Estimation_Details에서
                        'aov': 'AOV_EST',
                        'frequency': 'FREQUENCY_EST'
                    },
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("Method_2_BottomUp")
        
        # 제목
        ws['A1'] = "Method 2: Bottom-Up Approach (with Narrowing)"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:I1')
        
        ws['A2'] = "총 모집단 → Narrowing Filters → 타겟 고객 → 구매 행동"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:I2')
        
        # 헤더
        ws['A4'] = "Segment"
        ws['B4'] = "Total Population"
        ws['C4'] = "Narrowing Filters"
        ws['D4'] = "Narrowed Customers"
        ws['E4'] = "Purchase Rate"
        ws['F4'] = "AOV"
        ws['G4'] = "Frequency"
        ws['H4'] = "SAM"
        ws['I4'] = "Notes"
        
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=ExcelStyles.HEADER_FILL,
                end_color=ExcelStyles.HEADER_FILL,
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
        
        # 열 너비
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 25
        
        # 각 세그먼트
        segment_sam_ranges = []
        
        for idx, seg in enumerate(segments, start=1):
            i = 4 + idx  # Row 5, 6, 7...
            ws[f'A{i}'] = seg.get('name')
            ws[f'A{i}'].font = Font(bold=True)
            
            # B: Total Population
            ws[f'B{i}'] = self.fe.create_assumption_ref(seg.get('total_population'))
            ws[f'B{i}'].number_format = '#,##0'
            
            # C: Narrowing Filters (설명만)
            narrowing_filters = seg.get('narrowing_filters', [])
            filter_names = [f['name'] for f in narrowing_filters]
            ws[f'C{i}'] = ' × '.join(filter_names) if filter_names else '없음'
            ws[f'C{i}'].font = Font(size=9)
            
            # D: Narrowed Customers = Total Population × Filter1 × Filter2 × ...
            if narrowing_filters:
                filter_refs = [self.fe.create_assumption_ref(f['filter_id']) for f in narrowing_filters]
                narrowed_formula = f"=B{i}" + "".join([f"*{ref}" for ref in filter_refs])
            else:
                narrowed_formula = f"=B{i}"
            
            ws[f'D{i}'] = narrowed_formula
            ws[f'D{i}'].number_format = '#,##0'
            ws[f'D{i}'].fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )
            
            # E: Purchase Rate (Estimation_Details에서 참조)
            ws[f'E{i}'] = self.fe.create_assumption_ref(seg.get('purchase_rate'))
            ws[f'E{i}'].number_format = '0.0%'
            
            # F: AOV
            ws[f'F{i}'] = self.fe.create_assumption_ref(seg.get('aov'))
            ws[f'F{i}'].number_format = '#,##0'
            
            # G: Frequency
            ws[f'G{i}'] = self.fe.create_assumption_ref(seg.get('frequency'))
            ws[f'G{i}'].number_format = '0.0'
            
            # H: SAM = Narrowed Customers × Purchase Rate × AOV × Frequency
            ws[f'H{i}'] = f"=D{i}*E{i}*F{i}*G{i}"
            ws[f'H{i}'].number_format = '#,##0'
            ws[f'H{i}'].fill = PatternFill(
                start_color=ExcelStyles.CALC_FILL,
                end_color=ExcelStyles.CALC_FILL,
                fill_type="solid"
            )
            
            # I: Notes
            ws[f'I{i}'] = seg.get('notes', '')
            ws[f'I{i}'].font = Font(size=9, italic=True)
            
            # Named Range 정의
            seg_sam_name = f'M2_Seg{idx}_SAM'
            self.fe.define_named_range(seg_sam_name, 'Method_2_BottomUp', f'H{i}')
            segment_sam_ranges.append(seg_sam_name)
        
        # 총 SAM (Named Range 기반)
        last_row = 4 + len(segments)
        total_row = last_row + 2
        
        ws[f'A{total_row}'] = "Total SAM"
        ws[f'A{total_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{total_row}'].fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        
        # Named Range 기반 SUM
        if len(segment_sam_ranges) == 1:
            sum_formula = f"={segment_sam_ranges[0]}"
        else:
            sum_formula = f"=SUM({','.join(segment_sam_ranges)})"
        
        ws[f'H{total_row}'] = sum_formula
        ws[f'H{total_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'H{total_row}'].fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        ws[f'H{total_row}'].number_format = '#,##0'
        ws[f'F{total_row}'].number_format = '#,##0'
        
        # SAM Named Range
        self.fe.define_named_range(
            name='SAM_Method2',
            sheet='Method_2_BottomUp',
            cell=f'F{total_row}',
            scope='workbook'
        )
        
        print(f"   ✅ Method 2: Bottom-Up ({len(segments)}개 세그먼트)")


class Method3ProxyBuilder:
    """Method 3: Proxy Method"""
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, proxy_data: Dict) -> None:
        """
        Method 3 시트 생성 (v7.2.0 - 메타데이터 추가)
        
        Args:
            proxy_data: Proxy 데이터
                {
                    'proxy_market_name': '바이올린 구독 시장',  # 추가
                    'proxy_market': 'ASM_020',
                    'similarity_reason': '유사한 악기 구독 모델',  # 추가
                    'correlation': 'PROXY_CORR',
                    'correlation_basis': 'SNS 관심도 상관계수 0.3',  # 추가
                    'application_rate': 'PROXY_APP',
                    'application_basis': '시장 성숙도 차이 50% 보정'  # 추가
                }
        """
        
        ws = self.wb.create_sheet("Method_3_Proxy")
        
        # === 제목 ===
        ws['A1'] = "Method 3: Proxy Method (유사 시장 활용)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 25
        
        ws['A2'] = "유사 시장 규모를 기반으로 추정"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        
        # === Proxy 시장 정보 ===
        row = 4
        ws.cell(row=row, column=1).value = "📊 Proxy 시장"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True, color="0066CC")
        
        row += 1
        ws.cell(row=row, column=1).value = "Proxy 시장 이름:"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = proxy_data.get('proxy_market_name', '유사 시장')
        ws.cell(row=row, column=2).font = Font(bold=True, color="0066CC")
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "유사성 근거:"
        ws.cell(row=row, column=2).value = proxy_data.get('similarity_reason', '유사한 제품/서비스')
        ws.cell(row=row, column=2).font = Font(size=9)
        ws.merge_cells(f'B{row}:D{row}')
        
        # === 계산 섹션 ===
        row += 2
        ws.cell(row=row, column=1).value = "📐 계산"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True, color="0066CC")
        
        # Proxy Market Size
        row += 1
        proxy_size_row = row
        ws.cell(row=row, column=1).value = "Proxy Market Size"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = self.fe.create_assumption_ref(
            proxy_data.get('proxy_market', 'PROXY_SIZE')
        )
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=4).value = "유사 시장의 규모"
        ws.cell(row=row, column=4).font = Font(size=9, italic=True)
        
        # Correlation Coefficient
        row += 1
        ws.cell(row=row, column=1).value = "Correlation Coefficient"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = self.fe.create_assumption_ref(
            proxy_data.get('correlation', 'PROXY_CORR')
        )
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=4).value = proxy_data.get('correlation_basis', '상관관계 추정 근거')
        ws.cell(row=row, column=4).font = Font(size=9, italic=True)
        
        # Application Rate
        row += 1
        ws.cell(row=row, column=1).value = "Application Rate"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = self.fe.create_assumption_ref(
            proxy_data.get('application_rate', 'PROXY_APP')
        )
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=4).value = proxy_data.get('application_basis', '적용 비율 근거')
        ws.cell(row=row, column=4).font = Font(size=9, italic=True)
        
        # === SAM ===
        row += 2
        ws.cell(row=row, column=1).value = "SAM (Proxy)"
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        
        ws.cell(row=row, column=2).value = f"=B{proxy_size_row}*B{proxy_size_row+1}*B{proxy_size_row+2}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=row, column=2).fill = PatternFill(
            start_color="2E75B6", end_color="2E75B6", fill_type="solid"
        )
        
        ws.cell(row=row, column=4).value = "= Proxy Size × Correlation × Application"
        ws.cell(row=row, column=4).font = Font(size=9, italic=True)
        
        # SAM Named Range
        self.fe.define_named_range(
            name='SAM_Method3',
            sheet='Method_3_Proxy',
            cell=f'B{row}',
            scope='workbook'
        )
        
        # === 설명 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Correlation: 우리 시장과 Proxy 시장의 상관관계"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Application: Proxy 대비 우리 시장의 적용 비율 (성숙도, 규모 등 보정)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ Method 3: Proxy (메타데이터 포함)")


class Method4CompetitorBuilder:
    """Method 4: Competitor Revenue"""
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, competitors: List[Dict]) -> None:
        """
        Method 4 시트 생성
        
        Args:
            competitors: 경쟁사 목록
                [
                    {
                        'company': '경쟁사A',
                        'revenue': 'ASM_030',
                        'market_share': 'ASM_031'
                    },
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("Method_4_CompetitorRevenue")
        
        # 제목
        ws['A1'] = "Method 4: Competitor Revenue Approach"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 헤더
        ws['A3'] = "Company"
        ws['B3'] = "Revenue"
        ws['C3'] = "Market Share"
        
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 경쟁사 데이터 + Named Range
        comp_rev_ranges = []
        comp_share_ranges = []
        
        for idx, comp in enumerate(competitors, start=1):
            i = 3 + idx  # Row 4, 5, 6, ...
            
            ws[f'A{i}'] = comp.get('company')
            ws[f'B{i}'] = self.fe.create_assumption_ref(comp.get('revenue'))
            ws[f'C{i}'] = self.fe.create_assumption_ref(comp.get('market_share'))
            
            ws[f'B{i}'].number_format = '#,##0'
            ws[f'C{i}'].number_format = '0.0%'
            
            # Named Range 정의
            rev_name = f'M4_Comp{idx}_Rev'
            share_name = f'M4_Comp{idx}_Share'
            self.fe.define_named_range(rev_name, 'Method_4_CompetitorRevenue', f'B{i}')
            self.fe.define_named_range(share_name, 'Method_4_CompetitorRevenue', f'C{i}')
            comp_rev_ranges.append(rev_name)
            comp_share_ranges.append(share_name)
        
        # 합계 (Named Range 기반)
        last_row = 4 + len(competitors) - 1
        total_row = last_row + 1
        
        ws[f'A{total_row}'] = "Total"
        ws[f'A{total_row}'].font = Font(bold=True)
        
        if len(comp_rev_ranges) == 1:
            ws[f'B{total_row}'] = f"={comp_rev_ranges[0]}"
            ws[f'C{total_row}'] = f"={comp_share_ranges[0]}"
        else:
            ws[f'B{total_row}'] = f"=SUM({','.join(comp_rev_ranges)})"
            ws[f'C{total_row}'] = f"=SUM({','.join(comp_share_ranges)})"
        
        ws[f'B{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].font = Font(bold=True)
        
        # SAM 역산
        sam_row = total_row + 2
        
        ws[f'A{sam_row}'] = "SAM (역산)"
        ws[f'A{sam_row}'].font = Font(bold=True)
        
        ws[f'B{sam_row}'] = f"=B{total_row}/C{total_row}"
        ws[f'B{sam_row}'].number_format = '#,##0'
        ws[f'B{sam_row}'].font = Font(bold=True, size=12)
        ws[f'B{sam_row}'].fill = PatternFill(
            start_color=ExcelStyles.RESULT_FILL,
            end_color=ExcelStyles.RESULT_FILL,
            fill_type="solid"
        )
        
        # SAM Named Range
        self.fe.define_named_range(
            name='SAM_Method4',
            sheet='Method_4_CompetitorRevenue',
            cell=f'B{sam_row}',
            scope='workbook'
        )
        
        # 검증
        val_row = sam_row + 2
        ws[f'A{val_row}'] = "검증: Total Share <= 100%?"
        ws[f'B{val_row}'] = f'=IF(C{total_row}<=1, "✅", "❌ 점유율 합 > 100%")'
        
        print(f"   ✅ Method 4: Competitor Revenue ({len(competitors)}개 경쟁사)")

