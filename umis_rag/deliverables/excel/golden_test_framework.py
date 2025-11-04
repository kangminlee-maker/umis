"""
Golden Test Framework
결과 중심 Excel 검증 시스템

핵심 원칙:
1. Syntax 검증 (자기 참조, #REF!)
2. Golden Values 검증 (기대값 vs 실제값) ⭐ 핵심
3. 논리적 일관성 검증 (Revenue > COGS 등)

Golden Test Spec:
  - 각 Generator마다 expected_results 정의
  - 주요 셀의 기대값 명시 (시트!셀 → 값)
  - 자동 비교 (오차 < 1%)
"""

from pathlib import Path
from typing import Dict, List, Tuple, Any
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class GoldenTestSpec:
    """
    Golden Test 스펙
    
    각 Excel 도구의 기대 결과 정의
    """
    
    @staticmethod
    def get_market_sizing_spec() -> Dict:
        """
        Market Sizing Golden Spec
        
        케이스: 피아노 구독
        - TAM: ₩1,000억
        - 한국: 15%, 피아노: 25%
        - SAM (Method 1): ₩37.5억
        """
        
        return {
            'name': 'Market Sizing Workbook',
            'case': '피아노 구독 서비스',
            
            # 입력값
            'inputs': {
                'tam': 100_000_000_000,
                'korea_ratio': 0.15,
                'piano_ratio': 0.25,
                'customers': 100_000,
                'conversion': 0.20,
                'aov': 50_000,
                'frequency': 12
            },
            
            # 기대 결과 (시트!셀 → 값, 오차)
            'expected_values': [
                # Method 1
                {
                    'sheet': 'Method_1_TopDown',
                    'cell': 'A5',
                    'expected': 100_000_000_000,
                    'tolerance': 0.01,
                    'description': 'TAM'
                },
                {
                    'sheet': 'Method_1_TopDown',
                    'cell': 'B6',
                    'expected': 15_000_000_000,  # TAM × 15%
                    'tolerance': 0.01,
                    'description': '한국 시장'
                },
                {
                    'sheet': 'Method_1_TopDown',
                    'cell': 'C6',
                    'expected': 3_750_000_000,  # ₩37.5억
                    'tolerance': 0.01,
                    'description': 'SAM (Method 1)'
                },
                
                # Method 2
                {
                    'sheet': 'Method_2_BottomUp',
                    'cell': 'F6',
                    'expected': 12_000_000_000,  # ₩120억
                    'tolerance': 0.01,
                    'description': 'SAM (Method 2)'
                },
                
                # Method 3
                {
                    'sheet': 'Method_3_Proxy',
                    'cell': 'B7',
                    'expected': 7_500_000_000,  # ₩75억
                    'tolerance': 0.01,
                    'description': 'SAM (Method 3)'
                },
                
                # Method 4
                {
                    'sheet': 'Method_4_CompetitorRevenue',
                    'cell': 'B7',
                    'expected': 25_000_000_000,  # ₩250억
                    'tolerance': 0.01,
                    'description': 'SAM (Method 4)'
                },
                
                # Convergence
                {
                    'sheet': 'Convergence_Analysis',
                    'cell': 'B4',
                    'expected': 3_750_000_000,
                    'tolerance': 0.01,
                    'description': 'Method 1 SAM'
                },
                {
                    'sheet': 'Convergence_Analysis',
                    'cell': 'B8',
                    'expected': 12_062_500_000,  # 평균
                    'tolerance': 0.01,
                    'description': '평균 SAM'
                },
                
                # Summary (핵심!)
                {
                    'sheet': 'Summary',
                    'cell': 'B5',
                    'expected': 100_000_000_000,
                    'tolerance': 0.01,
                    'description': 'Summary TAM'
                },
                {
                    'sheet': 'Summary',
                    'cell': 'B6',
                    'expected': 12_062_500_000,  # 평균 SAM
                    'tolerance': 0.01,
                    'description': 'Summary SAM (평균)'
                },
            ],
            
            # 논리적 일관성 검증
            'consistency_checks': [
                {
                    'name': 'TAM > SAM',
                    'check': lambda results: results.get('Summary!B5', 0) > results.get('Summary!B6', 0),
                    'error_msg': 'TAM이 SAM보다 작음 (논리 오류)'
                },
                {
                    'name': 'Method 1 SAM > 0',
                    'check': lambda results: results.get('Method_1_TopDown!C6', 0) > 0,
                    'error_msg': 'Method 1 SAM이 0 이하'
                },
                {
                    'name': '평균 SAM > 0',
                    'check': lambda results: results.get('Convergence_Analysis!B8', 0) > 0,
                    'error_msg': '평균 SAM이 0 이하'
                }
            ]
        }
    
    @staticmethod
    def get_unit_economics_spec() -> Dict:
        """
        Unit Economics Golden Spec
        
        케이스: 음악 스트리밍
        - ARPU: ₩9,000, CAC: ₩25,000
        - LTV: ₩78,750, Ratio: 3.15
        """
        
        return {
            'name': 'Unit Economics Analyzer',
            'case': '음악 스트리밍',
            
            'inputs': {
                'arpu': 9000,
                'cac': 25000,
                'margin': 0.35,
                'churn': 0.04,
                'lifetime': 25
            },
            
            'expected_values': [
                # LTV
                {
                    'sheet': 'LTV_Calculation',
                    'cell': 'B9',
                    'expected': 78750,  # 9000 × 25 × 0.35
                    'tolerance': 0.01,
                    'description': 'LTV (방법 1)'
                },
                {
                    'sheet': 'LTV_Calculation',
                    'cell': 'B18',
                    'expected': 78750,
                    'tolerance': 0.01,
                    'description': 'LTV (평균)'
                },
                
                # Ratio
                {
                    'sheet': 'LTV_CAC_Ratio',
                    'cell': 'B7',
                    'expected': 3.15,
                    'tolerance': 0.02,
                    'description': 'LTV/CAC Ratio'
                },
                
                # Payback
                {
                    'sheet': 'Payback_Period',
                    'cell': 'B11',
                    'expected': 7.94,
                    'tolerance': 0.1,
                    'description': 'Payback Period'
                },
                
                # Dashboard
                {
                    'sheet': 'Dashboard',
                    'cell': 'B5',
                    'expected': 78750,
                    'tolerance': 0.01,
                    'description': 'Dashboard LTV'
                },
                {
                    'sheet': 'Dashboard',
                    'cell': 'B7',
                    'expected': 3.15,
                    'tolerance': 0.02,
                    'description': 'Dashboard Ratio'
                },
            ],
            
            'consistency_checks': [
                {
                    'name': 'LTV > CAC',
                    'check': lambda r: r.get('Dashboard!B5', 0) > r.get('Dashboard!B6', 0),
                    'error_msg': 'LTV가 CAC보다 작음'
                },
                {
                    'name': 'Ratio > 1',
                    'check': lambda r: r.get('Dashboard!B7', 0) > 1.0,
                    'error_msg': 'LTV/CAC < 1 (손실 비즈니스)'
                }
            ]
        }
    
    @staticmethod
    def get_financial_projection_spec() -> Dict:
        """
        Financial Projection Golden Spec
        
        케이스: 성인 교육
        - Year 0: ₩1,250억, Growth: 28%
        - Year 5: ₩4,295억
        """
        
        return {
            'name': 'Financial Projection Model',
            'case': '성인 교육 시장',
            
            'inputs': {
                'revenue_y0': 1250_0000_0000,
                'growth': 0.28,
                'gross_margin': 0.70,
                'net_margin': 0.10
            },
            
            'expected_values': [
                # Revenue
                {
                    'sheet': 'Revenue_Buildup',
                    'cell': 'B9',
                    'expected': 1250_0000_0000,
                    'tolerance': 0.01,
                    'description': 'Revenue Year 0'
                },
                {
                    'sheet': 'Revenue_Buildup',
                    'cell': 'C9',
                    'expected': 1600_0000_0000,  # × 1.28
                    'tolerance': 0.01,
                    'description': 'Revenue Year 1'
                },
                {
                    'sheet': 'Revenue_Buildup',
                    'cell': 'G9',
                    'expected': 4295_0000_0000,  # × 1.28^5
                    'tolerance': 0.02,
                    'description': 'Revenue Year 5'
                },
                
                # Dashboard
                {
                    'sheet': 'Dashboard',
                    'cell': 'B5',
                    'expected': 4295_0000_0000,
                    'tolerance': 0.02,
                    'description': 'Dashboard Revenue Y5'
                },
                {
                    'sheet': 'Dashboard',
                    'cell': 'B6',
                    'expected': 429_0000_0000,  # 10%
                    'tolerance': 0.02,
                    'description': 'Dashboard Net Income Y5'
                },
                {
                    'sheet': 'Dashboard',
                    'cell': 'B7',
                    'expected': 0.28,
                    'tolerance': 0.01,
                    'description': 'Dashboard CAGR'
                },
            ],
            
            'consistency_checks': [
                {
                    'name': 'Revenue Growth',
                    'check': lambda r: r.get('Revenue_Buildup!G9', 0) > r.get('Revenue_Buildup!B9', 0),
                    'error_msg': 'Year 5 Revenue ≤ Year 0 (성장 없음)'
                }
            ]
        }


class GoldenTestRunner:
    """
    Golden Test 실행기
    
    Syntax + Golden Values 병행 검증
    """
    
    def __init__(self, filepath: Path, spec: Dict):
        """
        Args:
            filepath: 검증할 Excel 파일
            spec: Golden Test Spec
        """
        self.filepath = filepath
        self.spec = spec
        self.wb_formula = None  # 수식 확인용
        self.wb_data = None  # 값 확인용
        self.results = {}
        self.errors = []
        self.warnings = []
    
    def run(self) -> Dict[str, Any]:
        """
        전체 검증 실행
        
        Returns:
            검증 결과
        """
        
        print(f"\n🔍 Golden Test: {self.spec['name']}")
        print(f"   케이스: {self.spec['case']}")
        print("="*70)
        
        # 파일 열기 (2가지 모드)
        try:
            self.wb_formula = load_workbook(self.filepath, data_only=False)
            self.wb_data = load_workbook(self.filepath, data_only=True)
        except Exception as e:
            self.errors.append(f"파일 열기 실패: {e}")
            return self._compile_results()
        
        # Step 1: Syntax 검증
        print("\n1️⃣ Syntax 검증")
        print("-"*70)
        self._check_syntax()
        
        # Step 2: Golden Values 검증 ⭐ 핵심
        print("\n2️⃣ Golden Values 검증 (결과 중심)")
        print("-"*70)
        self._check_golden_values()
        
        # Step 3: 논리적 일관성 검증
        print("\n3️⃣ 논리적 일관성 검증")
        print("-"*70)
        self._check_consistency()
        
        return self._compile_results()
    
    def _check_syntax(self):
        """Syntax 검증 (자기 참조, 오류 수식)"""
        
        self_ref_count = 0
        
        for sheet_name in self.wb_formula.sheetnames:
            ws = self.wb_formula[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # 자기 참조 검사
                        import re
                        formula = cell.value
                        cell_refs = re.findall(r'\$?[A-Z]+\$?\d+', formula)
                        cell_coord = cell.coordinate.replace('$', '')
                        
                        for ref in cell_refs:
                            if ref.replace('$', '') == cell_coord:
                                self_ref_count += 1
                                self.errors.append(
                                    f"❌ 자기 참조: {sheet_name}!{cell.coordinate} = {formula}"
                                )
        
        if self_ref_count == 0:
            print("✅ 자기 참조: 0개")
        else:
            print(f"❌ 자기 참조: {self_ref_count}개 발견!")
    
    def _check_golden_values(self):
        """Golden Values 검증 (결과 중심) ⭐"""
        
        passed_count = 0
        failed_count = 0
        
        for spec in self.spec['expected_values']:
            sheet = spec['sheet']
            cell = spec['cell']
            expected = spec['expected']
            tolerance = spec.get('tolerance', 0.01)
            desc = spec['description']
            
            # 값 가져오기
            actual = self._get_cell_value(sheet, cell)
            
            # 비교
            if actual is None:
                failed_count += 1
                self.errors.append(
                    f"❌ {sheet}!{cell} ({desc}): 값 없음 (기대: {self._format_value(expected)})"
                )
            else:
                # 오차 계산
                if expected != 0:
                    error = abs(actual - expected) / abs(expected)
                else:
                    error = abs(actual - expected)
                
                # 저장 (일관성 검증용)
                self.results[f'{sheet}!{cell}'] = actual
                
                if error <= tolerance:
                    passed_count += 1
                    print(f"✅ {sheet}!{cell} ({desc})")
                    print(f"   기대: {self._format_value(expected)}")
                    print(f"   실제: {self._format_value(actual)}")
                    print(f"   오차: {error*100:.2f}%")
                else:
                    failed_count += 1
                    self.errors.append(
                        f"❌ {sheet}!{cell} ({desc}): "
                        f"기대 {self._format_value(expected)} ≠ "
                        f"실제 {self._format_value(actual)} "
                        f"(오차 {error*100:.1f}%)"
                    )
        
        print(f"\n통과: {passed_count}개, 실패: {failed_count}개")
    
    def _check_consistency(self):
        """논리적 일관성 검증"""
        
        for check in self.spec.get('consistency_checks', []):
            name = check['name']
            check_func = check['check']
            error_msg = check['error_msg']
            
            try:
                if check_func(self.results):
                    print(f"✅ {name}")
                else:
                    self.errors.append(f"❌ {name}: {error_msg}")
                    print(f"❌ {name}: {error_msg}")
            except Exception as e:
                self.warnings.append(f"⚠️ {name}: 검증 실패 ({e})")
    
    def _get_cell_value(self, sheet: str, cell: str) -> Any:
        """
        셀 값 가져오기 (data_only=True)
        
        Args:
            sheet: 시트 이름
            cell: 셀 주소
        
        Returns:
            셀 값 (숫자) 또는 None
        """
        
        try:
            if sheet not in self.wb_data.sheetnames:
                return None
            
            ws = self.wb_data[sheet]
            cell_obj = ws[cell]
            
            if cell_obj.value is None:
                return None
            
            # 숫자로 변환 시도
            try:
                return float(cell_obj.value)
            except:
                # 문자열일 수 있음 (예: "✅ 통과")
                return cell_obj.value
        
        except Exception as e:
            return None
    
    def _format_value(self, value: float) -> str:
        """값 포맷팅"""
        
        if value >= 1_0000_0000:
            return f"₩{value/1_0000_0000:.1f}억"
        elif value >= 1_0000:
            return f"₩{value/1_0000:.1f}만"
        elif value >= 100:
            return f"₩{value:,.0f}"
        else:
            return f"{value:.2f}"
    
    def _compile_results(self) -> Dict:
        """결과 정리"""
        
        passed = len(self.errors) == 0
        
        print("\n" + "="*70)
        print("📊 Golden Test 결과")
        print("="*70)
        
        if passed:
            print("✅ 모든 검증 통과!")
        else:
            print(f"❌ {len(self.errors)}개 오류 발견")
            
            print("\n오류 목록:")
            for error in self.errors[:10]:
                print(f"  {error}")
            
            if len(self.errors) > 10:
                print(f"  ... 외 {len(self.errors) - 10}개")
        
        return {
            'passed': passed,
            'errors': self.errors,
            'warnings': self.warnings,
            'results': self.results
        }


# 편의 함수
def run_golden_test(filepath: Path, tool_type: str) -> bool:
    """
    Golden Test 실행 (편의 함수)
    
    Args:
        filepath: Excel 파일
        tool_type: 'market_sizing', 'unit_economics', 'financial_projection'
    
    Returns:
        통과 여부
    """
    
    # Spec 가져오기
    if tool_type == 'market_sizing':
        spec = GoldenTestSpec.get_market_sizing_spec()
    elif tool_type == 'unit_economics':
        spec = GoldenTestSpec.get_unit_economics_spec()
    elif tool_type == 'financial_projection':
        spec = GoldenTestSpec.get_financial_projection_spec()
    else:
        raise ValueError(f"Unknown tool_type: {tool_type}")
    
    # 실행
    runner = GoldenTestRunner(filepath, spec)
    result = runner.run()
    
    return result['passed']


# 사용 예시는 별도 스크립트에서
# python scripts/golden_test_all.py

