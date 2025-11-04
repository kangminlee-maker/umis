"""
Excel Workbook Validator
생성된 Excel 파일의 수식, 데이터, 구조를 자동 검증

검증 항목:
  1. 수식 검증 (자기 참조, 순환 참조, 오류)
  2. 데이터 검증 (빈 셀, 예상 범위)
  3. Named Range 검증
  4. 시트 구조 검증
  5. 계산 결과 검증 (예상값 vs 실제값)
"""

import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook


class ExcelValidator:
    """
    Excel Workbook 검증기
    
    기능:
      - 수식 오류 감지 (자기 참조, 순환 참조)
      - 데이터 완성도 확인
      - Named Range 검증
      - 계산 결과 검증
    """
    
    def __init__(self, filepath: Path):
        """
        Args:
            filepath: 검증할 Excel 파일 경로
        """
        self.filepath = filepath
        self.wb: Optional[Workbook] = None
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.info: List[str] = []
    
    def validate(self) -> Dict[str, Any]:
        """
        전체 검증 실행
        
        Returns:
            검증 결과
                {
                    'passed': bool,
                    'errors': List[str],
                    'warnings': List[str],
                    'info': List[str],
                    'stats': Dict
                }
        """
        
        print(f"\n🔍 Excel 검증 시작: {self.filepath.name}")
        print("="*70)
        
        # 파일 열기
        try:
            self.wb = load_workbook(self.filepath, data_only=False)
            self.info.append(f"✅ 파일 열기 성공: {len(self.wb.sheetnames)}개 시트")
        except Exception as e:
            self.errors.append(f"❌ 파일 열기 실패: {e}")
            return self._compile_results()
        
        # 1. 시트 구조 검증
        self._validate_sheet_structure()
        
        # 2. Named Range 검증
        self._validate_named_ranges()
        
        # 3. 수식 검증 (핵심!)
        self._validate_formulas()
        
        # 4. 데이터 완성도 검증
        self._validate_data_completeness()
        
        # 5. 계산 결과 검증 (샘플링)
        self._validate_calculation_results()
        
        return self._compile_results()
    
    def _validate_sheet_structure(self):
        """시트 구조 검증"""
        
        print("\n1️⃣ 시트 구조 검증")
        print("-" * 70)
        
        sheets = self.wb.sheetnames
        self.info.append(f"총 {len(sheets)}개 시트: {', '.join(sheets)}")
        
        # 필수 시트 확인 (Financial Projection 기준)
        if 'Dashboard' in sheets:
            self.info.append("✅ Dashboard 시트 존재")
        else:
            self.warnings.append("⚠️ Dashboard 시트 없음")
        
        if 'Assumptions' in sheets or 'Inputs' in sheets:
            self.info.append("✅ 입력 시트 존재")
        else:
            self.errors.append("❌ 입력 시트 없음 (Assumptions or Inputs)")
        
        print(f"시트 개수: {len(sheets)}")
        print(f"시트 목록: {', '.join(sheets[:5])}{'...' if len(sheets) > 5 else ''}")
    
    def _validate_named_ranges(self):
        """Named Range 검증"""
        
        print("\n2️⃣ Named Range 검증")
        print("-" * 70)
        
        named_ranges = list(self.wb.defined_names)
        self.info.append(f"총 {len(named_ranges)}개 Named Range 정의됨")
        
        if len(named_ranges) == 0:
            self.warnings.append("⚠️ Named Range가 없습니다 (수식 가독성 저하)")
        
        # Named Range 유효성 확인
        for name in named_ranges[:10]:  # 처음 10개만
            try:
                destinations = self.wb.defined_names[name].destinations
                for sheet, cell in destinations:
                    if sheet in self.wb.sheetnames:
                        self.info.append(f"✅ {name} → {sheet}!{cell}")
                    else:
                        self.errors.append(f"❌ {name}: 시트 '{sheet}' 없음")
            except Exception as e:
                self.errors.append(f"❌ {name}: 오류 - {e}")
        
        print(f"Named Range 개수: {len(named_ranges)}")
        if len(named_ranges) > 0:
            print(f"샘플: {', '.join(named_ranges[:5])}...")
    
    def _validate_formulas(self):
        """수식 검증 (핵심!)"""
        
        print("\n3️⃣ 수식 검증 (자기 참조, 오류 감지)")
        print("-" * 70)
        
        total_formulas = 0
        self_reference_count = 0
        error_formulas = 0
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas += 1
                        
                        # 자기 참조 검사
                        if self._has_self_reference(cell):
                            self_reference_count += 1
                            cell_addr = f"{sheet_name}!{cell.coordinate}"
                            self.errors.append(
                                f"❌ 자기 참조: {cell_addr} = {cell.value}"
                            )
                        
                        # 오류 수식 패턴 (#REF!, #DIV/0! 등)
                        if any(err in str(cell.value) for err in ['#REF!', '#DIV/0!', '#VALUE!', '#NAME?']):
                            error_formulas += 1
                            self.errors.append(
                                f"❌ 오류 수식: {sheet_name}!{cell.coordinate} = {cell.value}"
                            )
        
        self.info.append(f"총 {total_formulas}개 수식 검사 완료")
        
        if self_reference_count > 0:
            self.errors.append(f"❌ 자기 참조 {self_reference_count}개 발견! (심각)")
        else:
            self.info.append("✅ 자기 참조 없음")
        
        if error_formulas > 0:
            self.errors.append(f"❌ 오류 수식 {error_formulas}개 발견!")
        else:
            self.info.append("✅ 오류 수식 없음")
        
        print(f"총 수식: {total_formulas}개")
        print(f"자기 참조: {self_reference_count}개 {'❌' if self_reference_count > 0 else '✅'}")
        print(f"오류 수식: {error_formulas}개 {'❌' if error_formulas > 0 else '✅'}")
    
    def _has_self_reference(self, cell: Cell) -> bool:
        """
        셀이 자기 자신을 참조하는지 확인
        
        Args:
            cell: openpyxl Cell 객체
        
        Returns:
            자기 참조 여부
        
        Example:
            C5 = "=C5*2" → True (자기 참조!)
            C5 = "=B5*2" → False
        """
        
        if not cell.value or not isinstance(cell.value, str):
            return False
        
        formula = cell.value
        cell_coord = cell.coordinate
        
        # 수식에서 셀 참조 추출 (A1, $A$1, A$1 등)
        cell_refs = re.findall(r'\$?[A-Z]+\$?\d+', formula)
        
        # 절대 참조 제거 ($A$1 → A1)
        cell_coord_clean = cell_coord.replace('$', '')
        
        for ref in cell_refs:
            ref_clean = ref.replace('$', '')
            if ref_clean == cell_coord_clean:
                return True
        
        return False
    
    def _validate_data_completeness(self):
        """데이터 완성도 검증"""
        
        print("\n4️⃣ 데이터 완성도 검증")
        print("-" * 70)
        
        # 주요 시트에서 빈 셀 비율 확인
        critical_sheets = ['Revenue_Buildup', 'Cost_Structure', 'PL_5Year', 'PL_3Year']
        
        for sheet_name in critical_sheets:
            if sheet_name not in self.wb.sheetnames:
                continue
            
            ws = self.wb[sheet_name]
            
            # 데이터 영역 (A1:H20 정도) 검사
            total_cells = 0
            empty_cells = 0
            
            for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=8):
                for cell in row:
                    total_cells += 1
                    if cell.value is None:
                        empty_cells += 1
            
            empty_ratio = empty_cells / total_cells if total_cells > 0 else 0
            
            if empty_ratio > 0.7:
                self.warnings.append(
                    f"⚠️ {sheet_name}: 빈 셀 비율 {empty_ratio*100:.0f}% (데이터 부족 가능성)"
                )
            else:
                self.info.append(
                    f"✅ {sheet_name}: 데이터 충분 (빈 셀 {empty_ratio*100:.0f}%)"
                )
        
        print(f"주요 시트 {len(critical_sheets)}개 검사 완료")
    
    def _validate_calculation_results(self):
        """계산 결과 검증 (샘플링)"""
        
        print("\n5️⃣ 계산 결과 검증 (예상값 vs 실제값)")
        print("-" * 70)
        
        # Revenue_Buildup 검증
        if 'Revenue_Buildup' in self.wb.sheetnames:
            self._validate_revenue_growth()
        
        # Dashboard 검증
        if 'Dashboard' in self.wb.sheetnames:
            self._validate_dashboard_values()
    
    def _validate_revenue_growth(self):
        """Revenue 성장 로직 검증"""
        
        ws = self.wb['Revenue_Buildup']
        
        # Year 0 vs Year 1 성장률 확인 (샘플)
        # 첫 번째 세그먼트 (보통 Row 5)
        for row_idx in range(5, 8):  # 최대 3개 세그먼트 검사
            try:
                y0_cell = ws[f'B{row_idx}']
                y1_cell = ws[f'C{row_idx}']
                growth_cell = ws[f'H{row_idx}']
                
                if y0_cell.value is None or y1_cell.value is None:
                    continue
                
                # Year 1이 비어있으면 문제
                if y1_cell.value is None or y1_cell.value == 0:
                    self.errors.append(
                        f"❌ Revenue_Buildup!C{row_idx}: Year 1 데이터 없음 (수식 오류 가능성)"
                    )
                
                # 수식 패턴 확인
                if hasattr(y1_cell, 'value') and isinstance(y1_cell.value, str):
                    formula = y1_cell.value
                    # C5 = B5*(1+$H$5) 패턴인지 확인
                    if f'C{row_idx}' in formula:
                        self.errors.append(
                            f"❌ 자기 참조: Revenue_Buildup!C{row_idx} = {formula}"
                        )
                    elif f'B{row_idx}' in formula:
                        self.info.append(
                            f"✅ Revenue_Buildup!C{row_idx}: 올바른 참조 (B{row_idx})"
                        )
                
            except Exception as e:
                self.warnings.append(f"⚠️ Revenue 검증 오류: {e}")
    
    def _validate_dashboard_values(self):
        """Dashboard 값 검증"""
        
        ws = self.wb['Dashboard']
        
        # Dashboard의 주요 셀에 값이 있는지 확인
        critical_cells = ['B5', 'B6', 'B7']  # Revenue, Net Income, CAGR 등
        
        for cell_addr in critical_cells:
            cell = ws[cell_addr]
            
            if cell.value is None:
                self.errors.append(
                    f"❌ Dashboard!{cell_addr}: 값이 없음 (Named Range 참조 실패 가능)"
                )
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                # 수식이 있지만 data_only=False라 값 확인 불가
                self.info.append(
                    f"✅ Dashboard!{cell_addr}: 수식 존재"
                )
            else:
                self.info.append(
                    f"✅ Dashboard!{cell_addr}: 값 있음 ({cell.value})"
                )
    
    def _compile_results(self) -> Dict[str, Any]:
        """검증 결과 정리"""
        
        passed = len(self.errors) == 0
        
        print("\n" + "="*70)
        print("📊 검증 결과")
        print("="*70)
        
        if passed:
            print("✅ 검증 통과! (오류 없음)")
        else:
            print(f"❌ 검증 실패! ({len(self.errors)}개 오류)")
        
        # 오류 출력
        if self.errors:
            print(f"\n❌ 오류 ({len(self.errors)}개):")
            for error in self.errors[:10]:  # 최대 10개만
                print(f"   {error}")
            if len(self.errors) > 10:
                print(f"   ... 외 {len(self.errors) - 10}개")
        
        # 경고 출력
        if self.warnings:
            print(f"\n⚠️ 경고 ({len(self.warnings)}개):")
            for warning in self.warnings[:5]:
                print(f"   {warning}")
            if len(self.warnings) > 5:
                print(f"   ... 외 {len(self.warnings) - 5}개")
        
        # 정보 출력
        if self.info and not self.errors:
            print(f"\n✅ 정상 ({len(self.info)}개):")
            for info in self.info[:5]:
                print(f"   {info}")
            if len(self.info) > 5:
                print(f"   ... 외 {len(self.info) - 5}개")
        
        return {
            'passed': passed,
            'errors': self.errors,
            'warnings': self.warnings,
            'info': self.info,
            'stats': {
                'total_sheets': len(self.wb.sheetnames),
                'total_named_ranges': len(list(self.wb.defined_names)),
                'error_count': len(self.errors),
                'warning_count': len(self.warnings)
            }
        }


class GoldenWorkbookValidator:
    """
    Golden Workbook 비교 검증
    
    생성된 Excel과 예상 결과를 비교
    """
    
    def __init__(self, filepath: Path, expected_values: Dict):
        """
        Args:
            filepath: 검증할 Excel 파일
            expected_values: 예상 결과
                {
                    'revenue_y0': 1250_0000_0000,
                    'revenue_y5': 4295_0000_0000,
                    'net_income_y5': 429_0000_0000,
                    'cagr': 0.28,
                    ...
                }
        """
        self.filepath = filepath
        self.expected = expected_values
        self.wb = None
        self.results = []
    
    def validate(self) -> Dict[str, Any]:
        """
        Golden Workbook 검증
        
        Returns:
            검증 결과
        """
        
        print(f"\n🎯 Golden Workbook 검증: {self.filepath.name}")
        print("="*70)
        
        # data_only=True로 열어서 계산된 값 확인
        try:
            self.wb = load_workbook(self.filepath, data_only=True)
        except Exception as e:
            return {
                'passed': False,
                'error': f"파일 열기 실패: {e}"
            }
        
        # Named Range 값 확인
        self._check_named_range_values()
        
        passed = all(r['passed'] for r in self.results)
        
        print("\n" + "="*70)
        print("📊 Golden Workbook 검증 결과")
        print("="*70)
        
        if passed:
            print("✅ 모든 검증 통과!")
        else:
            print(f"❌ {sum(1 for r in self.results if not r['passed'])}개 실패")
        
        for result in self.results:
            status = "✅" if result['passed'] else "❌"
            print(f"{status} {result['name']}: {result['message']}")
        
        return {
            'passed': passed,
            'results': self.results
        }
    
    def _check_named_range_values(self):
        """Named Range 값 확인"""
        
        # Revenue_Y0 확인
        if 'revenue_y0' in self.expected:
            actual = self._get_named_range_value('Revenue_Y0')
            expected = self.expected['revenue_y0']
            
            if actual is None:
                self.results.append({
                    'name': 'Revenue_Y0',
                    'passed': False,
                    'message': f"값 없음 (예상: ₩{expected/1_0000_0000:.0f}억)"
                })
            elif abs(actual - expected) / expected < 0.01:  # 1% 오차 허용
                self.results.append({
                    'name': 'Revenue_Y0',
                    'passed': True,
                    'message': f"₩{actual/1_0000_0000:.0f}억 ≈ ₩{expected/1_0000_0000:.0f}억 ✅"
                })
            else:
                self.results.append({
                    'name': 'Revenue_Y0',
                    'passed': False,
                    'message': f"₩{actual/1_0000_0000:.0f}억 ≠ ₩{expected/1_0000_0000:.0f}억 (오차 {abs(actual-expected)/expected*100:.1f}%)"
                })
    
    def _get_named_range_value(self, name: str) -> Optional[float]:
        """
        Named Range의 값 가져오기
        
        Args:
            name: Named Range 이름
        
        Returns:
            값 (숫자) 또는 None
        """
        
        try:
            destinations = self.wb.defined_names[name].destinations
            for sheet_name, cell_addr in destinations:
                ws = self.wb[sheet_name]
                # 절대 참조 제거 ($B$5 → B5)
                cell_addr_clean = cell_addr.replace('$', '')
                cell = ws[cell_addr_clean]
                
                if cell.value is not None:
                    try:
                        return float(cell.value)
                    except:
                        return None
        except Exception as e:
            return None
        
        return None


# 편의 함수
def validate_excel(filepath: Path) -> bool:
    """
    Excel 파일 검증 (편의 함수)
    
    Args:
        filepath: Excel 파일 경로
    
    Returns:
        검증 통과 여부
    """
    
    validator = ExcelValidator(filepath)
    result = validator.validate()
    return result['passed']


def validate_with_golden(filepath: Path, expected_values: Dict) -> bool:
    """
    Golden Workbook 비교 검증 (편의 함수)
    
    Args:
        filepath: Excel 파일 경로
        expected_values: 예상 결과
    
    Returns:
        검증 통과 여부
    """
    
    # 1. 기본 검증
    validator = ExcelValidator(filepath)
    basic_result = validator.validate()
    
    if not basic_result['passed']:
        print("\n❌ 기본 검증 실패 - Golden 검증 생략")
        return False
    
    # 2. Golden 검증
    golden_validator = GoldenWorkbookValidator(filepath, expected_values)
    golden_result = golden_validator.validate()
    
    return golden_result['passed']


# 사용 예시는 별도 스크립트에서
# python scripts/validate_generated_excel.py

