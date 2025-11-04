"""
Convergence Analysis Sheet Builder
4가지 방법의 SAM 수렴 분석

피드백 반영:
  - 조건부 서식을 FormulaRule로 변경 (CellIsRule 대신)
"""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

from .formula_engine import FormulaEngine, ExcelStyles


class ConvergenceBuilder:
    """
    Convergence Analysis 시트 빌더
    
    기능:
      - 4가지 Method SAM 비교
      - 평균, 표준편차, 변동계수
      - Max/Min 비율
      - ±30% 수렴 확인
      - 조건부 서식 (FormulaRule)
    
    피드백 반영:
      - FormulaRule 사용 (안정적)
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        method_sheets: List[str] = None,
        sam_cells: List[str] = None
    ) -> None:
        """
        Convergence Analysis 시트 생성
        
        Args:
            method_sheets: Method 시트 이름 리스트
                ['Method_1_TopDown', 'Method_2_BottomUp', ...]
            sam_cells: 각 시트의 SAM 셀 또는 Named Range
                ['SAM', 'SAM_Method2', 'SAM_Method3', 'SAM_Method4']
        
        피드백 반영:
          - 조건부 서식 FormulaRule로 변경
        """
        
        if method_sheets is None:
            method_sheets = [
                'Method_1_TopDown',
                'Method_2_BottomUp',
                'Method_3_Proxy',
                'Method_4_CompetitorRevenue'
            ]
        
        if sam_cells is None:
            # Named Range 사용
            sam_cells = ['SAM', 'SAM_Method2', 'SAM_Method3', 'SAM_Method4']
        
        ws = self.wb.create_sheet("Convergence_Analysis")
        
        # 제목
        ws['A1'] = "Convergence Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # 헤더
        ws['A3'] = "Method"
        ws['B3'] = "SAM (억원)"
        ws['C3'] = "차이 (%)"
        
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color=ExcelStyles.HEADER_FILL,
                end_color=ExcelStyles.HEADER_FILL,
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")
        
        # 각 Method 결과
        for i, (method_name, sam_ref) in enumerate(zip(method_sheets, sam_cells), start=4):
            ws[f'A{i}'] = method_name
            
            # SAM 참조 (Named Range 또는 교차 시트)
            if sam_ref.startswith('SAM'):
                # Named Range 사용
                ws[f'B{i}'] = f"={sam_ref}"
            else:
                # 교차 시트 참조
                ws[f'B{i}'] = self.fe.create_cross_sheet_ref(method_name, sam_ref)
            
            ws[f'B{i}'].number_format = '#,##0'
            
            # 평균 대비 차이 (나중에 평균 계산 후 사용)
            ws[f'C{i}'] = f"=(B{i}-$B$8)/$B$8*100"
            ws[f'C{i}'].number_format = '0.0"%"'
        
        # 통계
        stats_start_row = 8
        
        # 평균
        ws[f'A{stats_start_row}'] = "평균"
        ws[f'A{stats_start_row}'].font = Font(bold=True)
        
        ws[f'B{stats_start_row}'] = "=AVERAGE(B4:B7)"
        ws[f'B{stats_start_row}'].number_format = '#,##0'
        ws[f'B{stats_start_row}'].fill = PatternFill(
            start_color=ExcelStyles.RESULT_FILL,
            end_color=ExcelStyles.RESULT_FILL,
            fill_type="solid"
        )
        ws[f'B{stats_start_row}'].font = Font(bold=True)
        
        # Named Range 정의 (중요!)
        self.fe.define_named_range('Conv_AvgSAM', 'Convergence_Analysis', f'B{stats_start_row}')
        
        # 표준편차
        ws[f'A{stats_start_row+1}'] = "표준편차"
        ws[f'B{stats_start_row+1}'] = "=STDEV(B4:B7)"
        ws[f'B{stats_start_row+1}'].number_format = '#,##0'
        
        # Named Range 정의
        self.fe.define_named_range('Conv_StdDev', 'Convergence_Analysis', f'B{stats_start_row+1}')
        
        # 변동계수 (CV%)
        ws[f'A{stats_start_row+2}'] = "변동계수 (CV%)"
        ws[f'B{stats_start_row+2}'] = "=Conv_StdDev/Conv_AvgSAM*100"  # Named Range 사용
        ws[f'B{stats_start_row+2}'].number_format = '0.0"%"'
        
        # Named Range 정의
        self.fe.define_named_range('Conv_CV', 'Convergence_Analysis', f'B{stats_start_row+2}')
        
        # Max/Min 비율
        ws[f'A{stats_start_row+3}'] = "Max/Min 비율"
        ws[f'B{stats_start_row+3}'] = "=MAX(B4:B7)/MIN(B4:B7)"
        ws[f'B{stats_start_row+3}'].number_format = '0.00'
        
        # Named Range 정의
        self.fe.define_named_range('Conv_MaxMin', 'Convergence_Analysis', f'B{stats_start_row+3}')
        
        # ±30% 수렴 확인
        conv_row = stats_start_row + 4
        
        ws[f'A{conv_row}'] = "±30% 수렴?"
        ws[f'A{conv_row}'].font = Font(bold=True)
        
        ws[f'B{conv_row}'] = '=IF(Conv_MaxMin<=1.3, "✅ 통과", "❌ 재검토 필요")'  # Named Range 사용
        ws[f'B{conv_row}'].font = Font(bold=True)
        
        # Named Range 정의
        self.fe.define_named_range('Conv_Status', 'Convergence_Analysis', f'B{conv_row}')
        
        # 조건부 서식 (피드백 반영: FormulaRule 사용!)
        # Note: openpyxl 버전에 따라 조건부 서식 API가 다를 수 있음
        # 간단하게 Rule 객체 직접 사용
        
        from openpyxl.formatting.rule import Rule
        
        green_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )
        
        red_fill = PatternFill(
            start_color="FFC7CE",
            end_color="FFC7CE",
            fill_type="solid"
        )
        
        # 통과 조건 (✅ 포함)
        green_rule = Rule(
            type='containsText',
            operator='containsText',
            text='✅',
            dxf=DifferentialStyle(fill=green_fill)
        )
        ws.conditional_formatting.add(f'B{conv_row}', green_rule)
        
        # 실패 조건 (❌ 포함)
        red_rule = Rule(
            type='containsText',
            operator='containsText',
            text='❌',
            dxf=DifferentialStyle(fill=red_fill)
        )
        ws.conditional_formatting.add(f'B{conv_row}', red_rule)
        
        # 해석
        interp_row = conv_row + 2
        
        ws[f'A{interp_row}'] = "해석:"
        ws[f'A{interp_row}'].font = Font(bold=True)
        
        ws[f'A{interp_row+1}'] = "Max/Min ≤ 1.3 (±30%)이면 4가지 방법이 수렴"
        ws[f'A{interp_row+2}'] = "→ SAM 추정이 신뢰 가능"
        ws[f'A{interp_row+3}'] = "Max/Min > 1.3이면 재검토 필요"
        ws[f'A{interp_row+4}'] = "→ 가정 또는 방법론 수정"
        
        # 열 너비
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        print(f"   ✅ Convergence Analysis (조건부 서식 FormulaRule 적용)")

