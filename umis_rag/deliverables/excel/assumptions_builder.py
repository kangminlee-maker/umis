"""
Assumptions Sheet Builder
가정 시트 생성기

시트 구조:
  - 헤더: ID, Category, Description, Value, Unit, Data_Type, Source, Confidence, Notes
  - 데이터 행: 각 가정
  - Named Range: 각 가정의 Value 셀
  - 서식: 입력 셀 강조, 추정치 코멘트
"""

from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.comments import Comment

from .formula_engine import FormulaEngine, ExcelStyles


class AssumptionsSheetBuilder:
    """
    Assumptions 시트 빌더
    
    기능:
      - 가정 목록 → Excel 시트
      - Named Range 자동 정의 (절대참조)
      - 입력 셀 서식
      - 추정치 코멘트
      - 시트 보호 (Value만 편집 가능)
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, assumptions: List[Dict]) -> None:
        """
        Assumptions 시트 생성
        
        Args:
            assumptions: 가정 목록
                [
                    {
                        'id': 'ASM_001',
                        'category': '인구',
                        'description': '타겟 고객 수',
                        'value': 10000,
                        'unit': '명',
                        'data_type': '직접데이터',  # or '추정치'
                        'source': 'SRC_20241031_001',
                        'confidence': 'High',
                        'notes': '통계청 공식'
                    },
                    ...
                ]
        """
        
        # 1. 시트 생성 (첫 번째 시트로)
        ws = self.wb.create_sheet("Assumptions", 0)
        
        # 2. 헤더 작성
        headers = [
            "ID", "Category", "Description", "Value",
            "Unit", "Data_Type", "Source", "Confidence", "Notes"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=ExcelStyles.HEADER_FILL,
                end_color=ExcelStyles.HEADER_FILL,
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 12  # ID
        ws.column_dimensions['B'].width = 15  # Category
        ws.column_dimensions['C'].width = 40  # Description
        ws.column_dimensions['D'].width = 15  # Value
        ws.column_dimensions['E'].width = 10  # Unit
        ws.column_dimensions['F'].width = 15  # Data_Type
        ws.column_dimensions['G'].width = 20  # Source
        ws.column_dimensions['H'].width = 12  # Confidence
        ws.column_dimensions['I'].width = 40  # Notes
        
        # 3. 데이터 행 작성
        for i, asm in enumerate(assumptions, start=2):
            ws.cell(row=i, column=1).value = asm.get('id')
            ws.cell(row=i, column=2).value = asm.get('category')
            ws.cell(row=i, column=3).value = asm.get('description')
            ws.cell(row=i, column=4).value = asm.get('value')
            ws.cell(row=i, column=5).value = asm.get('unit')
            ws.cell(row=i, column=6).value = asm.get('data_type')
            ws.cell(row=i, column=7).value = asm.get('source')
            ws.cell(row=i, column=8).value = asm.get('confidence')
            ws.cell(row=i, column=9).value = asm.get('notes', '')
            
            # Value 셀에 Named Range 정의 (절대참조!)
            asm_id = asm.get('id')
            if asm_id:
                self.fe.define_named_range(
                    name=asm_id,
                    sheet="Assumptions",
                    cell=f"D{i}",  # Value 컬럼
                    scope='workbook'  # Workbook-scope
                )
            
            # Value 셀 서식 (입력 셀)
            value_cell = ws.cell(row=i, column=4)
            value_cell.fill = PatternFill(
                start_color=ExcelStyles.INPUT_FILL,
                end_color=ExcelStyles.INPUT_FILL,
                fill_type="solid"
            )
            value_cell.number_format = '#,##0'
            value_cell.alignment = Alignment(horizontal="right")
            
            # 추정치인 경우 코멘트 추가
            if asm.get('data_type') == '추정치':
                comment_text = (
                    f"추정 논리:\n"
                    f"출처: {asm.get('source', 'N/A')}\n"
                    f"신뢰도: {asm.get('confidence', 'N/A')}\n\n"
                    f"Estimation_Details 시트 참조"
                )
                
                comment = Comment(comment_text, "Bill (Quantifier)")
                ws.cell(row=i, column=1).comment = comment
        
        # 4. 시트 보호 (선택적)
        # Note: 시트 보호는 사용자가 Excel에서 직접 설정 가능
        # ws.protection.sheet = True
        
        # Value 셀만 편집 가능하도록 설정 (시트 보호 시)
        # for row_idx in range(2, len(assumptions) + 2):
        #     ws.cell(row=row_idx, column=4).protection = Protection(locked=False)
        
        # 5. 상단 고정 (헤더 항상 보이게)
        ws.freeze_panes = 'A2'
        
        print(f"   ✅ Assumptions: {len(assumptions)}개 가정, {len(assumptions)}개 Named Range")


class EstimationDetailsBuilder:
    """
    Estimation Details 시트 빌더 (v7.2.0)
    
    YAML Spec 7개 섹션 완전 구현:
      1. 추정 필요 이유
      2. 사용한 데이터
      3. 추정 논리 (단계별)
      4. 신뢰도 평가
      5. 검증 방법
      6. 대체 접근법
      7. 사용 위치
    
    + Named Range 적용
    """
    
    def __init__(self, workbook: Workbook, formula_engine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, estimations: List[Dict]) -> None:
        """
        Estimation Details 시트 생성
        
        Args:
            estimations: 추정치 목록
                [
                    {
                        'id': 'FILTER_PIANO',
                        'description': '피아노 비중',
                        'value': 0.25,
                        'reason': '왜 추정이 필요한가',
                        'base_data': [{'name': '전체 악기 판매', 'value': 1000, 'source': 'SRC_xxx'}],
                        'logic_steps': ['Step 1: ...', 'Step 2: ...'],
                        'calculation': '1000 * 0.25',
                        'confidence': 'Medium',
                        'error_range': '±10%',
                        'verification': '상한/하한 테스트',
                        'alternatives': ['방법1: ...', '방법2: ...'],
                        'used_in': ['ASM_001', 'Method_1']
                    },
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("Estimation_Details")
        
        # === 제목 ===
        ws['A1'] = "Estimation Details (추정치 상세 문서화)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        ws['A2'] = "YAML Spec 7개 섹션 기반 - 모든 추정의 투명성 확보"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:F2')
        
        # 열 너비
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        # === 헤더 ===
        row = 4
        headers = ['EST_ID', 'Item', 'Final_Value', 'Used_In', 'Confidence', 'Error_Range']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color=ExcelStyles.HEADER_FILL,
                end_color=ExcelStyles.HEADER_FILL,
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal='center')
        
        # === 각 추정치 블록 ===
        current_row = row + 1
        
        for est in estimations:
            # 메인 정보
            est_id = est.get('id')
            
            ws.cell(row=current_row, column=1).value = est_id
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).fill = PatternFill(
                start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"
            )
            
            ws.cell(row=current_row, column=2).value = est.get('description')
            ws.cell(row=current_row, column=2).font = Font(bold=True)
            
            # Named Range 정의 (추정값)
            ws.cell(row=current_row, column=3).value = est.get('value')
            if isinstance(est.get('value'), (int, float)):
                if est.get('value') < 1:
                    ws.cell(row=current_row, column=3).number_format = '0.0%'
                else:
                    ws.cell(row=current_row, column=3).number_format = '#,##0'
            
            # Named Range 생성 (이미 존재하면 건너뜀)
            if est_id not in self.fe.named_ranges:
                self.fe.define_named_range(est_id, 'Estimation_Details', f'C{current_row}')
            
            ws.cell(row=current_row, column=4).value = est.get('used_in', '')
            ws.cell(row=current_row, column=4).font = Font(size=9)
            
            ws.cell(row=current_row, column=5).value = est.get('confidence', 'Medium')
            ws.cell(row=current_row, column=6).value = est.get('error_range', '±20%')
            
            current_row += 1
            
            # === 7개 섹션 시작 ===
            # [1] 추정 필요 이유
            ws.cell(row=current_row, column=1).value = "[1] 추정 이유"
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True, color="0066CC")
            ws.cell(row=current_row, column=2).value = est.get('reason', '직접 데이터 없음')
            ws.cell(row=current_row, column=2).font = Font(size=9)
            ws.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
            
            # [2] 사용한 데이터
            ws.cell(row=current_row, column=1).value = "[2] Base Data"
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True, color="0066CC")
            
            base_data = est.get('base_data', '')
            if isinstance(base_data, list):
                base_data_str = '; '.join([f"{d.get('name')}: {d.get('value')}" for d in base_data])
            else:
                base_data_str = base_data
            
            ws.cell(row=current_row, column=2).value = base_data_str
            ws.cell(row=current_row, column=2).font = Font(size=9)
            ws.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
            
            # [3] 추정 논리
            ws.cell(row=current_row, column=1).value = "[3] 추정 논리"
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True, color="0066CC")
            
            logic_steps = est.get('logic_steps', [])
            if logic_steps:
                logic_str = ' → '.join(logic_steps)
            else:
                logic_str = est.get('calculation', '계산식 없음')
            
            ws.cell(row=current_row, column=2).value = logic_str
            ws.cell(row=current_row, column=2).font = Font(size=9)
            ws.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
            
            # [4] 신뢰도 평가 (이미 위에 표시되어 있으므로 스킵 가능)
            
            # [5] 검증 방법
            ws.cell(row=current_row, column=1).value = "[5] 검증"
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True, color="0066CC")
            ws.cell(row=current_row, column=2).value = est.get('verification', '범위 체크 필요')
            ws.cell(row=current_row, column=2).font = Font(size=9)
            ws.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
            
            # [6] 대체 접근법
            ws.cell(row=current_row, column=1).value = "[6] 대체 방법"
            ws.cell(row=current_row, column=1).font = Font(size=9, bold=True, color="0066CC")
            
            alternatives = est.get('alternatives', [])
            if alternatives:
                alt_str = '; '.join(alternatives)
            else:
                alt_str = '현재 방법이 최선'
            
            ws.cell(row=current_row, column=2).value = alt_str
            ws.cell(row=current_row, column=2).font = Font(size=9)
            ws.merge_cells(f'B{current_row}:F{current_row}')
            current_row += 1
            
            # 구분선
            current_row += 1
        
        # 상단 고정
        ws.freeze_panes = 'A5'
        
        print(f"   ✅ Estimation Details: {len(estimations)}개 추정치")
        print(f"      - 7개 섹션 기반 상세 문서화")
        print(f"      - {len(estimations)}개 Named Range 생성")

