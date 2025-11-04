"""
Market Sizing Workbook Generator
Bill의 market_sizing.xlsx 자동 생성 (피드백 반영)

9개 시트:
  1. Assumptions
  2-5. Method_1_TopDown ~ Method_4_CompetitorRevenue
  6. Convergence_Analysis
  7-9. Scenarios, Validation_Log, Summary

피드백 반영:
  - fullCalcOnLoad=True 설정
  - Named Range Workbook-scope
  - 절대참조 사용
"""

from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from .formula_engine import FormulaEngine, ExcelStyles


class MarketSizingWorkbookGenerator:
    """
    Market Sizing Excel 자동 생성기
    
    피드백 반영된 개선사항:
      - Named Range 절대참조
      - fullCalcOnLoad=True
      - 검증 강화
    """
    
    def __init__(self):
        """초기화"""
        self.formula_engine: Optional[FormulaEngine] = None
    
    def generate(
        self,
        market_name: str,
        assumptions: List[Dict],
        tam: Dict,
        segments: List[Dict],
        proxy_data: Dict,
        competitors: List[Dict],
        output_dir: Path
    ) -> Path:
        """
        전체 워크북 생성
        
        Args:
            market_name: 시장 이름
            assumptions: 가정 목록
            tam: TAM 정의
            segments: 세그먼트 목록 (Bottom-Up용)
            proxy_data: Proxy 데이터
            competitors: 경쟁사 목록
            output_dir: 출력 디렉토리
        
        Returns:
            생성된 Excel 파일 경로
        
        피드백 반영:
          - fullCalcOnLoad=True 설정
          - Named Range 절대참조
        """
        
        print(f"🚀 Market Sizing Workbook 생성 시작")
        print(f"   시장: {market_name}")
        
        # 1. 워크북 초기화
        wb = Workbook()
        self.formula_engine = FormulaEngine(wb)
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 2. Assumptions 시트
        print(f"   1/9 Assumptions...")
        self._create_assumptions_sheet(wb, assumptions)
        
        # 3. Estimation Details (추정치가 있는 경우)
        estimations = [a for a in assumptions if a.get('data_type') == '추정치']
        if estimations:
            print(f"   2/9 Estimation Details...")
            self._create_estimation_details(wb, estimations)
        
        # 4-7. Method 시트들 (4가지)
        print(f"   3/9 Method 1: Top-Down...")
        self._create_method1_topdown(wb, tam)
        
        print(f"   4/9 Method 2: Bottom-Up...")
        self._create_method2_bottomup(wb, segments)
        
        print(f"   5/9 Method 3: Proxy...")
        self._create_method3_proxy(wb, proxy_data)
        
        print(f"   6/9 Method 4: Competitor Revenue...")
        self._create_method4_competitor(wb, competitors)
        
        # 8. Convergence Analysis
        print(f"   7/9 Convergence Analysis...")
        self._create_convergence_analysis(wb)
        
        # 9. Scenarios
        print(f"   8/9 Scenarios...")
        self._create_scenarios(wb)
        
        # 10. Validation Log
        print(f"   9/9 Validation Log...")
        self._create_validation_log(wb)
        
        # 11. 강제 재계산 설정 (피드백 반영!)
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True  # ⭐ 피드백 반영!
        
        # 12. 저장
        filename = f"market_sizing_{market_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = output_dir / filename
        
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        print(f"\n✅ Excel 생성 완료: {filepath}")
        print(f"📋 다음: Excel에서 열어서 함수 작동 확인")
        print(f"📋 다음: PDF로 저장 (백업)")
        
        return filepath
    
    def _create_assumptions_sheet(self, wb: Workbook, assumptions: List[Dict]):
        """Assumptions 시트 생성 (구현 예정)"""
        ws = wb.create_sheet("Assumptions", 0)
        ws['A1'] = "가정 시트 (구현 예정)"
    
    def _create_estimation_details(self, wb: Workbook, estimations: List[Dict]):
        """Estimation Details 시트"""
        ws = wb.create_sheet("Estimation_Details")
        ws['A1'] = "추정 상세 (구현 예정)"
    
    def _create_method1_topdown(self, wb: Workbook, tam: Dict):
        """Method 1: Top-Down"""
        ws = wb.create_sheet("Method_1_TopDown")
        ws['A1'] = "Top-Down 방법 (구현 예정)"
    
    def _create_method2_bottomup(self, wb: Workbook, segments: List[Dict]):
        """Method 2: Bottom-Up"""
        ws = wb.create_sheet("Method_2_BottomUp")
        ws['A1'] = "Bottom-Up 방법 (구현 예정)"
    
    def _create_method3_proxy(self, wb: Workbook, proxy_data: Dict):
        """Method 3: Proxy"""
        ws = wb.create_sheet("Method_3_Proxy")
        ws['A1'] = "Proxy 방법 (구현 예정)"
    
    def _create_method4_competitor(self, wb: Workbook, competitors: List[Dict]):
        """Method 4: Competitor Revenue"""
        ws = wb.create_sheet("Method_4_CompetitorRevenue")
        ws['A1'] = "경쟁사 역산 (구현 예정)"
    
    def _create_convergence_analysis(self, wb: Workbook):
        """Convergence Analysis"""
        ws = wb.create_sheet("Convergence_Analysis")
        ws['A1'] = "수렴 분석 (구현 예정)"
    
    def _create_scenarios(self, wb: Workbook):
        """Scenarios"""
        ws = wb.create_sheet("Scenarios")
        ws['A1'] = "시나리오 (구현 예정)"
    
    def _create_validation_log(self, wb: Workbook):
        """Validation Log"""
        ws = wb.create_sheet("Validation_Log")
        ws['A1'] = "검증 로그 (구현 예정)"


# 테스트는 별도 스크립트에서
# python scripts/test_excel_generation.py

