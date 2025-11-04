"""
Formula Reference Validator
수식 내 참조가 의도한 셀을 참조하는지 검증

핵심 원칙:
1. 수식 파싱: =Scenarios!B13에서 "Scenarios!B13" 추출
2. 참조 셀 내용 확인: B13에 뭐가 있나?
3. 의미 검증: B13이 "Proxy Corr"인데 "Average SAM"을 원했다면 오류!
4. 의도 매칭: 셀 내용 vs 수식의 의도

Example:
  Summary!B23 = "Best Case Average SAM"
  수식: =Scenarios!B13
  B13 내용: "Proxy Corr" (0.3)
  의도: "Average SAM" (₩120억)
  판정: ❌ 잘못된 참조! B13은 Proxy Corr이지 Average SAM이 아님
"""

import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class FormulaReferenceValidator:
    """
    수식 참조 검증기
    
    기능:
      - 수식에서 참조 추출
      - 참조 셀의 실제 내용 확인
      - 의도와 비교
    """
    
    def __init__(self, filepath: Path):
        """
        Args:
            filepath: 검증할 Excel 파일
        """
        self.filepath = filepath
        self.wb = None
        self.errors = []
        self.warnings = []
    
    def validate(self) -> Dict:
        """전체 검증 실행"""
        
        print(f"\n🔍 수식 참조 검증: {self.filepath.name}")
        print("="*70)
        
        # 파일 열기
        try:
            self.wb = load_workbook(self.filepath, data_only=False)
        except Exception as e:
            self.errors.append(f"파일 열기 실패: {e}")
            return self._compile_results()
        
        # 주요 시트별 검증
        self._validate_summary_references()
        self._validate_convergence_references()
        
        return self._compile_results()
    
    def _validate_summary_references(self):
        """Summary 시트의 참조 검증 (핵심!)"""
        
        print("\n1️⃣ Summary 시트 참조 검증")
        print("-"*70)
        
        if 'Summary' not in self.wb.sheetnames:
            self.warnings.append("Summary 시트 없음")
            return
        
        ws = self.wb['Summary']
        
        # 주요 셀 검증
        critical_cells = [
            {
                'cell': 'B5',
                'expected_name': 'TAM',
                'expected_source': 'TAM (Convergence 또는 Named Range)',
                'description': 'Summary TAM'
            },
            {
                'cell': 'B6',
                'expected_name': 'SAM',
                'expected_source': 'Convergence 평균 SAM',
                'description': 'Summary SAM (평균)'
            },
            # Method별 SAM (B10-B13)
            {
                'cell': 'B10',
                'expected_name': 'SAM (Method 1)',
                'expected_source': 'SAM Named Range',
                'description': 'Summary Method 1'
            },
            {
                'cell': 'B11',
                'expected_name': 'SAM (Method 2)',
                'expected_source': 'SAM_Method2',
                'description': 'Summary Method 2'
            },
            # Scenarios
            {
                'cell': 'B21',  # 대략 이 위치
                'expected_name': 'Best Case Average SAM',
                'expected_source': 'Scenarios Average SAM (Best)',
                'description': 'Best Case SAM'
            },
        ]
        
        for spec in critical_cells:
            cell_addr = spec['cell']
            cell = ws[cell_addr]
            
            if cell.value is None:
                continue
            
            if isinstance(cell.value, str) and cell.value.startswith('='):
                # 수식에서 참조 추출
                formula = cell.value
                refs = self._extract_references(formula)
                
                print(f"\n{cell_addr} ({spec['description']}):")
                print(f"  수식: {formula}")
                print(f"  참조: {refs}")
                
                # 각 참조의 내용 확인
                for ref in refs:
                    ref_content = self._get_reference_content(ref)
                    print(f"  → {ref}: {ref_content}")
                    
                    # 의도 검증
                    if spec['expected_name'] in ref or spec['expected_source'] in str(ref_content):
                        print(f"     ✅ 의도와 일치")
                    else:
                        # 잘못된 참조 의심
                        if ref_content and isinstance(ref_content, (int, float)):
                            # 숫자만 있으면 패스 (계산 결과)
                            continue
                        elif ref_content and isinstance(ref_content, str):
                            # 문자열이면 확인
                            if spec['expected_name'].lower() not in str(ref_content).lower():
                                self.errors.append(
                                    f"❌ {cell_addr}: {ref}는 '{ref_content}'인데, "
                                    f"'{spec['expected_name']}'을(를) 원함"
                                )
                                print(f"     ❌ 잘못된 참조! '{ref_content}'는 '{spec['expected_name']}'이 아님")
    
    def _validate_convergence_references(self):
        """Convergence 시트의 참조 검증"""
        
        print("\n2️⃣ Convergence 시트 참조 검증")
        print("-"*70)
        
        if 'Convergence_Analysis' not in self.wb.sheetnames:
            self.warnings.append("Convergence_Analysis 시트 없음")
            return
        
        ws = self.wb['Convergence_Analysis']
        
        # Method별 SAM이 올바른 Named Range 참조하는지
        sam_cells = [
            ('B4', 'SAM', 'Method 1 SAM'),
            ('B5', 'SAM_Method2', 'Method 2 SAM'),
            ('B6', 'SAM_Method3', 'Method 3 SAM'),
            ('B7', 'SAM_Method4', 'Method 4 SAM'),
        ]
        
        for cell_addr, expected_range, desc in sam_cells:
            cell = ws[cell_addr]
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                
                print(f"\n{cell_addr} ({desc}):")
                print(f"  수식: {formula}")
                
                if expected_range in formula:
                    print(f"  ✅ {expected_range} 참조 정상")
                else:
                    self.errors.append(
                        f"❌ {cell_addr}: {expected_range} 참조 없음 (수식: {formula})"
                    )
                    print(f"  ❌ {expected_range} 참조 없음!")
    
    def _extract_references(self, formula: str) -> List[str]:
        """
        수식에서 셀/Range 참조 추출
        
        Args:
            formula: Excel 수식
        
        Returns:
            참조 목록
        
        Examples:
            "=Scenarios!B13" → ["Scenarios!B13"]
            "=A5*B5" → ["A5", "B5"]
            "=SUM(B4:B7)" → ["B4:B7"]
        """
        
        refs = []
        
        # 시트 참조 (Sheet!Cell)
        sheet_refs = re.findall(r'([A-Za-z_]+)!\$?([A-Z]+)\$?(\d+)', formula)
        for sheet, col, row in sheet_refs:
            refs.append(f"{sheet}!{col}{row}")
        
        # 일반 셀 참조 (A1, $A$1 등)
        cell_refs = re.findall(r'(?<![A-Za-z_])\$?([A-Z]+)\$?(\d+)(?![A-Z])', formula)
        for col, row in cell_refs:
            if f"{col}{row}" not in ''.join(refs):  # 중복 제거
                refs.append(f"{col}{row}")
        
        # Named Range
        named_refs = re.findall(r'(?<==)([A-Za-z_][A-Za-z0-9_]*)', formula)
        for name in named_refs:
            if name.upper() not in ['SUM', 'AVERAGE', 'IF', 'IFERROR', 'MAX', 'MIN', 'STDEV']:
                refs.append(f"<NamedRange:{name}>")
        
        return refs
    
    def _get_reference_content(self, ref: str) -> any:
        """
        참조 셀의 실제 내용 확인
        
        Args:
            ref: 참조 (예: "Scenarios!B13", "A5", "<NamedRange:TAM>")
        
        Returns:
            셀 내용 (값 또는 수식)
        """
        
        try:
            # Named Range
            if ref.startswith('<NamedRange:'):
                range_name = ref.replace('<NamedRange:', '').replace('>', '')
                
                if range_name in self.wb.defined_names:
                    # Named Range의 실제 위치와 값
                    for sheet, cell in self.wb.defined_names[range_name].destinations:
                        ws = self.wb[sheet]
                        cell_clean = cell.replace('$', '')
                        return ws[cell_clean].value
                
                return f"<NotFound:{range_name}>"
            
            # 시트 참조 (Sheet!Cell)
            if '!' in ref:
                sheet_name, cell_addr = ref.split('!')
                
                if sheet_name in self.wb.sheetnames:
                    ws = self.wb[sheet_name]
                    cell = ws[cell_addr]
                    
                    # A열 (라벨)의 내용도 함께 반환
                    row = int(re.search(r'\d+', cell_addr).group())
                    label_cell = ws[f'A{row}']
                    
                    if label_cell.value:
                        return f"{label_cell.value} (값: {cell.value})"
                    else:
                        return cell.value
                
                return f"<SheetNotFound:{sheet_name}>"
            
            # 현재 시트 셀 참조
            # (현재 어느 시트인지 알 수 없으므로 생략)
            return f"<CurrentSheet:{ref}>"
        
        except Exception as e:
            return f"<Error:{e}>"
    
    def _compile_results(self) -> Dict:
        """결과 정리"""
        
        passed = len(self.errors) == 0
        
        print("\n" + "="*70)
        print("📊 참조 검증 결과")
        print("="*70)
        
        if passed:
            print("✅ 모든 참조 검증 통과!")
        else:
            print(f"❌ {len(self.errors)}개 오류 발견")
            
            print("\n오류 목록:")
            for error in self.errors:
                print(f"  {error}")
        
        if self.warnings:
            print(f"\n⚠️ 경고 ({len(self.warnings)}개):")
            for warning in self.warnings:
                print(f"  {warning}")
        
        return {
            'passed': passed,
            'errors': self.errors,
            'warnings': self.warnings
        }


# 편의 함수
def validate_formula_references(filepath: Path) -> bool:
    """
    수식 참조 검증 (편의 함수)
    
    Args:
        filepath: Excel 파일
    
    Returns:
        통과 여부
    """
    
    validator = FormulaReferenceValidator(filepath)
    result = validator.validate()
    return result['passed']


# 사용 예시
# python scripts/validate_formula_references.py

