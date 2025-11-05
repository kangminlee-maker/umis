"""
Should vs Will 시트 빌더
행동경제학 기반 규범 vs 현실 분석
"""

import sys
from pathlib import Path

# 프로젝트 루트 추가
if __name__ == "__main__":
    project_root = Path(__file__).parent.parent.parent.parent
    sys.path.insert(0, str(project_root))
    from umis_rag.deliverables.excel.formula_engine import FormulaEngine, ExcelStyles
else:
    from .formula_engine import FormulaEngine, ExcelStyles

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Optional


class ShouldVsWillBuilder:
    """
    Should vs Will 분석 시트
    
    Domain-Centric Reasoner의 s4 (Behavioral Economics) 결과 표시
    
    구조:
      - Should: 규범적 결론 (편향 제거)
      - Will: 현실적 예측 (편향 반영)
      - Gap: 차이 분석
      - 행동경제학 보정 계수
    """
    
    def __init__(self, workbook: Workbook, formula_engine: Optional[FormulaEngine] = None):
        """
        초기화
        
        Args:
            workbook: Excel Workbook
            formula_engine: FormulaEngine (선택)
        """
        self.wb = workbook
        self.formula_engine = formula_engine or FormulaEngine(workbook)
        self.styles = ExcelStyles()
    
    def create_sheet(
        self,
        should_vs_will_data: Optional[Dict] = None
    ) -> Worksheet:
        """
        Should vs Will 시트 생성
        
        Args:
            should_vs_will_data: {
                'items': [
                    {
                        'metric': str,
                        'should': float,
                        'will': float,
                        'reason': str,
                        'biases': [...]
                    },
                    ...
                ]
            }
        
        Returns:
            생성된 Worksheet
        """
        
        sheet = self.wb.create_sheet("Should_vs_Will")
        
        # 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === 제목 ===
        sheet.merge_cells('A1:F1')
        title_cell = sheet['A1']
        title_cell.value = "Should vs Will 분석 (행동경제학 보정)"
        title_cell.font = Font(bold=True, size=14, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 30
        
        # === 설명 ===
        sheet.merge_cells('A2:F2')
        desc_cell = sheet['A2']
        desc_cell.value = "Domain-Centric Reasoner s4 신호: 규범적 권고 vs 현실적 예측"
        desc_cell.font = Font(italic=True, size=10, color="7F7F7F")
        desc_cell.alignment = Alignment(horizontal='center')
        
        # === 헤더 ===
        headers = [
            ('A4', '항목', 25),
            ('B4', 'Should (규범적)', 18),
            ('C4', 'Will (현실적)', 18),
            ('D4', 'Gap (%)', 12),
            ('E4', '주요 원인', 20),
            ('F4', '조정 계수', 15)
        ]
        
        for col, header, width in headers:
            cell = sheet[col]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # 열 너비
            sheet.column_dimensions[col[0]].width = width
        
        sheet.row_dimensions[4].height = 25
        
        # === 데이터 행 ===
        
        if should_vs_will_data and should_vs_will_data.get('items'):
            items = should_vs_will_data['items']
            
            for idx, item in enumerate(items, start=5):
                row = idx
                
                # 항목
                sheet[f'A{row}'] = item.get('metric', '')
                sheet[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                sheet[f'A{row}'].border = border
                
                # Should
                should_val = item.get('should', 0)
                sheet[f'B{row}'] = should_val
                sheet[f'B{row}'].number_format = self._get_number_format(item.get('unit', 'number'))
                sheet[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
                sheet[f'B{row}'].border = border
                
                # Will
                will_val = item.get('will', 0)
                sheet[f'C{row}'] = will_val
                sheet[f'C{row}'].number_format = self._get_number_format(item.get('unit', 'number'))
                sheet[f'C{row}'].alignment = Alignment(horizontal='right', vertical='center')
                sheet[f'C{row}'].border = border
                
                # Gap (%) - Formula
                sheet[f'D{row}'] = f"=IF(B{row}=0, 0, (B{row}-C{row})/B{row}*100)"
                sheet[f'D{row}'].number_format = '0.0"%"'
                sheet[f'D{row}'].alignment = Alignment(horizontal='right', vertical='center')
                sheet[f'D{row}'].border = border
                
                # 조건부 서식 (Gap)
                gap_cell = sheet[f'D{row}']
                if should_val and will_val:
                    gap_pct = (should_val - will_val) / should_val * 100
                    if gap_pct > 30:
                        gap_cell.fill = PatternFill(start_color="FFE6E6", fill_type="solid")  # 연한 빨강
                    elif gap_pct < -30:
                        gap_cell.fill = PatternFill(start_color="E6FFE6", fill_type="solid")  # 연한 초록
                
                # 주요 원인
                reason = item.get('reason', '')
                sheet[f'E{row}'] = reason
                sheet[f'E{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sheet[f'E{row}'].border = border
                
                # 조정 계수
                biases = item.get('biases', [])
                if biases:
                    bias_text = ', '.join([f"{b['bias']}: ×{b['factor']}" for b in biases])
                    sheet[f'F{row}'] = bias_text
                else:
                    sheet[f'F{row}'] = '없음 (편향 없음)'
                sheet[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                sheet[f'F{row}'].border = border
                sheet[f'F{row}'].font = Font(size=9)
        
        else:
            # 예시 데이터
            self._add_example_data(sheet)
        
        # === 요약 섹션 ===
        self._add_summary_section(sheet, should_vs_will_data)
        
        # === 행동경제학 가이드 ===
        self._add_behavioral_econ_guide(sheet)
        
        return sheet
    
    def _get_number_format(self, unit: str) -> str:
        """단위에 따른 숫자 포맷"""
        if unit == 'percentage' or unit == '%':
            return '0.0"%"'
        elif unit == 'krw' or unit == 'currency':
            return '#,##0'
        elif unit == 'krw_billion':
            return '#,##0" 억"'
        else:
            return '#,##0.0'
    
    def _add_example_data(self, sheet: Worksheet):
        """예시 데이터 추가"""
        
        examples = [
            {
                'metric': '플랫폼 수수료율 (%)',
                'should': 0.075,  # 7.5%
                'will': 0.095,    # 9.5%
                'reason': '시장 지배력 (독과점)',
                'unit': 'percentage',
                'biases': [{'bias': 'market_power', 'factor': 1.27}]
            },
            {
                'metric': '시니어 케어 로봇 시장 (억 원)',
                'should': 5000,
                'will': 2850,
                'reason': '기술 거부감 + 가격 부담',
                'unit': 'krw_billion',
                'biases': [
                    {'bias': 'tech_resistance', 'factor': 0.3},
                    {'bias': 'price_burden', 'factor': 0.6}
                ]
            },
            {
                'metric': '구독 전환율 (%)',
                'should': 0.30,
                'will': 0.15,
                'reason': '현상유지 편향',
                'unit': 'percentage',
                'biases': [{'bias': 'status_quo_bias', 'factor': 0.5}]
            }
        ]
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for idx, ex in enumerate(examples, start=5):
            row = idx
            
            # 데이터 입력
            sheet[f'A{row}'] = ex['metric']
            sheet[f'B{row}'] = ex['should']
            sheet[f'C{row}'] = ex['will']
            sheet[f'D{row}'] = f"=IF(B{row}=0, 0, (B{row}-C{row})/B{row}*100)"
            sheet[f'E{row}'] = ex['reason']
            
            if ex['biases']:
                bias_text = ', '.join([f"{b['bias']}: ×{b['factor']}" for b in ex['biases']])
                sheet[f'F{row}'] = bias_text
            else:
                sheet[f'F{row}'] = '없음'
            
            # 포맷팅
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = sheet[f'{col}{row}']
                cell.border = border
            
            sheet[f'B{row}'].number_format = self._get_number_format(ex['unit'])
            sheet[f'C{row}'].number_format = self._get_number_format(ex['unit'])
            sheet[f'D{row}'].number_format = '0.0"%"'
            
            # Gap 색상
            gap_val = (ex['should'] - ex['will']) / ex['should'] * 100
            if gap_val > 30:
                sheet[f'D{row}'].fill = PatternFill(start_color="FFE6E6", fill_type="solid")
            elif gap_val < -30:
                sheet[f'D{row}'].fill = PatternFill(start_color="E6FFE6", fill_type="solid")
    
    def _add_summary_section(self, sheet: Worksheet, data: Optional[Dict]):
        """요약 섹션 추가"""
        
        start_row = 10  # 예시 데이터 후
        
        # 제목
        sheet[f'A{start_row}'] = "요약"
        sheet[f'A{start_row}'].font = Font(bold=True, size=12)
        
        # 평균 Gap
        sheet[f'A{start_row+1}'] = "평균 Gap (%)"
        sheet[f'B{start_row+1}'] = "=AVERAGE(D5:D7)"  # 예시 3개 평균
        sheet[f'B{start_row+1}'].number_format = '0.0"%"'
        sheet[f'B{start_row+1}'].font = Font(bold=True)
        
        # 최대 Gap
        sheet[f'A{start_row+2}'] = "최대 Gap (%)"
        sheet[f'B{start_row+2}'] = "=MAX(D5:D7)"
        sheet[f'B{start_row+2}'].number_format = '0.0"%"'
        
        # 최소 Gap
        sheet[f'A{start_row+3}'] = "최소 Gap (%)"
        sheet[f'B{start_row+3}'] = "=MIN(D5:D7)"
        sheet[f'B{start_row+3}'].number_format = '0.0"%"'
    
    def _add_behavioral_econ_guide(self, sheet: Worksheet):
        """행동경제학 가이드 추가"""
        
        start_row = 15
        
        # 제목
        sheet.merge_cells(f'A{start_row}:F{start_row}')
        title = sheet[f'A{start_row}']
        title.value = "행동경제학 편향 가이드"
        title.font = Font(bold=True, size=12, color="1F4E78")
        title.alignment = Alignment(horizontal='center')
        
        # 가이드 내용
        guide_data = [
            ('편향', '설명', '조정 계수', '적용 예시'),
            ('손실회피', '손실 = 이득 × 2.5', '×0.4-0.6', '가격 인상 저항'),
            ('현상유지', '전환 저항, 현상 유지 선호', '×0.5', '구독/플랫폼 전환'),
            ('시장 지배력', '독과점 → 가격 결정력', '×1.0-1.3', '플랫폼 수수료'),
            ('기술 거부감', '노인층, 보수 산업', '×0.3', '로봇, AI 채택'),
            ('가격 부담', '고가 제품 구매 주저', '×0.6', '500만원+ 제품')
        ]
        
        for idx, row_data in enumerate(guide_data, start=start_row+1):
            row = idx
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row, col_idx, value)
                
                if idx == start_row + 1:  # 헤더
                    cell.font = Font(bold=True, size=10)
                    cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
                else:
                    cell.font = Font(size=9)
                
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 열 너비 조정
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        
        # === 사용 가이드 ===
        guide_row = start_row + 8
        
        sheet.merge_cells(f'A{guide_row}:F{guide_row}')
        guide_title = sheet[f'A{guide_row}']
        guide_title.value = "💡 사용 가이드"
        guide_title.font = Font(bold=True, size=11)
        
        guide_texts = [
            "• Should: 이상적/규범적 결론 (편향 제거) → 정책 권고, 목표 설정",
            "• Will: 현실적 예측 (편향 반영) → 실제 채택률, 매출 예측",
            "• Gap > 30%: 큰 차이 → 개선 기회 또는 현실 수용 필요",
            "• Gap < 10%: 작은 차이 → 편향 영향 미미"
        ]
        
        for idx, text in enumerate(guide_texts, start=guide_row+1):
            sheet.merge_cells(f'A{idx}:F{idx}')
            cell = sheet[f'A{idx}']
            cell.value = text
            cell.font = Font(size=9, italic=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        return sheet


# 독립 실행 테스트
if __name__ == "__main__":
    import sys
    from pathlib import Path
    
    # 프로젝트 루트 추가
    project_root = Path(__file__).parent.parent.parent.parent
    sys.path.insert(0, str(project_root))
    
    print("\n" + "=" * 60)
    print("Should vs Will Builder 테스트")
    print("=" * 60)
    
    wb = Workbook()
    
    # 예시 데이터
    test_data = {
        'items': [
            {
                'metric': '플랫폼 수수료율 (%)',
                'should': 0.075,
                'will': 0.095,
                'reason': '시장 지배력 85%',
                'unit': 'percentage',
                'biases': [{'bias': 'market_power', 'factor': 1.27}]
            },
            {
                'metric': '시니어 케어 로봇 시장 (억)',
                'should': 5000,
                'will': 2850,
                'reason': '기술 거부감 + 가격 부담',
                'unit': 'krw_billion',
                'biases': [
                    {'bias': 'tech_resistance', 'factor': 0.3},
                    {'bias': 'price_burden', 'factor': 0.6}
                ]
            }
        ]
    }
    
    builder = ShouldVsWillBuilder(wb)
    sheet = builder.create_sheet(test_data)
    
    print(f"\n✅ 시트 생성 완료: {sheet.title}")
    print(f"   데이터 행: {len(test_data['items'])}개")
    
    # 저장
    output_path = Path("test_output/should_vs_will_test.xlsx")
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    
    print(f"   저장: {output_path}")
    print("\n✅ 테스트 완료")

