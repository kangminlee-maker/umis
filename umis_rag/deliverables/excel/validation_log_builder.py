"""
Validation Log Sheet Builder
데이터 검증 이력 및 출처 기록

구조:
  - 검증 항목
  - 데이터 출처
  - 검증 상태
  - 검증자/일자
  - 노트
"""

from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

from .formula_engine import ExcelStyles


def _get_header_styles():
    """헤더 스타일 생성 헬퍼"""
    header_font = Font(size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, thin_border


class ValidationLogBuilder:
    """
    Validation Log 시트 빌더
    
    기능:
      - 검증 이력 기록
      - 데이터 출처 추적
      - 검증 상태 시각화
      - 검증 완료율 계산
    """
    
    def __init__(self, workbook: Workbook):
        """
        Args:
            workbook: openpyxl Workbook
        """
        self.wb = workbook
    
    def create_sheet(
        self,
        validation_items: Optional[List[Dict]] = None
    ) -> None:
        """
        Validation Log 시트 생성
        
        Args:
            validation_items: 검증 항목 목록
                [
                    {
                        'category': 'Assumption',
                        'item': 'TAM (Total Addressable Market)',
                        'source': 'Statista 2024',
                        'status': 'Validated',
                        'validator': 'Rachel',
                        'date': '2024-11-04',
                        'notes': '신뢰도 A등급'
                    },
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("Validation_Log")
        
        # === 1. 제목 ===
        ws['A1'] = "Data Validation Log"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=9, italic=True, color="666666")
        ws.merge_cells('A2:G2')
        
        # === 2. 컬럼 헤더 ===
        headers = [
            'Category',
            'Item',
            'Data Source',
            'Status',
            'Validator',
            'Date',
            'Notes'
        ]
        
        header_font, header_fill, thin_border = _get_header_styles()
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 15  # Category
        ws.column_dimensions['B'].width = 30  # Item
        ws.column_dimensions['C'].width = 25  # Source
        ws.column_dimensions['D'].width = 12  # Status
        ws.column_dimensions['E'].width = 15  # Validator
        ws.column_dimensions['F'].width = 12  # Date
        ws.column_dimensions['G'].width = 40  # Notes
        
        # === 3. 기본 검증 항목 또는 사용자 제공 항목 ===
        if validation_items is None:
            validation_items = self._get_default_validation_items()
        
        _, _, thin_border = _get_header_styles()
        
        row = 5
        for item in validation_items:
            # A: Category
            ws.cell(row=row, column=1).value = item.get('category', '')
            ws.cell(row=row, column=1).font = Font(size=10)
            ws.cell(row=row, column=1).border = thin_border
            
            # B: Item
            ws.cell(row=row, column=2).value = item.get('item', '')
            ws.cell(row=row, column=2).font = Font(size=10)
            ws.cell(row=row, column=2).border = thin_border
            
            # C: Source
            ws.cell(row=row, column=3).value = item.get('source', '')
            ws.cell(row=row, column=3).font = Font(size=10)
            ws.cell(row=row, column=3).border = thin_border
            
            # D: Status
            status = item.get('status', 'Pending')
            ws.cell(row=row, column=4).value = status
            ws.cell(row=row, column=4).font = Font(size=10, bold=True)
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4).border = thin_border
            
            # E: Validator
            ws.cell(row=row, column=5).value = item.get('validator', '')
            ws.cell(row=row, column=5).font = Font(size=10)
            ws.cell(row=row, column=5).border = thin_border
            
            # F: Date
            ws.cell(row=row, column=6).value = item.get('date', '')
            ws.cell(row=row, column=6).font = Font(size=10)
            ws.cell(row=row, column=6).border = thin_border
            
            # G: Notes
            ws.cell(row=row, column=7).value = item.get('notes', '')
            ws.cell(row=row, column=7).font = Font(size=9)
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=7).border = thin_border
            
            row += 1
        
        # === 4. 통계 요약 ===
        row += 1
        ws.cell(row=row, column=1).value = "Validation Summary"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        total_items = len(validation_items)
        validated_items = sum(1 for item in validation_items if item.get('status') == 'Validated')
        pending_items = total_items - validated_items
        
        ws.cell(row=row, column=1).value = "Total Items:"
        ws.cell(row=row, column=2).value = total_items
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "Validated:"
        ws.cell(row=row, column=2).value = validated_items
        ws.cell(row=row, column=2).font = Font(bold=True, color="006100")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1).value = "Pending:"
        ws.cell(row=row, column=2).value = pending_items
        ws.cell(row=row, column=2).font = Font(bold=True, color="9C0006")
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1).value = "Completion Rate:"
        completion_rate = (validated_items / total_items * 100) if total_items > 0 else 0
        ws.cell(row=row, column=2).value = f"{completion_rate:.1f}%"
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        # === 5. 조건부 서식 (Status 컬럼) ===
        # Validated = 녹색
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        validated_rule = FormulaRule(
            formula=[f'$D5="Validated"'],
            stopIfTrue=True,
            fill=green_fill,
            font=green_font
        )
        ws.conditional_formatting.add(f'D5:D{row-5}', validated_rule)
        
        # Pending = 노란색
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(color="9C6500")
        pending_rule = FormulaRule(
            formula=[f'$D5="Pending"'],
            stopIfTrue=True,
            fill=yellow_fill,
            font=yellow_font
        )
        ws.conditional_formatting.add(f'D5:D{row-5}', pending_rule)
        
        # Rejected = 빨간색
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006", bold=True)
        rejected_rule = FormulaRule(
            formula=[f'$D5="Rejected"'],
            stopIfTrue=True,
            fill=red_fill,
            font=red_font
        )
        ws.conditional_formatting.add(f'D5:D{row-5}', rejected_rule)
        
        print(f"   ✅ Validation Log 시트 생성 완료")
        print(f"      - {total_items}개 검증 항목")
        print(f"      - 완료율: {completion_rate:.1f}%")
    
    def _get_default_validation_items(self) -> List[Dict]:
        """
        기본 검증 항목 반환
        
        Returns:
            검증 항목 목록
        """
        
        return [
            {
                'category': 'Market Size',
                'item': 'TAM (Total Addressable Market)',
                'source': 'Statista 2024',
                'status': 'Validated',
                'validator': 'Rachel',
                'date': '2024-11-04',
                'notes': '신뢰도 A등급, 공식 보고서 확인'
            },
            {
                'category': 'Market Size',
                'item': 'SAM (Serviceable Addressable Market)',
                'source': 'Calculated (4 methods)',
                'status': 'Validated',
                'validator': 'Bill',
                'date': '2024-11-04',
                'notes': 'Convergence ±30% 통과'
            },
            {
                'category': 'Assumption',
                'item': 'Market Share Target',
                'source': 'Competitive Analysis',
                'status': 'Pending',
                'validator': '',
                'date': '',
                'notes': '경쟁사 점유율 추가 검증 필요'
            },
            {
                'category': 'Assumption',
                'item': 'Growth Rate (CAGR)',
                'source': 'Industry Report',
                'status': 'Validated',
                'validator': 'Rachel',
                'date': '2024-11-04',
                'notes': 'Gartner 2024 보고서 기준'
            },
            {
                'category': 'Benchmark',
                'item': 'Customer Acquisition Cost (CAC)',
                'source': 'Industry Benchmark',
                'status': 'Validated',
                'validator': 'Bill',
                'date': '2024-11-04',
                'notes': '유사 기업 평균치 적용'
            },
            {
                'category': 'Benchmark',
                'item': 'Churn Rate',
                'source': 'ProfitWell 2024',
                'status': 'Pending',
                'validator': '',
                'date': '',
                'notes': '서비스별 Churn 추가 확인'
            },
            {
                'category': 'Proxy Data',
                'item': 'Adjacent Market Size',
                'source': 'Bloomberg Terminal',
                'status': 'Validated',
                'validator': 'Rachel',
                'date': '2024-11-04',
                'notes': '유사 시장 3개 데이터 평균'
            },
            {
                'category': 'Competitor',
                'item': 'Competitor Revenue Data',
                'source': 'DART / IR',
                'status': 'Validated',
                'validator': 'Rachel',
                'date': '2024-11-04',
                'notes': '공시 데이터 검증 완료'
            }
        ]


# 테스트는 별도 스크립트에서
# python scripts/test_excel_generation.py

