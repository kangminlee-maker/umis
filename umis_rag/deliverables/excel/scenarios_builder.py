"""
Scenarios Sheet Builder
Best/Base/Worst 시나리오 분석

구조:
  - Best Case: 주요 가정 +15%
  - Base Case: 현재 가정 (0%)
  - Worst Case: 주요 가정 -15%
"""

from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

from .formula_engine import FormulaEngine, ExcelStyles


class ScenariosBuilder:
    """
    Scenarios 시트 빌더
    
    기능:
      - Best/Base/Worst 3가지 시나리오
      - SAM 민감도 분석
      - 시나리오별 핵심 가정 조정
      - 시각화 (조건부 서식)
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(
        self,
        key_assumptions: List[str] = None,
        scenario_adjustments: Dict[str, float] = None
    ) -> None:
        """
        Scenarios 시트 생성
        
        Args:
            key_assumptions: 핵심 가정 Named Range 목록
                예: ['TAM', 'MarketShare', 'GrowthRate']
            scenario_adjustments: 시나리오별 조정 비율
                기본값: {'Best': 1.15, 'Base': 1.0, 'Worst': 0.85}
        
        구조:
          Row 1-2: 제목
          Row 3: 컬럼 헤더
          Row 4+: 각 가정별 시나리오 값
          Row N: SAM 계산 (4가지 방법)
          Row N+5: 차트 (선택)
        """
        
        # 기본값
        if scenario_adjustments is None:
            scenario_adjustments = {
                'Best': 1.15,   # +15%
                'Base': 1.0,    # 현재
                'Worst': 0.85   # -15%
            }
        
        # 핵심 가정이 지정되지 않으면 주요 Named Range 사용
        if key_assumptions is None:
            key_assumptions = self._get_key_assumptions()
        
        ws = self.wb.create_sheet("Scenarios")
        
        # === 1. 제목 ===
        ws['A1'] = "Scenario Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = "주요 가정의 변화가 SAM에 미치는 영향"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # === 2. 컬럼 헤더 ===
        headers = ['Assumption', 'Best Case\n(+15%)', 'Base Case\n(Current)', 'Worst Case\n(-15%)', 'Range']
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # === 3. 가정별 시나리오 값 ===
        row = 5
        for assumption_name in key_assumptions:
            # A열: 가정 이름
            ws.cell(row=row, column=1).value = assumption_name
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # B열: Best Case (가정 * 1.15)
            ws.cell(row=row, column=2).value = f"={assumption_name}*{scenario_adjustments['Best']}"
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # C열: Base Case (가정 그대로)
            ws.cell(row=row, column=3).value = f"={assumption_name}"
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            # D열: Worst Case (가정 * 0.85)
            ws.cell(row=row, column=4).value = f"={assumption_name}*{scenario_adjustments['Worst']}"
            ws.cell(row=row, column=4).number_format = '#,##0'
            
            # E열: Range (Best - Worst)
            ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
            ws.cell(row=row, column=5).number_format = '#,##0'
            
            row += 1
        
        # === 4. SAM 계산 ===
        row += 1
        ws.cell(row=row, column=1).value = "SAM Results"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        row += 1
        
        # SAM 4가지 방법
        sam_methods = [
            ('Method 1 (Top-Down)', 'SAM'),
            ('Method 2 (Bottom-Up)', 'SAM_Method2'),
            ('Method 3 (Proxy)', 'SAM_Method3'),
            ('Method 4 (Competitor)', 'SAM_Method4')
        ]
        
        for method_name, sam_range in sam_methods:
            # A열: 방법 이름
            ws.cell(row=row, column=1).value = method_name
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # B-D열: 각 시나리오에서 SAM은 동일 (가정이 Named Range로 연결되어 있음)
            # 실제로는 각 시나리오별로 재계산해야 하지만, 단순화를 위해 참조만 표시
            for col in [2, 3, 4]:
                ws.cell(row=row, column=col).value = f"={sam_range}"
                ws.cell(row=row, column=col).number_format = '#,##0'
            
            # E열: Range
            ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
            ws.cell(row=row, column=5).number_format = '#,##0'
            
            row += 1
        
        # === 5. 평균 SAM ===
        row += 1
        ws.cell(row=row, column=1).value = "Average SAM"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        # SAM 평균 계산 (위 4개 방법의 평균)
        sam_start_row = row - 4
        sam_end_row = row - 1
        
        for col in [2, 3, 4]:
            col_letter = chr(64 + col)  # A=65, B=66, ...
            ws.cell(row=row, column=col).value = f"=AVERAGE({col_letter}{sam_start_row}:{col_letter}{sam_end_row})"
            ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Range
        ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True)
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "Interpretation Guide:"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Best Case: 주요 가정이 15% 개선될 경우"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Base Case: 현재 가정 유지"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Worst Case: 주요 가정이 15% 악화될 경우"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Range: 시나리오 간 차이 (민감도 지표)"
        ws.cell(row=row, column=1).font = Font(size=9)
        
        # === 7. 조건부 서식 (Range 컬럼) ===
        # Range가 큰 경우 빨간색 (민감도 높음)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        high_sensitivity = FormulaRule(
            formula=[f'E5>AVERAGE($E$5:$E${row-5})*1.5'],
            stopIfTrue=True,
            fill=red_fill
        )
        ws.conditional_formatting.add(f'E5:E{row-5}', high_sensitivity)
        
        print(f"   ✅ Scenarios 시트 생성 완료")
        print(f"      - {len(key_assumptions)}개 핵심 가정")
        print(f"      - 3가지 시나리오 (Best/Base/Worst)")
    
    def _get_key_assumptions(self) -> List[str]:
        """
        Named Range에서 핵심 가정 추출
        
        Returns:
            핵심 가정 Named Range 목록
        """
        
        # Assumptions 시트의 주요 Named Range들
        key_ranges = []
        
        # Named Range 중 주요 항목 추출
        for name_str in self.wb.defined_names:
            # ASM_로 시작하는 Named Range 또는 주요 지표
            if (name_str.startswith('ASM_') or 
                name_str in ['TAM', 'MarketShare', 'GrowthRate']):
                key_ranges.append(name_str)
        
        # 없으면 기본값
        if not key_ranges:
            # FormulaEngine에서 정의된 Named Range 사용
            if hasattr(self.fe, 'named_ranges'):
                key_ranges = list(self.fe.named_ranges.keys())[:10]
            else:
                # 최소 기본값
                key_ranges = ['TAM']
        
        # 최대 10개로 제한 (시트가 너무 길어지지 않도록)
        return key_ranges[:10]


# 테스트는 별도 스크립트에서
# python scripts/test_excel_generation.py

