"""
Unit Economics Scenarios Sheet Builder
시나리오별 Unit Economics 분석

Sheet 8: Scenarios
- Conservative (보수적)
- Base (기본)
- Optimistic (낙관적)
- 각 시나리오별 LTV, CAC, Ratio, Payback
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class UEScenariosBuilder:
    """
    Unit Economics Scenarios 시트 빌더
    
    기능:
      - 3가지 시나리오 (Conservative/Base/Optimistic)
      - 시나리오별 핵심 지표
      - 비교 분석
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """Scenarios 시트 생성"""
        
        ws = self.wb.create_sheet("UE_Scenarios")
        
        # === 1. 제목 ===
        ws['A1'] = "Unit Economics Scenarios"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "보수적/기본/낙관적 시나리오별 Unit Economics 비교"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Metric', 'Conservative', 'Base', 'Optimistic', 'Range']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 입력 변수 조정 ===
        row += 1
        ws.cell(row=row, column=1).value = "Input Variables"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # ARPU
        row += 1
        ws.cell(row=row, column=1).value = "ARPU (월)"
        ws.cell(row=row, column=2).value = "=ARPU*0.85"  # Conservative: -15%
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "=ARPU"  # Base
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = "=ARPU*1.15"  # Optimistic: +15%
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # CAC
        row += 1
        ws.cell(row=row, column=1).value = "CAC"
        ws.cell(row=row, column=2).value = "=CAC*1.15"  # Conservative: +15% (높음)
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "=CAC"  # Base
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = "=CAC*0.85"  # Optimistic: -15% (낮음)
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # Churn
        row += 1
        ws.cell(row=row, column=1).value = "Monthly Churn"
        ws.cell(row=row, column=2).value = "=MonthlyChurn*1.15"  # Conservative: +15% (높음)
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "=MonthlyChurn"  # Base
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = "=MonthlyChurn*0.85"  # Optimistic: -15% (낮음)
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
        ws.cell(row=row, column=5).number_format = '0.0%'
        
        # Gross Margin
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=2).value = "=GrossMargin*0.9"  # Conservative: -10%
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "=GrossMargin"  # Base
        ws.cell(row=row, column=3).number_format = '0.0%'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = "=GrossMargin*1.1"  # Optimistic: +10%
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '0.0%'
        
        # === 4. 결과 지표 ===
        row += 2
        ws.cell(row=row, column=1).value = "Resulting Metrics"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # LTV
        arpu_row = row - 4
        margin_row = row - 1
        churn_row = row - 2
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # Conservative
        ws.cell(row=row, column=2).value = f"=B{arpu_row}*CustomerLifetime*B{margin_row}"
        ws.cell(row=row, column=2).number_format = '#,##0'
        
        # Base
        ws.cell(row=row, column=3).value = "=LTV"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Optimistic
        ws.cell(row=row, column=4).value = f"=D{arpu_row}*CustomerLifetime*D{margin_row}"
        ws.cell(row=row, column=4).number_format = '#,##0'
        
        # Range
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '#,##0'
        
        # LTV/CAC Ratio
        cac_row = row - 3
        row += 1
        ws.cell(row=row, column=1).value = "LTV/CAC Ratio"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # Conservative
        ws.cell(row=row, column=2).value = f"=B{row-1}/B{cac_row}"
        ws.cell(row=row, column=2).number_format = '0.00'
        
        # Base
        ws.cell(row=row, column=3).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=3).number_format = '0.00'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Optimistic
        ws.cell(row=row, column=4).value = f"=D{row-1}/D{cac_row}"
        ws.cell(row=row, column=4).number_format = '0.00'
        
        # Range
        ws.cell(row=row, column=5).value = f"=D{row}-B{row}"
        ws.cell(row=row, column=5).number_format = '0.00'
        
        # Payback Period
        row += 1
        ws.cell(row=row, column=1).value = "Payback Period"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # Conservative
        ws.cell(row=row, column=2).value = f"=B{cac_row}/(B{arpu_row}*B{margin_row})"
        ws.cell(row=row, column=2).number_format = '0.0'
        
        # Base
        ws.cell(row=row, column=3).value = "=PaybackPeriod"
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # Optimistic
        ws.cell(row=row, column=4).value = f"=D{cac_row}/(D{arpu_row}*D{margin_row})"
        ws.cell(row=row, column=4).number_format = '0.0'
        
        # Range
        ws.cell(row=row, column=5).value = f"=B{row}-D{row}"
        ws.cell(row=row, column=5).number_format = '0.0'
        
        # === 5. 시나리오 평가 ===
        row += 2
        ws.cell(row=row, column=1).value = "Scenario Evaluation"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        ratio_row = row - 2
        
        row += 1
        ws.cell(row=row, column=1).value = "Conservative 달성 가능?"
        ws.cell(row=row, column=2).value = f'=IF(B{ratio_row}>=3, "✅ Yes", "❌ No")'
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:C{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Optimistic 목표치"
        ws.cell(row=row, column=2).value = f'=IF(D{ratio_row}>=5, "✅ Excellent", "Good")'
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:C{row}')
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Conservative: ARPU -15%, CAC +15%, Churn +15% (최악)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Base: 현재 가정 유지"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Optimistic: ARPU +15%, CAC -15%, Churn -15% (최선)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Conservative에서도 LTV/CAC > 3.0이면 안정적인 비즈니스 모델"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ UE Scenarios 시트 생성 완료")
        print(f"      - 3가지 시나리오 (Conservative/Base/Optimistic)")
        print(f"      - LTV, LTV/CAC, Payback 비교")


# 테스트는 별도 스크립트에서

