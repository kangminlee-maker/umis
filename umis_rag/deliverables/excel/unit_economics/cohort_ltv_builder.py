"""
Cohort LTV Sheet Builder
월별 코호트 LTV 추적 시트

Sheet 8: Cohort_LTV
- 월별 코호트 (2023-01 ~ 2024-12)
- 코호트별 LTV
- Cohort Improvement Rate
- Trend Analysis
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class CohortLTVBuilder:
    """
    Cohort LTV 시트 빌더
    
    기능:
      - 월별 코호트 LTV 추적
      - Cohort Improvement Rate
      - Trend 시각화
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, months: int = 12) -> None:
        """
        Cohort LTV 시트 생성
        
        Args:
            months: 추적할 코호트 개수 (기본 12개월)
        """
        
        ws = self.wb.create_sheet("Cohort_LTV")
        
        # === 1. 제목 ===
        ws['A1'] = "Cohort LTV Tracking"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "월별 코호트의 LTV 변화 추적 (개선 추세 확인)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Cohort', 'Cohort LTV', 'vs Baseline', 'Improvement', 'Note']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 코호트 데이터 ===
        input_fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        
        # 첫 번째 코호트 (Baseline)
        row += 1
        baseline_row = row
        ws.cell(row=row, column=1).value = "2024-01 (Baseline)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ws.cell(row=row, column=2).value = "=LTV"  # 현재 LTV = Baseline
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        ws.cell(row=row, column=3).value = 0
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=4).value = "0%"
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=5).value = "기준 코호트"
        ws.cell(row=row, column=5).font = Font(size=9, italic=True)
        
        # 나머지 코호트 (입력 필요)
        cohorts = [
            '2024-02', '2024-03', '2024-04', '2024-05', '2024-06',
            '2024-07', '2024-08', '2024-09', '2024-10', '2024-11', '2024-12'
        ]
        
        for cohort_name in cohorts[:months-1]:
            row += 1
            
            # A: Cohort
            ws.cell(row=row, column=1).value = cohort_name
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # B: Cohort LTV (입력 필요 - 실제 데이터로 채움)
            ws.cell(row=row, column=2).value = None  # 수동 입력
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # C: vs Baseline
            ws.cell(row=row, column=3).value = f"=IFERROR(B{row}-B${baseline_row}, 0)"
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            # D: Improvement
            ws.cell(row=row, column=4).value = f"=IFERROR(C{row}/B${baseline_row}*100, 0)"
            ws.cell(row=row, column=4).number_format = '0.0"%"'
            
            # E: Note
            ws.cell(row=row, column=5).value = (
                f'=IF(D{row}>=20, "✅ 목표 달성 (20%+)", '
                f'IF(D{row}>=10, "양호 (10%+)", "개선 필요"))'
            )
            ws.cell(row=row, column=5).font = Font(size=9)
        
        # === 4. 평균 개선률 ===
        row += 1
        last_cohort_row = row - 1
        first_cohort_row = baseline_row + 1
        
        # Named Range for Cohort Improvements (D 컬럼, 각 코호트 개선률)
        cohort_improvement_names = []
        for idx, cohort_row in enumerate(range(first_cohort_row, last_cohort_row + 1), start=1):
            nr_name = f'Cohort_Improvement_{idx}'
            self.fe.define_named_range(nr_name, 'Cohort_LTV', f'D{cohort_row}')
            cohort_improvement_names.append(nr_name)
        
        ws.cell(row=row, column=1).value = "평균 Improvement"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # AVERAGE using Named Ranges
        ws.cell(row=row, column=4).value = f"=AVERAGE({','.join(cohort_improvement_names)})"
        ws.cell(row=row, column=4).number_format = '0.0"%"'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        
        # === 5. 목표 ===
        row += 2
        ws.cell(row=row, column=1).value = "Cohort Improvement 목표"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• 각 코호트 10-20% 개선 목표"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 개선 방법: Onboarding 최적화, Feature 개선, Churn 예방"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 실제 코호트 데이터를 B열에 입력하여 추적"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ Cohort LTV 시트 생성 완료")
        print(f"      - {months}개 코호트 추적")
        print(f"      - 개선률 자동 계산")


# 테스트는 별도 스크립트에서

