"""
LTV Calculation Sheet Builder
고객 생애 가치 (LTV) 계산 시트

Sheet 2: LTV Calculation
- LTV Formula 1: ARPU × Lifetime × Gross Margin
- LTV Formula 2: ARPU × Margin / Churn
- Confidence Interval
- 벤치마크 비교
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class LTVBuilder:
    """
    LTV Calculation 시트 빌더
    
    기능:
      - 2가지 LTV 계산 방식
      - Confidence Interval 계산
      - 벤치마크 비교
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """LTV Calculation 시트 생성"""
        
        ws = self.wb.create_sheet("LTV_Calculation")
        
        # === 1. 제목 ===
        ws['A1'] = "LTV (Customer Lifetime Value) Calculation"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "고객 생애 가치 = 고객 1명이 생애 동안 창출하는 총 수익"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        
        # === 2. Formula 1: ARPU × Lifetime × Margin ===
        row = 4
        ws.cell(row=row, column=1).value = "방법 1: Lifetime 기반"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Formula:"
        ws.cell(row=row, column=2).value = "LTV = ARPU × Lifetime × Gross Margin"
        ws.cell(row=row, column=2).font = Font(size=10, italic=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        # 계산
        calc_fill = PatternFill(start_color=ExcelStyles.CALC_FILL, end_color=ExcelStyles.CALC_FILL, fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=1).value = "ARPU (월)"
        ws.cell(row=row, column=2).value = "=ARPU"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        row += 1
        ws.cell(row=row, column=1).value = "Customer Lifetime"
        ws.cell(row=row, column=2).value = "=CustomerLifetime"
        ws.cell(row=row, column=2).number_format = '#,##0.0'
        ws.cell(row=row, column=3).value = "months"
        
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=2).value = "=GrossMargin"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "%"
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV (방법 1)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # LTV 계산
        ltv_formula = self.fe.create_ltv_formula('ARPU', 'CustomerLifetime', 'GrossMargin')
        ws.cell(row=row, column=2).value = ltv_formula
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=11, bold=True)
        ws.cell(row=row, column=3).value = "원"
        
        # Named Range for LTV1
        self.fe.define_named_range('LTV_Method1', 'LTV_Calculation', f'B{row}')
        
        # === 3. Formula 2: ARPU × Margin / Churn ===
        row += 2
        ws.cell(row=row, column=1).value = "방법 2: Churn 기반"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Formula:"
        ws.cell(row=row, column=2).value = "LTV = ARPU × Gross Margin / Monthly Churn"
        ws.cell(row=row, column=2).font = Font(size=10, italic=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "ARPU (월)"
        ws.cell(row=row, column=2).value = "=ARPU"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=2).value = "=GrossMargin"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "%"
        
        row += 1
        ws.cell(row=row, column=1).value = "Monthly Churn Rate"
        ws.cell(row=row, column=2).value = "=MonthlyChurn"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "%"
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV (방법 2)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # LTV 계산 (Churn 기반)
        ltv_formula2 = self.fe.create_ltv_from_churn('ARPU', 'GrossMargin', 'MonthlyChurn')
        ws.cell(row=row, column=2).value = ltv_formula2
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=11, bold=True)
        ws.cell(row=row, column=3).value = "원"
        
        # Named Range for LTV2
        self.fe.define_named_range('LTV_Method2', 'LTV_Calculation', f'B{row}')
        
        # === 4. 평균 LTV (최종) ===
        row += 2
        ws.cell(row=row, column=1).value = "최종 LTV (2가지 방법 평균)"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        
        # 평균 계산
        ws.cell(row=row, column=2).value = "=AVERAGE(LTV_Method1, LTV_Method2)"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=12, bold=True, color="FFFFFF")
        ws.cell(row=row, column=3).value = "원"
        
        # Named Range for Final LTV
        self.fe.define_named_range('LTV', 'LTV_Calculation', f'B{row}')
        
        # === 5. Confidence Interval ===
        row += 2
        ws.cell(row=row, column=1).value = "신뢰 구간 (±15%)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "Lower Bound (85%)"
        ws.cell(row=row, column=2).value = "=LTV*0.85"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        row += 1
        ws.cell(row=row, column=1).value = "Upper Bound (115%)"
        ws.cell(row=row, column=2).value = "=LTV*1.15"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• LTV = 고객 1명이 생애 동안 창출하는 총 매출 (Margin 반영)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 방법 1과 2의 평균을 사용하여 안정적인 추정"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• LTV > CAC × 3 이상이면 건강한 비즈니스 (다음 시트에서 확인)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ LTV Calculation 시트 생성 완료")
        print(f"      - 2가지 계산 방법 (Lifetime / Churn)")
        print(f"      - Named Range: LTV_Method1, LTV_Method2, LTV")


# 테스트는 별도 스크립트에서
# python scripts/test_unit_economics.py

