"""
Benchmark Comparison Sheet Builder
업계 벤치마크 비교 시트

Sheet 9: Benchmark_Comparison
- 업계별 LTV/CAC 벤치마크
- 우리 지표 vs 업계 평균
- Gap Analysis
- Positioning
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class BenchmarkBuilder:
    """
    Benchmark Comparison 시트 빌더
    
    기능:
      - 업계 벤치마크 비교
      - Gap Analysis
      - Positioning 평가
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, industry: str = 'SaaS') -> None:
        """
        Benchmark Comparison 시트 생성
        
        Args:
            industry: 산업 (SaaS, E-commerce, Subscription 등)
        """
        
        ws = self.wb.create_sheet("Benchmark_Comparison")
        
        # === 1. 제목 ===
        ws['A1'] = "Industry Benchmark Comparison"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = f"우리 Unit Economics vs 업계 벤치마크"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # === 2. 컬럼 헤더 ===
        row = 4
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Metric', 'Our Value', 'Industry Avg', 'Gap', 'Assessment']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === 3. 벤치마크 데이터 ===
        benchmarks = self._get_industry_benchmarks(industry)
        
        # LTV/CAC Ratio
        row += 1
        ws.cell(row=row, column=1).value = "LTV/CAC Ratio"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ws.cell(row=row, column=2).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        ws.cell(row=row, column=3).value = benchmarks['ltv_cac_avg']
        ws.cell(row=row, column=3).number_format = '0.00'
        
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = '0.00'
        
        ws.cell(row=row, column=5).value = f'=IF(D{row}>0, "우리가 높음 ✅", "개선 필요")'
        ws.cell(row=row, column=5).font = Font(size=9)
        
        # CAC Payback
        row += 1
        ws.cell(row=row, column=1).value = "CAC Payback (months)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ws.cell(row=row, column=2).value = "=PaybackPeriod"
        ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        ws.cell(row=row, column=3).value = benchmarks['payback_avg']
        ws.cell(row=row, column=3).number_format = '0.0'
        
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = '0.0'
        
        ws.cell(row=row, column=5).value = f'=IF(D{row}<0, "우리가 빠름 ✅", "개선 필요")'
        ws.cell(row=row, column=5).font = Font(size=9)
        
        # Monthly Churn
        row += 1
        ws.cell(row=row, column=1).value = "Monthly Churn Rate"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ws.cell(row=row, column=2).value = "=MonthlyChurn"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        ws.cell(row=row, column=3).value = benchmarks['churn_avg']
        ws.cell(row=row, column=3).number_format = '0.0%'
        
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        ws.cell(row=row, column=5).value = f'=IF(D{row}<0, "우리가 낮음 ✅", "개선 필요")'
        ws.cell(row=row, column=5).font = Font(size=9)
        
        # Gross Margin
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        ws.cell(row=row, column=2).value = "=GrossMargin"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=2).font = Font(bold=True)
        
        ws.cell(row=row, column=3).value = benchmarks['margin_avg']
        ws.cell(row=row, column=3).number_format = '0.0%'
        
        ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
        ws.cell(row=row, column=4).number_format = '0.0%'
        
        ws.cell(row=row, column=5).value = f'=IF(D{row}>0, "우리가 높음 ✅", "개선 필요")'
        ws.cell(row=row, column=5).font = Font(size=9)
        
        # === 4. 종합 평가 ===
        row += 2
        ws.cell(row=row, column=1).value = "종합 평가"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "우리 경쟁력:"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # 4개 지표 중 우수한 개수 계산
        ltv_row = row - 4
        payback_row = row - 3
        churn_row = row - 2
        margin_row = row - 1
        
        ws.cell(row=row, column=2).value = (
            f'=COUNTIF(E{ltv_row}:E{margin_row}, "*✅*")&"/"&4&" 지표 우수"'
        )
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "포지셔닝:"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        ws.cell(row=row, column=2).value = (
            f'=IF(COUNTIF(E{ltv_row}:E{margin_row}, "*✅*")>=3, '
            f'"업계 평균 이상 (경쟁력 우수)", '
            f'IF(COUNTIF(E{ltv_row}:E{margin_row}, "*✅*")>=2, '
            f'"업계 평균 수준", "개선 필요"))'
        )
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        # === 5. 업계별 벤치마크 참고표 ===
        row += 2
        ws.cell(row=row, column=1).value = "업계별 벤치마크 (참고)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        header_row = row
        ws.cell(row=row, column=1).value = "Industry"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "LTV/CAC"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "Payback"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        
        ws.cell(row=row, column=4).value = "Churn"
        ws.cell(row=row, column=4).font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        
        ws.cell(row=row, column=5).value = "Margin"
        ws.cell(row=row, column=5).font = header_font
        ws.cell(row=row, column=5).fill = header_fill
        
        # 업계별 데이터
        industries = [
            {'name': 'SaaS (Enterprise)', 'ltv_cac': 5.0, 'payback': 12, 'churn': 0.02, 'margin': 0.80},
            {'name': 'SaaS (SMB)', 'ltv_cac': 3.0, 'payback': 8, 'churn': 0.05, 'margin': 0.75},
            {'name': 'E-commerce', 'ltv_cac': 2.5, 'payback': 6, 'churn': 0.10, 'margin': 0.30},
            {'name': '구독 서비스', 'ltv_cac': 3.5, 'payback': 9, 'churn': 0.04, 'margin': 0.50},
            {'name': 'OTT/스트리밍', 'ltv_cac': 3.2, 'payback': 8, 'churn': 0.04, 'margin': 0.35},
        ]
        
        for ind in industries:
            row += 1
            ws.cell(row=row, column=1).value = ind['name']
            ws.cell(row=row, column=1).font = Font(size=9)
            
            ws.cell(row=row, column=2).value = ind['ltv_cac']
            ws.cell(row=row, column=2).number_format = '0.0'
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=3).value = ind['payback']
            ws.cell(row=row, column=3).number_format = '0'
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=4).value = ind['churn']
            ws.cell(row=row, column=4).number_format = '0.0%'
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=5).value = ind['margin']
            ws.cell(row=row, column=5).number_format = '0.0%'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Gap > 0: 우리가 업계 평균보다 우수"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 4개 지표 중 3개 이상 우수: 경쟁력 우수"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Churn, Payback은 낮을수록 좋음 (음수 Gap이 좋음)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ Benchmark Comparison 시트 생성 완료")
        print(f"      - 업계: {industry}")
        print(f"      - 5개 업계 벤치마크 참고")
    
    def _get_industry_benchmarks(self, industry: str) -> dict:
        """
        업계별 평균 벤치마크 반환
        
        Args:
            industry: 산업명
        
        Returns:
            벤치마크 딕셔너리
        """
        
        benchmarks_db = {
            'SaaS': {
                'ltv_cac_avg': 4.0,
                'payback_avg': 10.0,
                'churn_avg': 0.035,
                'margin_avg': 0.75
            },
            'E-commerce': {
                'ltv_cac_avg': 2.5,
                'payback_avg': 6.0,
                'churn_avg': 0.10,
                'margin_avg': 0.30
            },
            'Subscription': {
                'ltv_cac_avg': 3.5,
                'payback_avg': 9.0,
                'churn_avg': 0.04,
                'margin_avg': 0.50
            },
            'Streaming': {
                'ltv_cac_avg': 3.2,
                'payback_avg': 8.0,
                'churn_avg': 0.04,
                'margin_avg': 0.35
            }
        }
        
        return benchmarks_db.get(industry, benchmarks_db['SaaS'])


# 테스트는 별도 스크립트에서

