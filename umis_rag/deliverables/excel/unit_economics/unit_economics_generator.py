"""
Unit Economics Workbook Generator (완성 버전)
단위 경제성 분석 Excel 자동 생성

버전: Batch 3 완성 (10개 시트)
"""

from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime

from openpyxl import Workbook

from ..formula_engine import FormulaEngine
from .inputs_builder import InputsBuilder
from .ltv_builder import LTVBuilder
from .cac_builder import CACBuilder
from .ratio_builder import RatioBuilder
from .payback_builder import PaybackBuilder
from .sensitivity_builder import SensitivityBuilder
from .ue_scenarios_builder import UEScenariosBuilder
from .benchmark_builder import BenchmarkBuilder
from .cohort_ltv_builder import CohortLTVBuilder
from .dashboard_builder import UEDashboardBuilder


class UnitEconomicsGenerator:
    """
    Unit Economics Excel 자동 생성기 (완성)
    
    생성 시트 (10개):
      1. Dashboard (요약)
      2. Inputs
      3. LTV_Calculation
      4. CAC_Analysis
      5. LTV_CAC_Ratio
      6. Payback_Period
      7. Sensitivity_Analysis
      8. UE_Scenarios
      9. Cohort_LTV
      10. Benchmark_Comparison
    """
    
    def __init__(self):
        """초기화"""
        self.formula_engine: Optional[FormulaEngine] = None
    
    def generate(
        self,
        market_name: str,
        inputs_data: Dict,
        channels_data: List[Dict] = None,
        industry: str = 'SaaS',
        cohort_months: int = 12,
        output_dir: Path = Path('.')
    ) -> Path:
        """
        Unit Economics Workbook 생성 (완성)
        
        Args:
            market_name: 시장/비즈니스 이름
            inputs_data: 입력 데이터
                {
                    'arpu': 9000,
                    'cac': 25000,
                    'gross_margin': 0.35,
                    'monthly_churn': 0.04,
                    'customer_lifetime': 25,
                    'sm_spend_monthly': 5000000,
                    'new_customers_monthly': 200
                }
            channels_data: 채널별 CAC 데이터 (선택)
            industry: 업계 (SaaS, E-commerce, Subscription 등)
            cohort_months: 코호트 추적 개월 수
            output_dir: 출력 디렉토리
        
        Returns:
            생성된 Excel 파일 경로
        """
        
        print(f"🚀 Unit Economics Workbook 생성 시작")
        print(f"   시장: {market_name}")
        print(f"   버전: 완성 (10개 시트)")
        
        # 1. 워크북 초기화
        wb = Workbook()
        self.formula_engine = FormulaEngine(wb)
        
        # 기본 시트 제거
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # 2. Sheet 2: Inputs
        print(f"   2/10 Inputs...")
        inputs_builder = InputsBuilder(wb, self.formula_engine)
        inputs_builder.create_sheet(inputs_data)
        
        # 3. Sheet 3: LTV Calculation
        print(f"   3/10 LTV Calculation...")
        ltv_builder = LTVBuilder(wb, self.formula_engine)
        ltv_builder.create_sheet()
        
        # 4. Sheet 4: CAC Analysis
        print(f"   4/10 CAC Analysis...")
        cac_builder = CACBuilder(wb, self.formula_engine)
        cac_builder.create_sheet(channels_data)
        
        # 5. Sheet 5: LTV/CAC Ratio
        print(f"   5/10 LTV/CAC Ratio...")
        ratio_builder = RatioBuilder(wb, self.formula_engine)
        ratio_builder.create_sheet()
        
        # 6. Sheet 6: Payback Period
        print(f"   6/10 Payback Period...")
        payback_builder = PaybackBuilder(wb, self.formula_engine)
        payback_builder.create_sheet()
        
        # 7. Sheet 7: Sensitivity Analysis
        print(f"   7/10 Sensitivity Analysis...")
        sensitivity_builder = SensitivityBuilder(wb, self.formula_engine)
        sensitivity_builder.create_sheet()
        
        # 8. Sheet 8: Scenarios
        print(f"   8/10 UE Scenarios...")
        scenarios_builder = UEScenariosBuilder(wb, self.formula_engine)
        scenarios_builder.create_sheet()
        
        # 9. Sheet 9: Cohort LTV (Batch 3)
        print(f"   9/10 Cohort LTV...")
        cohort_builder = CohortLTVBuilder(wb, self.formula_engine)
        cohort_builder.create_sheet(cohort_months)
        
        # 10. Sheet 10: Benchmark Comparison (Batch 3)
        print(f"   10/10 Benchmark Comparison...")
        benchmark_builder = BenchmarkBuilder(wb, self.formula_engine)
        benchmark_builder.create_sheet(industry)
        
        # 11. Sheet 1: Dashboard (Batch 3, 맨 앞으로 이동)
        print(f"   1/10 Dashboard...")
        dashboard_builder = UEDashboardBuilder(wb, self.formula_engine)
        dashboard_builder.create_sheet(market_name)
        
        # 12. 강제 재계산 설정
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        # 13. 저장
        filename = f"unit_economics_{market_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = output_dir / filename
        
        output_dir.mkdir(parents=True, exist_ok=True)
        wb.save(filepath)
        
        print(f"\n✅ Excel 생성 완료: {filepath}")
        print(f"📊 시트: {len(wb.sheetnames)}개 (Dashboard, Inputs, LTV, CAC, Ratio, Payback, Sensitivity, Scenarios, Cohort, Benchmark)")
        print(f"📋 Named Range: {len(self.formula_engine.named_ranges)}개")
        print(f"🎉 Unit Economics Analyzer 완성!")
        
        return filepath


# 테스트는 별도 스크립트에서
# python scripts/test_unit_economics.py

