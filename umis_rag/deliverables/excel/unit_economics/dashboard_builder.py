"""
Unit Economics Dashboard Builder
요약 대시보드 시트

Sheet 10: Dashboard
- 핵심 지표 요약
- Traffic Light
- 권장사항
- 한 눈에 보는 건강도
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

from ..formula_engine import FormulaEngine, ExcelStyles


class UEDashboardBuilder:
    """
    Unit Economics Dashboard 시트 빌더
    
    기능:
      - 핵심 지표 요약
      - Traffic Light (색상 코딩)
      - 종합 평가
      - Action Items
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, market_name: str = "Target Market") -> None:
        """
        Dashboard 시트 생성
        
        Args:
            market_name: 시장/비즈니스 이름
        """
        
        ws = self.wb.create_sheet("Dashboard", 0)  # 첫 번째 시트로
        
        # === 1. 대시보드 제목 ===
        ws['A1'] = "Unit Economics Dashboard"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 35
        
        ws['A2'] = market_name
        ws['A2'].font = Font(size=12, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        
        # === 2. 핵심 지표 (Big Numbers) ===
        row = 4
        ws.cell(row=row, column=1).value = "📊 핵심 지표"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        # LTV
        row += 1
        ws.cell(row=row, column=1).value = "Customer Lifetime Value (LTV)"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=LTV"
        ws.cell(row=row, column=2).number_format = '₩#,##0'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # CAC
        row += 1
        ws.cell(row=row, column=1).value = "Customer Acquisition Cost (CAC)"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=CAC"
        ws.cell(row=row, column=2).number_format = '₩#,##0'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # LTV/CAC Ratio (가장 중요!)
        row += 1
        ws.cell(row=row, column=1).value = "LTV/CAC Ratio"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        
        ratio_cell = f'B{row}'
        ws.cell(row=row, column=2).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).font = Font(size=18, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        # Traffic Light (LTV/CAC)
        excellent_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        excellent_font = Font(color="FFFFFF", bold=True, size=18)
        excellent_rule = FormulaRule(
            formula=[f'{ratio_cell}>=5'],
            stopIfTrue=True,
            fill=excellent_fill,
            font=excellent_font
        )
        ws.conditional_formatting.add(ratio_cell, excellent_rule)
        
        good_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        good_font = Font(color="FFFFFF", bold=True, size=18)
        good_rule = FormulaRule(
            formula=[f'AND({ratio_cell}>=3, {ratio_cell}<5)'],
            stopIfTrue=True,
            fill=good_fill,
            font=good_font
        )
        ws.conditional_formatting.add(ratio_cell, good_rule)
        
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_font = Font(color="000000", bold=True, size=18)
        warning_rule = FormulaRule(
            formula=[f'AND({ratio_cell}>=1.5, {ratio_cell}<3)'],
            stopIfTrue=True,
            fill=warning_fill,
            font=warning_font
        )
        ws.conditional_formatting.add(ratio_cell, warning_rule)
        
        poor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        poor_font = Font(color="FFFFFF", bold=True, size=18)
        poor_rule = FormulaRule(
            formula=[f'{ratio_cell}<1.5'],
            stopIfTrue=True,
            fill=poor_fill,
            font=poor_font
        )
        ws.conditional_formatting.add(ratio_cell, poor_rule)
        
        # 평가
        ws.cell(row=row, column=3).value = (
            f'=IF(LTV_CAC_Ratio>=5, "우수", '
            f'IF(LTV_CAC_Ratio>=3, "양호", '
            f'IF(LTV_CAC_Ratio>=1.5, "주의", "위험")))'
        )
        ws.cell(row=row, column=3).font = Font(size=11, bold=True)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        # Payback Period
        row += 1
        ws.cell(row=row, column=1).value = "CAC Payback Period"
        ws.cell(row=row, column=1).font = Font(size=11)
        
        ws.cell(row=row, column=2).value = "=PaybackPeriod"
        ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws.cell(row=row, column=3).value = "개월"
        ws.cell(row=row, column=3).font = Font(size=10)
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        # === 3. 종합 건강도 ===
        row += 2
        ws.cell(row=row, column=1).value = "🏥 비즈니스 건강도"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "종합 평가:"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # 2가지 지표 모두 통과 확인
        ws.cell(row=row, column=2).value = (
            f'=IF(AND(LTV_CAC_Ratio>=3, PaybackPeriod<=12), '
            f'"✅ 건강한 비즈니스", '
            f'IF(OR(LTV_CAC_Ratio<1.5, PaybackPeriod>18), '
            f'"❌ 비즈니스 모델 재검토", "⚠️ 개선 필요"))'
        )
        ws.cell(row=row, column=2).font = Font(size=11, bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
        ws.merge_cells(f'B{row}:E{row}')
        
        # === 4. 핵심 권장사항 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 핵심 권장사항"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "1. LTV 개선:"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=2).value = (
            f'=IF(MonthlyChurn>0.05, "Churn 감소 필요 (현재 "&TEXT(MonthlyChurn,"0.0%")&")", '
            f'"Churn 양호 ✅")'
        )
        ws.cell(row=row, column=2).font = Font(size=9)
        ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "2. CAC 최적화:"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=2).value = (
            f'=IF(PaybackPeriod>12, "마케팅 효율화 필요 (Payback "&TEXT(PaybackPeriod,"0.0")&"개월)", '
            f'"마케팅 효율 양호 ✅")'
        )
        ws.cell(row=row, column=2).font = Font(size=9)
        ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "3. Sensitivity:"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        ws.cell(row=row, column=2).value = "Sensitivity_Analysis 시트에서 가장 중요한 변수 확인"
        ws.cell(row=row, column=2).font = Font(size=9)
        ws.merge_cells(f'B{row}:E{row}')
        
        # === 5. 다음 액션 ===
        row += 2
        ws.cell(row=row, column=1).value = "📋 다음 액션"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        actions = [
            "1. Excel 전체 시트 검토 (Inputs → LTV → CAC → Ratio → Payback → Sensitivity → Scenarios)",
            "2. 실제 데이터로 검증 (현재는 테스트 데이터)",
            "3. Cohort_LTV 시트에 실제 코호트 데이터 입력",
            "4. Benchmark_Comparison에서 업계 대비 포지셔닝 확인",
            "5. Scenarios 시트에서 최악/최선 시나리오 확인"
        ]
        
        for action in actions:
            row += 1
            ws.cell(row=row, column=1).value = action
            ws.cell(row=row, column=1).font = Font(size=9)
            ws.merge_cells(f'A{row}:E{row}')
        
        # === 6. 시트 참조 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "📊 상세 분석 시트"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True, color="666666")
        
        sheets_guide = [
            "• Inputs: 핵심 지표 입력 (노란색 셀만 수정)",
            "• LTV_Calculation: LTV 계산 상세 (2가지 방법)",
            "• CAC_Analysis: CAC 계산 상세 (채널별 분석)",
            "• LTV_CAC_Ratio: 비율 분석 + Traffic Light",
            "• Payback_Period: 회수 기간 + 월별 Timeline",
            "• Sensitivity_Analysis: 변수별 영향도 + 2-Way Matrix",
            "• UE_Scenarios: 3가지 시나리오 비교",
            "• Cohort_LTV: 코호트 개선 추적",
            "• Benchmark_Comparison: 업계 벤치마크 비교"
        ]
        
        for guide in sheets_guide:
            row += 1
            ws.cell(row=row, column=1).value = guide
            ws.cell(row=row, column=1).font = Font(size=9, color="666666")
            ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ Dashboard 시트 생성 완료")
        print(f"      - 핵심 지표 Big Numbers")
        print(f"      - Traffic Light (자동 색상)")
        print(f"      - 권장사항 + 다음 액션")


# 테스트는 별도 스크립트에서

