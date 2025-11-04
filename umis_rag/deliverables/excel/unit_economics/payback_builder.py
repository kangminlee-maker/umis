"""
Payback Period Sheet Builder
CAC 회수 기간 분석 시트

Sheet 5: Payback_Period
- CAC Payback Period 계산
- 월별 Cash Flow Timeline
- 목표 대비 평가 (< 12개월)
- 누적 Cash Flow
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

from ..formula_engine import FormulaEngine, ExcelStyles


class PaybackBuilder:
    """
    Payback Period 시트 빌더
    
    기능:
      - Payback Period 계산
      - 월별 Cash Flow
      - 누적 Cash Flow
      - 목표 대비 평가
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """Payback Period 시트 생성"""
        
        ws = self.wb.create_sheet("Payback_Period")
        
        # === 1. 제목 ===
        ws['A1'] = "CAC Payback Period Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "고객 획득 비용(CAC)을 회수하는 데 걸리는 시간"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:E2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        
        # === 2. Payback Period 계산 ===
        row = 4
        ws.cell(row=row, column=1).value = "Payback Period 계산"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Formula:"
        ws.cell(row=row, column=2).value = "Payback = CAC / (ARPU × Gross Margin)"
        ws.cell(row=row, column=2).font = Font(size=10, italic=True)
        ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "CAC (Customer Acquisition Cost)"
        ws.cell(row=row, column=2).value = "=CAC"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        row += 1
        ws.cell(row=row, column=1).value = "ARPU (월)"
        ws.cell(row=row, column=2).value = "=ARPU"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=2).value = "=GrossMargin"
        ws.cell(row=row, column=2).number_format = '0.0%'
        ws.cell(row=row, column=3).value = "%"
        
        row += 1
        ws.cell(row=row, column=1).value = "월별 Contribution Margin"
        ws.cell(row=row, column=2).value = "=ARPU*GrossMargin"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원/월"
        ws.cell(row=row, column=4).value = "고객 1명이 매월 창출하는 순수익"
        ws.merge_cells(f'D{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Payback Period"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        
        # Payback 계산
        payback_formula = self.fe.create_payback_formula('CAC', 'ARPU', 'GrossMargin')
        ws.cell(row=row, column=2).value = payback_formula
        ws.cell(row=row, column=2).number_format = '0.0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=12, bold=True)
        ws.cell(row=row, column=3).value = "개월"
        
        # Named Range
        payback_cell = f'B{row}'
        self.fe.define_named_range('PaybackPeriod', 'Payback_Period', payback_cell)
        
        # === 3. 평가 ===
        row += 1
        ws.cell(row=row, column=1).value = "평가"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        evaluation_formula = (
            f'=IF(PaybackPeriod<=6, "우수 (< 6개월)", '
            f'IF(PaybackPeriod<=12, "양호 (< 12개월)", '
            f'IF(PaybackPeriod<=18, "주의 (< 18개월)", "위험 (> 18개월)")))'
        )
        ws.cell(row=row, column=2).value = evaluation_formula
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        # 조건부 서식
        # < 6개월: 진한 녹색
        excellent_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        excellent_font = Font(color="FFFFFF", bold=True)
        excellent_rule = FormulaRule(
            formula=[f'{payback_cell}<=6'],
            stopIfTrue=True,
            fill=excellent_fill,
            font=excellent_font
        )
        ws.conditional_formatting.add(payback_cell, excellent_rule)
        
        # 6-12개월: 녹색
        good_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        good_rule = FormulaRule(
            formula=[f'AND({payback_cell}>6, {payback_cell}<=12)'],
            stopIfTrue=True,
            fill=good_fill
        )
        ws.conditional_formatting.add(payback_cell, good_rule)
        
        # 12-18개월: 노란색
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_rule = FormulaRule(
            formula=[f'AND({payback_cell}>12, {payback_cell}<=18)'],
            stopIfTrue=True,
            fill=warning_fill
        )
        ws.conditional_formatting.add(payback_cell, warning_rule)
        
        # > 18개월: 빨간색
        poor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        poor_font = Font(color="FFFFFF", bold=True)
        poor_rule = FormulaRule(
            formula=[f'{payback_cell}>18'],
            stopIfTrue=True,
            fill=poor_fill,
            font=poor_font
        )
        ws.conditional_formatting.add(payback_cell, poor_rule)
        
        # === 4. 월별 Cash Flow Timeline ===
        row += 2
        ws.cell(row=row, column=1).value = "월별 Cash Flow Timeline"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        # 헤더
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        ws.cell(row=row, column=1).value = "Month"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "Monthly CF"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "Cumulative CF"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        
        ws.cell(row=row, column=4).value = "Status"
        ws.cell(row=row, column=4).font = header_font
        ws.cell(row=row, column=4).fill = header_fill
        
        # Month 0: CAC 지출
        row += 1
        ws.cell(row=row, column=1).value = 0
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).value = "=-CAC"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "=-CAC"
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4).value = "고객 획득 (CAC 지출)"
        
        # Month 1-24
        for month in range(1, 25):
            row += 1
            ws.cell(row=row, column=1).value = month
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            
            # Monthly CF = ARPU × Margin
            ws.cell(row=row, column=2).value = "=ARPU*GrossMargin"
            ws.cell(row=row, column=2).number_format = '#,##0'
            
            # Cumulative CF
            prev_cumulative = f'C{row-1}'
            ws.cell(row=row, column=3).value = f"={prev_cumulative}+B{row}"
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            # Status
            ws.cell(row=row, column=4).value = f'=IF(C{row}>=0, "✅ Payback 완료", "진행 중")'
            
            # 조건부 서식 (누적 CF가 0 이상이면 녹색)
            positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            positive_rule = FormulaRule(
                formula=[f'C{row}>=0'],
                stopIfTrue=True,
                fill=positive_fill
            )
            ws.conditional_formatting.add(f'C{row}', positive_rule)
        
        # === 5. 업계 벤치마크 ===
        row += 2
        ws.cell(row=row, column=1).value = "업계 벤치마크"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        benchmarks = [
            {'level': 'Best-in-Class', 'payback': '< 6개월', 'color': '00B050'},
            {'level': 'Good (목표)', 'payback': '< 12개월', 'color': '92D050'},
            {'level': 'Acceptable', 'payback': '12-18개월', 'color': 'FFC000'},
            {'level': 'Poor', 'payback': '> 18개월', 'color': 'FF0000'},
        ]
        
        for benchmark in benchmarks:
            row += 1
            ws.cell(row=row, column=1).value = benchmark['level']
            ws.cell(row=row, column=1).font = Font(size=9, bold=True)
            
            ws.cell(row=row, column=2).value = benchmark['payback']
            ws.cell(row=row, column=2).font = Font(size=9)
            
            # 색상 표시
            color_fill = PatternFill(start_color=benchmark['color'], end_color=benchmark['color'], fill_type="solid")
            ws.cell(row=row, column=1).fill = color_fill
            if benchmark['color'] in ['00B050', 'FF0000']:
                ws.cell(row=row, column=1).font = Font(size=9, bold=True, color="FFFFFF")
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• Payback Period = CAC를 회수하는 데 걸리는 시간"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 짧을수록 좋음: 빠른 회수 = 현금 흐름 개선"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• LTV/CAC와 함께 평가: 높은 비율 + 짧은 Payback = 이상적"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:E{row}')
        
        print(f"   ✅ Payback Period 시트 생성 완료")
        print(f"      - 월별 Cash Flow Timeline (24개월)")
        print(f"      - Named Range: PaybackPeriod")


# 테스트는 별도 스크립트에서

