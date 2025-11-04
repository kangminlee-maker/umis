"""
CAC Analysis Sheet Builder
고객 획득 비용 (CAC) 분석 시트

Sheet 3: CAC Analysis
- Total S&M Spend
- New Customers Acquired
- CAC = Total Spend / New Customers
- CAC by Channel (선택)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class CACBuilder:
    """
    CAC Analysis 시트 빌더
    
    기능:
      - CAC 계산
      - 채널별 CAC (선택)
      - CAC 벤치마크 비교
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, channels_data: list = None) -> None:
        """
        CAC Analysis 시트 생성
        
        Args:
            channels_data: 채널별 데이터 (선택)
                [
                    {'channel': '검색 광고', 'spend': 3000000, 'customers': 100},
                    {'channel': 'SNS 광고', 'spend': 2000000, 'customers': 80},
                    ...
                ]
        """
        
        ws = self.wb.create_sheet("CAC_Analysis")
        
        # === 1. 제목 ===
        ws['A1'] = "CAC (Customer Acquisition Cost) Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "고객 획득 비용 = 고객 1명을 획득하는 데 드는 마케팅 비용"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        
        # === 2. 전체 CAC 계산 ===
        row = 4
        ws.cell(row=row, column=1).value = "전체 CAC 계산"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Formula:"
        ws.cell(row=row, column=2).value = "CAC = Total S&M Spend / New Customers"
        ws.cell(row=row, column=2).font = Font(size=10, italic=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Total S&M Spend (Monthly)"
        ws.cell(row=row, column=2).value = "=SMSpend"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "Inputs 시트에서 참조"
        
        row += 1
        ws.cell(row=row, column=1).value = "New Customers (Monthly)"
        ws.cell(row=row, column=2).value = "=NewCustomers"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "명"
        ws.cell(row=row, column=4).value = "Inputs 시트에서 참조"
        
        row += 1
        ws.cell(row=row, column=1).value = "CAC (전체)"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        # CAC 계산
        cac_formula = self.fe.create_cac_formula('SMSpend', 'NewCustomers')
        ws.cell(row=row, column=2).value = cac_formula
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=2).fill = PatternFill(start_color=ExcelStyles.RESULT_FILL, end_color=ExcelStyles.RESULT_FILL, fill_type="solid")
        ws.cell(row=row, column=2).font = Font(size=11, bold=True)
        ws.cell(row=row, column=3).value = "원"
        
        # Named Range for CAC_Calculated
        self.fe.define_named_range('CAC_Calculated', 'CAC_Analysis', f'B{row}')
        
        # === 3. CAC 비교 (Inputs vs Calculated) ===
        row += 2
        ws.cell(row=row, column=1).value = "CAC 비교"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "CAC (Inputs 시트)"
        ws.cell(row=row, column=2).value = "=CAC"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "직접 입력한 값"
        
        row += 1
        ws.cell(row=row, column=1).value = "CAC (계산)"
        ws.cell(row=row, column=2).value = "=CAC_Calculated"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "S&M Spend로 역산"
        
        row += 1
        ws.cell(row=row, column=1).value = "차이 (%)"
        ws.cell(row=row, column=2).value = "=IFERROR((CAC_Calculated-CAC)/CAC*100, 0)"
        ws.cell(row=row, column=2).number_format = '0.0"%"'
        ws.cell(row=row, column=3).value = "%"
        ws.cell(row=row, column=4).value = "10% 이내 권장"
        
        # === 4. 채널별 CAC (선택) ===
        if channels_data and len(channels_data) > 0:
            row += 2
            ws.cell(row=row, column=1).value = "채널별 CAC 분석"
            ws.cell(row=row, column=1).font = Font(size=11, bold=True)
            ws.merge_cells(f'A{row}:D{row}')
            
            # 헤더
            row += 1
            header_font = Font(size=10, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
            
            ws.cell(row=row, column=1).value = "채널"
            ws.cell(row=row, column=1).font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            
            ws.cell(row=row, column=2).value = "지출"
            ws.cell(row=row, column=2).font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            
            ws.cell(row=row, column=3).value = "신규 고객"
            ws.cell(row=row, column=3).font = header_font
            ws.cell(row=row, column=3).fill = header_fill
            
            ws.cell(row=row, column=4).value = "CAC"
            ws.cell(row=row, column=4).font = header_font
            ws.cell(row=row, column=4).fill = header_fill
            
            # 데이터
            input_fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
            
            for channel in channels_data:
                row += 1
                ws.cell(row=row, column=1).value = channel['channel']
                ws.cell(row=row, column=1).font = Font(size=10)
                
                ws.cell(row=row, column=2).value = channel['spend']
                ws.cell(row=row, column=2).fill = input_fill
                ws.cell(row=row, column=2).number_format = '#,##0'
                
                ws.cell(row=row, column=3).value = channel['customers']
                ws.cell(row=row, column=3).fill = input_fill
                ws.cell(row=row, column=3).number_format = '#,##0'
                
                # CAC 계산
                ws.cell(row=row, column=4).value = f"=IFERROR(B{row}/C{row}, 0)"
                ws.cell(row=row, column=4).number_format = '#,##0'
        
        # === 5. 업계 벤치마크 ===
        row += 2
        ws.cell(row=row, column=1).value = "업계 벤치마크 (참고)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        benchmarks = [
            {'industry': 'SaaS (B2B)', 'cac': '₩300,000 - ₩1,000,000'},
            {'industry': 'SaaS (SMB)', 'cac': '₩100,000 - ₩300,000'},
            {'industry': 'E-commerce', 'cac': '₩20,000 - ₩50,000'},
            {'industry': '구독 서비스', 'cac': '₩15,000 - ₩40,000'},
        ]
        
        for benchmark in benchmarks:
            row += 1
            ws.cell(row=row, column=1).value = benchmark['industry']
            ws.cell(row=row, column=1).font = Font(size=9, color="666666")
            ws.cell(row=row, column=2).value = benchmark['cac']
            ws.cell(row=row, column=2).font = Font(size=9, color="666666")
        
        # === 6. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• CAC = 고객 1명을 획득하는 데 드는 총 마케팅 비용 (광고, 프로모션, 영업)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• CAC가 낮을수록 좋지만, 채널별로 CAC가 다를 수 있음 (검색 < SNS < 오프라인)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• LTV/CAC > 3.0 이상이면 건강한 비즈니스 (다음 시트에서 확인)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ CAC Analysis 시트 생성 완료")
        print(f"      - Named Range: CAC_Calculated")
        if channels_data:
            print(f"      - 채널별 CAC: {len(channels_data)}개 채널")


# 테스트는 별도 스크립트에서
# python scripts/test_unit_economics.py

