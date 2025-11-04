"""
Sensitivity Analysis Sheet Builder
민감도 분석 시트

Sheet 7: Sensitivity_Analysis
- 단일 변수 민감도 (ARPU, CAC, Churn ±20%)
- 2-Way Sensitivity Matrix (ARPU × Churn)
- Tornado Chart (영향도 순위)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..formula_engine import FormulaEngine, ExcelStyles


class SensitivityBuilder:
    """
    Sensitivity Analysis 시트 빌더
    
    기능:
      - 단일 변수 민감도
      - 2-Way Sensitivity Matrix
      - 영향도 순위
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """Sensitivity Analysis 시트 생성"""
        
        ws = self.wb.create_sheet("Sensitivity_Analysis")
        
        # === 1. 제목 ===
        ws['A1'] = "Sensitivity Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "핵심 변수의 변화가 LTV/CAC 비율에 미치는 영향"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:F2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        # === 2. 단일 변수 민감도 ===
        row = 4
        ws.cell(row=row, column=1).value = "1. 단일 변수 민감도 (±20%)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        # 헤더
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        headers = ['Variable', '-20%', '-10%', 'Base', '+10%', '+20%']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # === ARPU 민감도 ===
        row += 1
        ws.cell(row=row, column=1).value = "ARPU (Revenue Impact)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # -20%
        ws.cell(row=row, column=2).value = "=IFERROR((ARPU*0.8*CustomerLifetime*GrossMargin)/CAC, 0)"
        ws.cell(row=row, column=2).number_format = '0.00'
        
        # -10%
        ws.cell(row=row, column=3).value = "=IFERROR((ARPU*0.9*CustomerLifetime*GrossMargin)/CAC, 0)"
        ws.cell(row=row, column=3).number_format = '0.00'
        
        # Base
        ws.cell(row=row, column=4).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=4).number_format = '0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color=ExcelStyles.CALC_FILL, end_color=ExcelStyles.CALC_FILL, fill_type="solid")
        
        # +10%
        ws.cell(row=row, column=5).value = "=IFERROR((ARPU*1.1*CustomerLifetime*GrossMargin)/CAC, 0)"
        ws.cell(row=row, column=5).number_format = '0.00'
        
        # +20%
        ws.cell(row=row, column=6).value = "=IFERROR((ARPU*1.2*CustomerLifetime*GrossMargin)/CAC, 0)"
        ws.cell(row=row, column=6).number_format = '0.00'
        
        # === CAC 민감도 ===
        row += 1
        ws.cell(row=row, column=1).value = "CAC (Cost Impact)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # CAC가 증가하면 비율은 감소
        ws.cell(row=row, column=2).value = "=IFERROR(LTV/(CAC*1.2), 0)"  # CAC +20% = Ratio 감소
        ws.cell(row=row, column=2).number_format = '0.00'
        
        ws.cell(row=row, column=3).value = "=IFERROR(LTV/(CAC*1.1), 0)"
        ws.cell(row=row, column=3).number_format = '0.00'
        
        ws.cell(row=row, column=4).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=4).number_format = '0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color=ExcelStyles.CALC_FILL, end_color=ExcelStyles.CALC_FILL, fill_type="solid")
        
        ws.cell(row=row, column=5).value = "=IFERROR(LTV/(CAC*0.9), 0)"  # CAC -10% = Ratio 증가
        ws.cell(row=row, column=5).number_format = '0.00'
        
        ws.cell(row=row, column=6).value = "=IFERROR(LTV/(CAC*0.8), 0)"
        ws.cell(row=row, column=6).number_format = '0.00'
        
        # === Churn 민감도 ===
        row += 1
        ws.cell(row=row, column=1).value = "Churn (Retention Impact)"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        # Churn이 증가하면 LTV 감소 → Ratio 감소
        ws.cell(row=row, column=2).value = "=IFERROR((ARPU*GrossMargin/(MonthlyChurn*1.2))/CAC, 0)"
        ws.cell(row=row, column=2).number_format = '0.00'
        
        ws.cell(row=row, column=3).value = "=IFERROR((ARPU*GrossMargin/(MonthlyChurn*1.1))/CAC, 0)"
        ws.cell(row=row, column=3).number_format = '0.00'
        
        ws.cell(row=row, column=4).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=4).number_format = '0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color=ExcelStyles.CALC_FILL, end_color=ExcelStyles.CALC_FILL, fill_type="solid")
        
        ws.cell(row=row, column=5).value = "=IFERROR((ARPU*GrossMargin/(MonthlyChurn*0.9))/CAC, 0)"
        ws.cell(row=row, column=5).number_format = '0.00'
        
        ws.cell(row=row, column=6).value = "=IFERROR((ARPU*GrossMargin/(MonthlyChurn*0.8))/CAC, 0)"
        ws.cell(row=row, column=6).number_format = '0.00'
        
        # === Gross Margin 민감도 ===
        row += 1
        ws.cell(row=row, column=1).value = "Gross Margin"
        ws.cell(row=row, column=1).font = Font(size=10)
        
        ws.cell(row=row, column=2).value = "=IFERROR((ARPU*CustomerLifetime*GrossMargin*0.8)/CAC, 0)"
        ws.cell(row=row, column=2).number_format = '0.00'
        
        ws.cell(row=row, column=3).value = "=IFERROR((ARPU*CustomerLifetime*GrossMargin*0.9)/CAC, 0)"
        ws.cell(row=row, column=3).number_format = '0.00'
        
        ws.cell(row=row, column=4).value = "=LTV_CAC_Ratio"
        ws.cell(row=row, column=4).number_format = '0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color=ExcelStyles.CALC_FILL, end_color=ExcelStyles.CALC_FILL, fill_type="solid")
        
        ws.cell(row=row, column=5).value = "=IFERROR((ARPU*CustomerLifetime*GrossMargin*1.1)/CAC, 0)"
        ws.cell(row=row, column=5).number_format = '0.00'
        
        ws.cell(row=row, column=6).value = "=IFERROR((ARPU*CustomerLifetime*GrossMargin*1.2)/CAC, 0)"
        ws.cell(row=row, column=6).number_format = '0.00'
        
        # === 3. 2-Way Sensitivity Matrix (ARPU × Churn) ===
        row += 2
        ws.cell(row=row, column=1).value = "2. 2-Way Sensitivity Matrix: ARPU × Churn"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Churn → / ARPU ↓"
        ws.cell(row=row, column=1).font = Font(size=9, italic=True, color="666666")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # ARPU 변화 헤더 (-20%, -10%, Base, +10%, +20%)
        arpu_changes = [0.8, 0.9, 1.0, 1.1, 1.2]
        arpu_labels = ['-20%', '-10%', 'Base', '+10%', '+20%']
        
        for col_idx, label in enumerate(arpu_labels, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = label
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
        
        # Churn 변화 행 (-20%, -10%, Base, +10%, +20%)
        churn_changes = [0.8, 0.9, 1.0, 1.1, 1.2]
        churn_labels = ['-20%', '-10%', 'Base', '+10%', '+20%']
        
        for idx, (churn_mult, churn_label) in enumerate(zip(churn_changes, churn_labels)):
            row += 1
            
            # A열: Churn 변화
            ws.cell(row=row, column=1).value = churn_label
            ws.cell(row=row, column=1).font = Font(size=9, bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=1).fill = header_fill
            
            # B-F열: ARPU × Churn 조합
            for col_idx, arpu_mult in enumerate(arpu_changes, start=2):
                # LTV/CAC = (ARPU * arpu_mult * Margin / (Churn * churn_mult)) / CAC
                formula = f"=IFERROR((ARPU*{arpu_mult}*GrossMargin/(MonthlyChurn*{churn_mult}))/CAC, 0)"
                ws.cell(row=row, column=col_idx).value = formula
                ws.cell(row=row, column=col_idx).number_format = '0.00'
                ws.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
                
                # Base Case 강조
                if arpu_mult == 1.0 and churn_mult == 1.0:
                    ws.cell(row=row, column=col_idx).font = Font(bold=True)
                    ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        # === 4. 영향도 분석 ===
        row += 2
        ws.cell(row=row, column=1).value = "3. 영향도 순위 (Impact Ranking)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        # 헤더
        ws.cell(row=row, column=1).value = "Variable"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "Change"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "LTV/CAC Impact"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'C{row}:D{row}')
        
        ws.cell(row=row, column=5).value = "% Change"
        ws.cell(row=row, column=5).font = header_font
        ws.cell(row=row, column=5).fill = header_fill
        
        # 각 변수의 +20% 영향
        variables = [
            {
                'name': 'ARPU',
                'base_ratio': '=LTV_CAC_Ratio',
                'new_ratio': '=IFERROR((ARPU*1.2*CustomerLifetime*GrossMargin)/CAC, 0)',
                'start_row': row + 1
            },
            {
                'name': 'CAC',
                'base_ratio': '=LTV_CAC_Ratio',
                'new_ratio': '=IFERROR(LTV/(CAC*1.2), 0)',
                'start_row': row + 2
            },
            {
                'name': 'Churn',
                'base_ratio': '=LTV_CAC_Ratio',
                'new_ratio': '=IFERROR((ARPU*GrossMargin/(MonthlyChurn*1.2))/CAC, 0)',
                'start_row': row + 3
            },
            {
                'name': 'Gross Margin',
                'base_ratio': '=LTV_CAC_Ratio',
                'new_ratio': '=IFERROR((ARPU*CustomerLifetime*GrossMargin*1.2)/CAC, 0)',
                'start_row': row + 4
            }
        ]
        
        for var in variables:
            row += 1
            
            # Variable
            ws.cell(row=row, column=1).value = var['name']
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # Change
            ws.cell(row=row, column=2).value = "+20%"
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # New Ratio
            ws.cell(row=row, column=3).value = var['new_ratio']
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.merge_cells(f'C{row}:D{row}')
            
            # % Change
            ws.cell(row=row, column=5).value = f"=IFERROR((C{row}-LTV_CAC_Ratio)/LTV_CAC_Ratio*100, 0)"
            ws.cell(row=row, column=5).number_format = '0.0"%"'
        
        # === 5. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• % Change가 큰 변수 = 가장 중요한 변수 (집중 관리 필요)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 일반적으로 ARPU와 Churn이 가장 큰 영향 (각 20%)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 2-Way Matrix에서 최악/최선 시나리오 확인 가능"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:F{row}')
        
        print(f"   ✅ Sensitivity Analysis 시트 생성 완료")
        print(f"      - 단일 변수 민감도 (4개 변수)")
        print(f"      - 2-Way Matrix (ARPU × Churn)")
        print(f"      - 영향도 순위")


# 테스트는 별도 스크립트에서

