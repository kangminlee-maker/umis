"""
Unit Economics Inputs Sheet Builder
입력 데이터 시트 생성

Sheet 1: Inputs
- ARPU (Average Revenue Per User)
- CAC (Customer Acquisition Cost)
- Gross Margin (%)
- Monthly Churn Rate (%)
- Customer Lifetime (months)
- Total S&M Spend (월별)
- New Customers (월별)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..formula_engine import FormulaEngine, ExcelStyles


class InputsBuilder:
    """
    Unit Economics Inputs 시트 빌더
    
    기능:
      - 핵심 지표 입력 (ARPU, CAC, Churn, Margin)
      - Named Range 정의
      - 입력 가이드 제공
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self, inputs_data: dict = None) -> None:
        """
        Inputs 시트 생성
        
        Args:
            inputs_data: 입력 데이터 딕셔너리
                {
                    'arpu': 9000,
                    'cac': 25000,
                    'gross_margin': 0.35,
                    'monthly_churn': 0.04,
                    'customer_lifetime': 25,
                    'sm_spend_monthly': 5000000,
                    'new_customers_monthly': 200
                }
        """
        
        # 기본값
        if inputs_data is None:
            inputs_data = {
                'arpu': 10000,
                'cac': 30000,
                'gross_margin': 0.40,
                'monthly_churn': 0.05,
                'customer_lifetime': 20,
                'sm_spend_monthly': 10000000,
                'new_customers_monthly': 300
            }
        
        ws = self.wb.create_sheet("Inputs", 0)  # 첫 번째 시트
        
        # === 1. 제목 ===
        ws['A1'] = "Unit Economics Inputs"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "핵심 지표 입력 (노란색 셀만 수정)"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # === 2. 컬럼 헤더 ===
        headers = ['Metric', 'Value', 'Unit', 'Description']
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 40
        
        # === 3. 핵심 지표 입력 ===
        input_fill = PatternFill(start_color=ExcelStyles.INPUT_FILL, end_color=ExcelStyles.INPUT_FILL, fill_type="solid")
        
        metrics = [
            {
                'name': 'ARPU',
                'label': 'ARPU (Average Revenue Per User)',
                'value': inputs_data.get('arpu', 10000),
                'unit': '원/월',
                'description': '고객 1명당 평균 월 매출',
                'cell': 'B5'
            },
            {
                'name': 'CAC',
                'label': 'CAC (Customer Acquisition Cost)',
                'value': inputs_data.get('cac', 30000),
                'unit': '원',
                'description': '고객 1명 획득 비용',
                'cell': 'B6'
            },
            {
                'name': 'GrossMargin',
                'label': 'Gross Margin',
                'value': inputs_data.get('gross_margin', 0.40),
                'unit': '%',
                'description': '매출총이익률 (Revenue - COGS) / Revenue',
                'cell': 'B7'
            },
            {
                'name': 'MonthlyChurn',
                'label': 'Monthly Churn Rate',
                'value': inputs_data.get('monthly_churn', 0.05),
                'unit': '%',
                'description': '월별 고객 이탈률',
                'cell': 'B8'
            },
            {
                'name': 'CustomerLifetime',
                'label': 'Customer Lifetime',
                'value': inputs_data.get('customer_lifetime', 20),
                'unit': 'months',
                'description': '평균 고객 생애 (개월)',
                'cell': 'B9'
            },
        ]
        
        row = 5
        for metric in metrics:
            # A열: Metric 이름
            ws.cell(row=row, column=1).value = metric['label']
            ws.cell(row=row, column=1).font = Font(size=10)
            
            # B열: Value (입력)
            ws.cell(row=row, column=2).value = metric['value']
            ws.cell(row=row, column=2).fill = input_fill
            ws.cell(row=row, column=2).font = Font(size=10, bold=True)
            
            # 숫자 포맷
            if metric['unit'] == '%':
                ws.cell(row=row, column=2).number_format = '0.0%'
            elif metric['unit'] in ['원', '원/월']:
                ws.cell(row=row, column=2).number_format = '#,##0'
            else:
                ws.cell(row=row, column=2).number_format = '#,##0.0'
            
            # C열: Unit
            ws.cell(row=row, column=3).value = metric['unit']
            ws.cell(row=row, column=3).font = Font(size=9, color="666666")
            
            # D열: Description
            ws.cell(row=row, column=4).value = metric['description']
            ws.cell(row=row, column=4).font = Font(size=9)
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            
            # Named Range 정의
            self.fe.define_named_range(
                name=metric['name'],
                sheet='Inputs',
                cell=metric['cell']
            )
            
            row += 1
        
        # === 4. 월별 S&M 데이터 (선택) ===
        row += 1
        ws.cell(row=row, column=1).value = "Monthly S&M Data (Optional)"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "Total S&M Spend (Monthly)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = inputs_data.get('sm_spend_monthly', 10000000)
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = '원'
        ws.cell(row=row, column=4).value = 'Sales & Marketing 월별 총 지출'
        
        # Named Range
        self.fe.define_named_range('SMSpend', 'Inputs', f'B{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "New Customers (Monthly)"
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).value = inputs_data.get('new_customers_monthly', 300)
        ws.cell(row=row, column=2).fill = input_fill
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = '명'
        ws.cell(row=row, column=4).value = '월별 신규 고객 수'
        
        # Named Range
        self.fe.define_named_range('NewCustomers', 'Inputs', f'B{row}')
        
        # === 5. 입력 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "📋 입력 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• ARPU: 고객 1명의 평균 월 매출 (구독료, 평균 구매액 등)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• CAC: 마케팅 비용 / 신규 고객 수 (광고, 프로모션, 영업 비용)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Gross Margin: (매출 - 원가) / 매출 (라이선스료, COGS 제외 후)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Churn: 당월 해지 고객 / 전월 고객 수 (월별 이탈률)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• Lifetime: 1 / Monthly Churn 또는 실제 평균 생애 (개월)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ Inputs 시트 생성 완료")
        print(f"      - 7개 Named Range 정의 (ARPU, CAC, GrossMargin, MonthlyChurn, CustomerLifetime, SMSpend, NewCustomers)")


# 테스트는 별도 스크립트에서
# python scripts/test_unit_economics.py

