"""
LTV/CAC Ratio Sheet Builder
LTV/CAC 비율 분석 시트

Sheet 4: LTV_CAC_Ratio
- LTV/CAC 비율 계산
- 업계 벤치마크 비교 (3.0, 5.0)
- Traffic Light (조건부 서식)
- 해석 가이드
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule

from ..formula_engine import FormulaEngine, ExcelStyles


class RatioBuilder:
    """
    LTV/CAC Ratio 시트 빌더
    
    기능:
      - LTV/CAC 비율 계산
      - 벤치마크 비교
      - Traffic Light (조건부 서식)
      - 개선 권장사항
    """
    
    def __init__(self, workbook: Workbook, formula_engine: FormulaEngine):
        """
        Args:
            workbook: openpyxl Workbook
            formula_engine: FormulaEngine 인스턴스
        """
        self.wb = workbook
        self.fe = formula_engine
    
    def create_sheet(self) -> None:
        """LTV/CAC Ratio 시트 생성"""
        
        ws = self.wb.create_sheet("LTV_CAC_Ratio")
        
        # === 1. 제목 ===
        ws['A1'] = "LTV/CAC Ratio Analysis"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "고객 생애 가치 / 고객 획득 비용 = 사업 건강도 핵심 지표"
        ws['A2'].font = Font(size=10, italic=True, color="666666")
        ws.merge_cells('A2:D2')
        
        # 컬럼 폭
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        
        # === 2. LTV/CAC 비율 계산 ===
        row = 4
        ws.cell(row=row, column=1).value = "비율 계산"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV (Customer Lifetime Value)"
        ws.cell(row=row, column=2).value = "=LTV"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "LTV_Calculation 시트에서 참조"
        
        row += 1
        ws.cell(row=row, column=1).value = "CAC (Customer Acquisition Cost)"
        ws.cell(row=row, column=2).value = "=CAC"
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3).value = "원"
        ws.cell(row=row, column=4).value = "Inputs 시트에서 참조"
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV/CAC Ratio"
        ws.cell(row=row, column=1).font = Font(size=12, bold=True)
        
        # 비율 계산
        ratio_formula = self.fe.create_ratio_formula('LTV', 'CAC')
        ws.cell(row=row, column=2).value = ratio_formula
        ws.cell(row=row, column=2).number_format = '0.00'
        ws.cell(row=row, column=2).font = Font(size=14, bold=True)
        ws.cell(row=row, column=3).value = "배"
        
        # Named Range for Ratio
        ratio_cell = f'B{row}'
        self.fe.define_named_range('LTV_CAC_Ratio', 'LTV_CAC_Ratio', ratio_cell)
        
        # === 3. Traffic Light (조건부 서식) ===
        # Excellent (> 5.0): 진한 녹색
        excellent_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        excellent_font = Font(color="FFFFFF", bold=True, size=14)
        excellent_rule = FormulaRule(
            formula=[f'B{row}>=5'],
            stopIfTrue=True,
            fill=excellent_fill,
            font=excellent_font
        )
        ws.conditional_formatting.add(f'B{row}', excellent_rule)
        
        # Good (3.0 - 5.0): 녹색
        good_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        good_font = Font(color="FFFFFF", bold=True, size=14)
        good_rule = FormulaRule(
            formula=[f'AND(B{row}>=3, B{row}<5)'],
            stopIfTrue=True,
            fill=good_fill,
            font=good_font
        )
        ws.conditional_formatting.add(f'B{row}', good_rule)
        
        # Warning (1.5 - 3.0): 노란색
        warning_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        warning_font = Font(color="000000", bold=True, size=14)
        warning_rule = FormulaRule(
            formula=[f'AND(B{row}>=1.5, B{row}<3)'],
            stopIfTrue=True,
            fill=warning_fill,
            font=warning_font
        )
        ws.conditional_formatting.add(f'B{row}', warning_rule)
        
        # Poor (< 1.5): 빨간색
        poor_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        poor_font = Font(color="FFFFFF", bold=True, size=14)
        poor_rule = FormulaRule(
            formula=[f'B{row}<1.5'],
            stopIfTrue=True,
            fill=poor_fill,
            font=poor_font
        )
        ws.conditional_formatting.add(f'B{row}', poor_rule)
        
        # === 4. 평가 (자동) ===
        row += 1
        ws.cell(row=row, column=1).value = "평가"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        evaluation_formula = (
            f'=IF(LTV_CAC_Ratio>=5, "우수 (Excellent)", '
            f'IF(LTV_CAC_Ratio>=3, "양호 (Good)", '
            f'IF(LTV_CAC_Ratio>=1.5, "주의 (Warning)", "위험 (Poor)")))'
        )
        ws.cell(row=row, column=2).value = evaluation_formula
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        
        # === 5. 업계 벤치마크 ===
        row += 2
        ws.cell(row=row, column=1).value = "업계 벤치마크"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        benchmarks = [
            {'level': 'Excellent (우수)', 'ratio': '> 5.0', 'description': '매우 건강한 비즈니스', 'color': '00B050'},
            {'level': 'Good (양호)', 'ratio': '3.0 - 5.0', 'description': '건강한 비즈니스 (목표)', 'color': '92D050'},
            {'level': 'Warning (주의)', 'ratio': '1.5 - 3.0', 'description': '개선 필요', 'color': 'FFC000'},
            {'level': 'Poor (위험)', 'ratio': '< 1.5', 'description': '비즈니스 모델 재검토', 'color': 'FF0000'},
        ]
        
        row += 1
        # 헤더
        header_font = Font(size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color=ExcelStyles.HEADER_FILL, end_color=ExcelStyles.HEADER_FILL, fill_type="solid")
        
        ws.cell(row=row, column=1).value = "등급"
        ws.cell(row=row, column=1).font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        
        ws.cell(row=row, column=2).value = "비율"
        ws.cell(row=row, column=2).font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        ws.cell(row=row, column=3).value = "평가"
        ws.cell(row=row, column=3).font = header_font
        ws.cell(row=row, column=3).fill = header_fill
        ws.merge_cells(f'C{row}:D{row}')
        
        for benchmark in benchmarks:
            row += 1
            ws.cell(row=row, column=1).value = benchmark['level']
            ws.cell(row=row, column=1).font = Font(size=9, bold=True)
            
            ws.cell(row=row, column=2).value = benchmark['ratio']
            ws.cell(row=row, column=2).font = Font(size=9)
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            ws.cell(row=row, column=3).value = benchmark['description']
            ws.cell(row=row, column=3).font = Font(size=9)
            ws.merge_cells(f'C{row}:D{row}')
            
            # 색상 표시
            color_fill = PatternFill(start_color=benchmark['color'], end_color=benchmark['color'], fill_type="solid")
            ws.cell(row=row, column=1).fill = color_fill
            if benchmark['color'] in ['00B050', 'FF0000']:
                ws.cell(row=row, column=1).font = Font(size=9, bold=True, color="FFFFFF")
        
        # === 6. 개선 권장사항 ===
        row += 2
        ws.cell(row=row, column=1).value = "💡 개선 권장사항"
        ws.cell(row=row, column=1).font = Font(size=11, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "LTV/CAC < 3.0 인 경우:"
        ws.cell(row=row, column=1).font = Font(size=9, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        recommendations = [
            "1. LTV 개선: Churn 감소, ARPU 증가, Upsell/Cross-sell",
            "2. CAC 감소: 마케팅 효율화, 채널 최적화, 바이럴 강화",
            "3. Gross Margin 개선: 원가 절감, 가격 인상",
            "4. 비즈니스 모델 재검토: 타겟 고객, 가치 제안 점검"
        ]
        
        for rec in recommendations:
            row += 1
            ws.cell(row=row, column=1).value = rec
            ws.cell(row=row, column=1).font = Font(size=9)
            ws.merge_cells(f'A{row}:D{row}')
        
        # === 7. 해석 가이드 ===
        row += 2
        ws.cell(row=row, column=1).value = "📊 해석 가이드"
        ws.cell(row=row, column=1).font = Font(size=10, bold=True)
        
        row += 1
        ws.cell(row=row, column=1).value = "• LTV/CAC = 고객에게서 얻는 가치 / 획득에 드는 비용"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 3.0 이상: 고객 1명당 획득 비용의 3배 이상 수익"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = "• 비율만으로는 부족 → Payback Period도 함께 확인 (다음 시트)"
        ws.cell(row=row, column=1).font = Font(size=9)
        ws.merge_cells(f'A{row}:D{row}')
        
        print(f"   ✅ LTV/CAC Ratio 시트 생성 완료")
        print(f"      - Traffic Light 조건부 서식 (4단계)")
        print(f"      - Named Range: LTV_CAC_Ratio")


# 테스트는 별도 스크립트에서

