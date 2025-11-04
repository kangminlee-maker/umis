#!/usr/bin/env python3
"""
모든 범위 하드코딩 찾기
B4:B7, C17:C20 같은 패턴을 모두 찾아서 보고
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent

def find_hardcoded_ranges(filepath: Path):
    """
    Excel 파일에서 범위 하드코딩 찾기
    """
    
    print(f"\n🔍 {filepath.name}")
    print("="*70)
    
    wb = load_workbook(filepath, data_only=False)
    
    hardcoded_ranges = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    
                    # 범위 패턴 찾기 (B4:B7, $A$5:$A$10 등)
                    ranges = re.findall(r'\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+', formula)
                    
                    if ranges:
                        hardcoded_ranges.append({
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'formula': formula,
                            'ranges': ranges
                        })
    
    if hardcoded_ranges:
        print(f"\n❌ {len(hardcoded_ranges)}개 범위 하드코딩 발견:\n")
        
        for item in hardcoded_ranges[:20]:  # 최대 20개만
            print(f"{item['sheet']}!{item['cell']}:")
            print(f"  수식: {item['formula']}")
            print(f"  범위: {', '.join(item['ranges'])}")
            print()
        
        if len(hardcoded_ranges) > 20:
            print(f"... 외 {len(hardcoded_ranges) - 20}개")
        
        return False
    else:
        print("\n✅ 범위 하드코딩 없음!")
        return True


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🔍 모든 Excel 파일에서 범위 하드코딩 찾기")
    print("="*70)
    print("\n목적: B4:B7, C17:C20 같은 범위를 모두 찾아서 보고")
    print("이유: Named Range로 바꿔야 함\n")
    
    test_output_dir = project_root / 'test_output'
    
    files = [
        'market_sizing_piano_subscription_20251104.xlsx',
        'unit_economics_music_streaming_20251104.xlsx',
        'financial_projection_korean_adult_education_20251104.xlsx',
    ]
    
    results = {}
    
    for filename in files:
        filepath = test_output_dir / filename
        if filepath.exists():
            results[filename] = find_hardcoded_ranges(filepath)
        else:
            print(f"\n⚠️ {filename}: 파일 없음")
    
    # 최종 결과
    print("\n" + "="*70)
    print("📊 전체 결과")
    print("="*70)
    
    for filename, passed in results.items():
        status = "✅" if passed else "❌"
        print(f"{status} {filename}: {'Clean' if passed else '범위 하드코딩 있음'}")
    
    if all(results.values()):
        print("\n✅ 모든 파일 Clean!")
        print("\n💡 진정한 Named Range 100% 달성")
        sys.exit(0)
    else:
        print(f"\n❌ {sum(1 for r in results.values() if not r)}개 파일에 범위 하드코딩")
        print("\n📋 수정 필요:")
        print("  1. 범위를 Named Range로 전환")
        print("  2. Builder 코드 수정")
        print("  3. 재생성 후 다시 검사")
        sys.exit(1)

