#!/usr/bin/env python3
"""
Named Range 확인
SAM이 어디를 참조하는지 확인
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'

if not filepath.exists():
    print("❌ 파일 없음")
    sys.exit(1)

wb = load_workbook(filepath, data_only=False)

print("\n" + "="*70)
print("🔍 Named Range 상세 확인")
print("="*70 + "\n")

# 주요 Named Range 확인
important_ranges = ['SAM', 'SAM_Method2', 'SAM_Method3', 'SAM_Method4', 'TAM', 'TAM_VALUE']

for range_name in important_ranges:
    if range_name in wb.defined_names:
        defn = wb.defined_names[range_name]
        
        print(f"{range_name}:")
        
        try:
            # destinations는 (sheet_name, cell_address) 튜플의 generator
            for sheet_name, cell_addr in defn.destinations:
                print(f"  → {sheet_name}!{cell_addr}")
                
                # 실제 셀 값 확인
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # $C$6 → C6로 변환
                    cell_addr_clean = cell_addr.replace('$', '')
                    cell = ws[cell_addr_clean]
                    print(f"     값/수식: {cell.value}")
        except Exception as e:
            print(f"  오류: {e}")
        
        print()

# Method_1의 C6 확인
print("="*70)
print("Method_1_TopDown!C6 확인:")
print("-"*70)

if 'Method_1_TopDown' in wb.sheetnames:
    ws = wb['Method_1_TopDown']
    c6 = ws['C6']
    
    print(f"C6 값: {c6.value}")
    print(f"C6 수식: {c6.value if isinstance(c6.value, str) else '숫자'}")
    
    # A5, B5, C5도 확인
    print(f"\nA5: {ws['A5'].value}")
    print(f"B5: {ws['B5'].value}")  
    print(f"C5: {ws['C5'].value}")
    print(f"B6: {ws['B6'].value}")
    print(f"C6: {ws['C6'].value}")
    
    print("\n기대 구조:")
    print("  A5 = =TAM_VALUE (TAM 값)")
    print("  B5 = =FILTER_KOREA (비율 15%)")
    print("  C5 = =FILTER_PIANO (비율 25%)")
    print("  B6 = =A5*B5 (TAM × 15%)")
    print("  C6 = =B6*C5 (B6 × 25% = SAM)")
    print("\n  SAM Named Range → Method_1_TopDown!C6")

