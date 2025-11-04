#!/usr/bin/env python3
"""
Market Sizing Excel 수식 진단
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def diagnose_market_sizing():
    """Market Sizing 예제 파일 진단"""
    
    filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'
    
    if not filepath.exists():
        print(f"❌ 파일 없음: {filepath}")
        return
    
    print("\n" + "="*70)
    print("🔍 Market Sizing 수식 진단")
    print("="*70)
    
    wb = load_workbook(filepath, data_only=False)
    
    print(f"\n시트 목록: {wb.sheetnames}\n")
    
    # 1. Assumptions 시트
    if 'Assumptions' in wb.sheetnames:
        ws = wb['Assumptions']
        
        print("1️⃣ Assumptions 시트")
        print("-"*70)
        
        # 처음 5개 가정
        print("\n가정 데이터 (Row 5-10):")
        for row_idx in range(5, 11):
            a = ws[f'A{row_idx}'].value
            b = ws[f'B{row_idx}'].value
            c = ws[f'C{row_idx}'].value
            d = ws[f'D{row_idx}'].value
            
            print(f"  Row {row_idx}: {a} | {b} | {c} | {d}")
            
            # D 컬럼 (Value)이 비어있으면 문제
            if b and d is None:
                print(f"    ⚠️ Value 없음!")
    
    # 2. Method_1_TopDown 시트
    if 'Method_1_TopDown' in wb.sheetnames:
        ws = wb['Method_1_TopDown']
        
        print("\n2️⃣ Method_1_TopDown 시트")
        print("-"*70)
        
        # TAM 시작 (보통 B5)
        print("\nTAM 및 Narrowing:")
        for row_idx in range(5, 12):
            a = ws[f'A{row_idx}'].value
            b = ws[f'B{row_idx}'].value
            c = ws[f'C{row_idx}'].value
            
            if a:
                print(f"  Row {row_idx}: {a}")
                print(f"    B{row_idx} (값/수식): {b}")
                if c:
                    print(f"    C{row_idx} (비율): {c}")
    
    # 3. Method_2_BottomUp 시트
    if 'Method_2_BottomUp' in wb.sheetnames:
        ws = wb['Method_2_BottomUp']
        
        print("\n3️⃣ Method_2_BottomUp 시트")
        print("-"*70)
        
        print("\nSegment 계산:")
        for row_idx in range(5, 15):
            a = ws[f'A{row_idx}'].value
            b = ws[f'B{row_idx}'].value
            
            if a:
                print(f"  Row {row_idx}: {a} | {b}")
    
    # 4. Convergence_Analysis 시트
    if 'Convergence_Analysis' in wb.sheetnames:
        ws = wb['Convergence_Analysis']
        
        print("\n4️⃣ Convergence_Analysis 시트")
        print("-"*70)
        
        print("\n4가지 Method SAM:")
        for row_idx in range(5, 20):
            a = ws[f'A{row_idx}'].value
            b = ws[f'B{row_idx}'].value
            c = ws[f'C{row_idx}'].value
            
            if a and ('Method' in str(a) or '평균' in str(a) or 'Max/Min' in str(a)):
                print(f"  Row {row_idx}: {a}")
                print(f"    B{row_idx}: {b}")
                if c:
                    print(f"    C{row_idx}: {c}")
    
    # 5. Summary 시트
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        
        print("\n5️⃣ Summary 시트")
        print("-"*70)
        
        print("\n핵심 지표:")
        for row_idx in range(4, 15):
            a = ws[f'A{row_idx}'].value
            b = ws[f'B{row_idx}'].value
            
            if a and b:
                print(f"  Row {row_idx}: {a} = {b}")
    
    # Named Range 확인
    print("\n6️⃣ Named Ranges")
    print("-"*70)
    
    named_ranges = list(wb.defined_names)
    print(f"\n총 {len(named_ranges)}개 Named Range:")
    for name in named_ranges[:15]:
        print(f"  - {name}")
    
    if len(named_ranges) > 15:
        print(f"  ... 외 {len(named_ranges) - 15}개")


if __name__ == "__main__":
    diagnose_market_sizing()
    
    print("\n" + "="*70)
    print("📋 진단 완료")
    print("="*70)
    print("\n확인할 사항:")
    print("  1. Assumptions의 Value (D열)이 채워져 있는가?")
    print("  2. Method 시트들의 SAM이 계산되었는가?")
    print("  3. Convergence에 4가지 SAM 값이 있는가?")
    print("  4. Summary에 값이 표시되는가?")

