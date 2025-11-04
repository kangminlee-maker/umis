#!/usr/bin/env python3
"""
Market Sizing v7.2.0 테스트
- Estimation Details 7개 섹션
- Bottom-Up Narrowing
- Proxy 메타데이터
"""

from openpyxl import Workbook
from umis_rag.deliverables.excel.formula_engine import FormulaEngine
from umis_rag.deliverables.excel.assumptions_builder import EstimationDetailsBuilder
from umis_rag.deliverables.excel.method_builders import Method2BottomUpBuilder, Method3ProxyBuilder


def test_estimation_details():
    """Estimation Details 7개 섹션 테스트"""
    
    wb = Workbook()
    wb.remove(wb.active)
    fe = FormulaEngine(wb)
    
    # Estimation Details Builder
    builder = EstimationDetailsBuilder(wb, fe)
    
    # 추정치 데이터 (7개 섹션 포함)
    estimations = [
        {
            'id': 'FILTER_PIANO',
            'description': '피아노 비중 (전체 악기 대비)',
            'value': 0.25,
            'confidence': 'Medium',
            'error_range': '±10%',
            'used_in': 'Method_1, Method_2',
            
            # 7개 섹션
            'reason': '직접 데이터 없음. 악기별 판매 데이터 미공개',
            'base_data': [
                {'name': '전체 악기 시장', 'value': '1000억원', 'source': 'SRC_20250104_001'},
                {'name': '피아노 검색량', 'value': '월 10만건', 'source': 'Google Trends'}
            ],
            'logic_steps': [
                '전체 악기 검색량 대비 피아노 25%',
                '판매액도 검색량과 유사하다고 가정',
                '최종: 25%'
            ],
            'calculation': '피아노 검색량 / 전체 악기 검색량 = 0.25',
            'verification': '상한 35% (현악기 전체), 하한 15% (건반 악기 내 비중)',
            'alternatives': [
                '방법1: 수입 통계 → 데이터 부족',
                '방법2: 제조사 인터뷰 → 시간 부족',
                '현재 방법이 가용 데이터로 최선'
            ]
        },
        {
            'id': 'PURCHASE_RATE_EST',
            'description': '구독 전환율 (타겟 고객 대비)',
            'value': 0.15,
            'confidence': 'Medium',
            'error_range': '±20%',
            'used_in': 'Method_2',
            
            'reason': '신규 서비스로 전환율 데이터 없음',
            'base_data': [
                {'name': '유사 서비스 (기타 구독)', 'value': '12%', 'source': '경쟁사 공시'},
                {'name': '음악 앱 구독률', 'value': '18%', 'source': 'Statista'}
            ],
            'logic_steps': [
                '유사 서비스 기타 구독: 12%',
                '음악 앱: 18%',
                '중간값: 15%'
            ],
            'calculation': '(12% + 18%) / 2 = 15%',
            'verification': '상한 25% (프리미엄 서비스), 하한 5% (보수적)',
            'alternatives': [
                '방법1: 설문조사 → 예산 부족',
                '방법2: 베타 테스트 → 시간 부족',
                '벤치마크 방법 채택'
            ]
        },
        {
            'id': 'AOV_EST',
            'description': 'AOV (Average Order Value)',
            'value': 50000,
            'confidence': 'High',
            'error_range': '±5%',
            'used_in': 'Method_2',
            
            'reason': '가격 정책 미확정',
            'base_data': [
                {'name': '경쟁사 A 가격', 'value': '49,000원', 'source': '웹사이트'},
                {'name': '경쟁사 B 가격', 'value': '55,000원', 'source': '웹사이트'}
            ],
            'logic_steps': [
                '경쟁사 평균: 52,000원',
                '우리 포지셔닝: 중간가',
                '최종: 50,000원'
            ],
            'calculation': '(49,000 + 55,000) / 2 ≈ 50,000',
            'verification': '시장 가격 범위 내 (40K ~ 60K)',
            'alternatives': [
                '방법1: 가격 민감도 조사 → 시간 부족',
                '경쟁사 가격이 신뢰도 높음'
            ]
        },
        {
            'id': 'FREQUENCY_EST',
            'description': '연간 구매 빈도',
            'value': 12,
            'confidence': 'High',
            'error_range': '±0',
            'used_in': 'Method_2',
            
            'reason': '구독 모델 특성상 명확',
            'base_data': [
                {'name': '구독 주기', 'value': '월 구독', 'source': '비즈니스 모델'}
            ],
            'logic_steps': [
                '월 구독 = 연 12회'
            ],
            'calculation': '12 (월 구독)',
            'verification': '구독 모델 정의상 확정',
            'alternatives': []
        }
    ]
    
    builder.create_sheet(estimations)
    
    wb.save('test_output/market_sizing_estimation_details_v7_2.xlsx')
    print(f"✅ Estimation Details 테스트 완료")
    print(f"   - 4개 추정치")
    print(f"   - 7개 섹션 기반")
    print(f"   - Named Range: {len(estimations)}개")


def test_bottomup_narrowing():
    """Bottom-Up Narrowing 테스트"""
    
    wb = Workbook()
    wb.remove(wb.active)
    fe = FormulaEngine(wb)
    
    # Assumptions 시트 (간단 버전)
    ws_asm = wb.create_sheet("Assumptions")
    ws_asm['A1'] = 'ID'
    ws_asm['B1'] = 'Value'
    
    ws_asm['A2'] = 'TOTAL_POPULATION'
    ws_asm['B2'] = 5000000  # 500만명
    fe.define_named_range('TOTAL_POPULATION', 'Assumptions', 'B2')
    
    ws_asm['A3'] = 'FILTER_KOREA'
    ws_asm['B3'] = 0.15  # 한국 15%
    fe.define_named_range('FILTER_KOREA', 'Assumptions', 'B3')
    
    ws_asm['A4'] = 'FILTER_PIANO'
    ws_asm['B4'] = 0.25  # 피아노 25%
    fe.define_named_range('FILTER_PIANO', 'Assumptions', 'B4')
    
    ws_asm['A5'] = 'PURCHASE_RATE_EST'
    ws_asm['B5'] = 0.15
    fe.define_named_range('PURCHASE_RATE_EST', 'Assumptions', 'B5')
    
    ws_asm['A6'] = 'AOV_EST'
    ws_asm['B6'] = 50000
    fe.define_named_range('AOV_EST', 'Assumptions', 'B6')
    
    ws_asm['A7'] = 'FREQUENCY_EST'
    ws_asm['B7'] = 12
    fe.define_named_range('FREQUENCY_EST', 'Assumptions', 'B7')
    
    # Bottom-Up Builder
    builder = Method2BottomUpBuilder(wb, fe)
    
    segments = [
        {
            'name': '개인 구독',
            'total_population': 'TOTAL_POPULATION',
            'narrowing_filters': [
                {'name': '지역 (한국)', 'filter_id': 'FILTER_KOREA'},
                {'name': '피아노 관심', 'filter_id': 'FILTER_PIANO'}
            ],
            'purchase_rate': 'PURCHASE_RATE_EST',
            'aov': 'AOV_EST',
            'frequency': 'FREQUENCY_EST',
            'notes': '개인 대상 월 구독'
        }
    ]
    
    builder.create_sheet(segments)
    
    wb.save('test_output/market_sizing_bottomup_narrowing_v7_2.xlsx')
    print(f"✅ Bottom-Up Narrowing 테스트 완료")
    print(f"   - Narrowing 2단계 (한국 15% × 피아노 25%)")
    print(f"   - Narrowed Customers = 5,000,000 × 0.15 × 0.25 = 187,500명")
    print(f"   - SAM = 187,500 × 0.15 × 50,000 × 12 = 약 168억원")


def test_proxy_metadata():
    """Proxy 메타데이터 테스트"""
    
    wb = Workbook()
    wb.remove(wb.active)
    fe = FormulaEngine(wb)
    
    # Assumptions 시트
    ws_asm = wb.create_sheet("Assumptions")
    ws_asm['A1'] = 'ID'
    ws_asm['B1'] = 'Value'
    
    ws_asm['A2'] = 'PROXY_SIZE'
    ws_asm['B2'] = 50000000000  # 500억
    fe.define_named_range('PROXY_SIZE', 'Assumptions', 'B2')
    
    ws_asm['A3'] = 'PROXY_CORR'
    ws_asm['B3'] = 0.3
    fe.define_named_range('PROXY_CORR', 'Assumptions', 'B3')
    
    ws_asm['A4'] = 'PROXY_APP'
    ws_asm['B4'] = 0.5
    fe.define_named_range('PROXY_APP', 'Assumptions', 'B4')
    
    # Proxy Builder
    builder = Method3ProxyBuilder(wb, fe)
    
    proxy_data = {
        'proxy_market_name': '바이올린 구독 서비스 시장',
        'proxy_market': 'PROXY_SIZE',
        'similarity_reason': '같은 현악기 구독 모델, 고가 악기, 월 구독형',
        'correlation': 'PROXY_CORR',
        'correlation_basis': 'SNS 검색 트렌드 상관계수 0.3 (Google Trends 분석)',
        'application_rate': 'PROXY_APP',
        'application_basis': '바이올린 대비 피아노 시장 50% 수준 (성숙도 차이)'
    }
    
    builder.create_sheet(proxy_data)
    
    wb.save('test_output/market_sizing_proxy_metadata_v7_2.xlsx')
    print(f"✅ Proxy 메타데이터 테스트 완료")
    print(f"   - Proxy 시장: 바이올린 구독 (500억)")
    print(f"   - 유사성 근거 포함")
    print(f"   - SAM = 500억 × 0.3 × 0.5 = 75억")


if __name__ == '__main__':
    print("\n" + "="*70)
    print("Market Sizing v7.2.0 테스트")
    print("="*70 + "\n")
    
    print("1️⃣  Estimation Details (7개 섹션)")
    test_estimation_details()
    
    print("\n2️⃣  Bottom-Up (Narrowing 로직)")
    test_bottomup_narrowing()
    
    print("\n3️⃣  Proxy (메타데이터)")
    test_proxy_metadata()
    
    print("\n" + "="*70)
    print("✅ 모든 테스트 완료!")
    print("="*70)

