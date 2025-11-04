#!/usr/bin/env python3
"""
Assumptions 값 확인
모든 Named Range의 실제 값 확인
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'

wb = load_workbook(filepath, data_only=False)

print("\n" + "="*70)
print("🔍 Assumptions 시트 전체 데이터")
print("="*70 + "\n")

if 'Assumptions' in wb.sheetnames:
    ws = wb['Assumptions']
    
    # Row 1-20 출력
    print("Row 1-15:")
    print("-"*70)
    
    for row_idx in range(1, 16):
        values = []
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row_idx}']
            values.append(str(cell.value) if cell.value is not None else '')
        
        if any(values):
            print(f"Row {row_idx}: {' | '.join(values)}")

print("\n" + "="*70)
print("🔍 주요 Named Range 실제 값")
print("="*70 + "\n")

# TAM_VALUE 확인
print("TAM_VALUE:")
if 'TAM_VALUE' in wb.defined_names:
    for sheet, cell in wb.defined_names['TAM_VALUE'].destinations:
        ws = wb[sheet]
        cell_clean = cell.replace('$', '')
        value = ws[cell_clean].value
        print(f"  {sheet}!{cell} = {value}")
        print(f"  타입: {type(value)}")
        print(f"  기대값: 100,000,000,000 (₩1,000억)")

print("\nFILTER_KOREA:")
if 'FILTER_KOREA' in wb.defined_names:
    for sheet, cell in wb.defined_names['FILTER_KOREA'].destinations:
        ws = wb[sheet]
        cell_clean = cell.replace('$', '')
        value = ws[cell_clean].value
        print(f"  {sheet}!{cell} = {value}")
        print(f"  타입: {type(value)}")
        print(f"  기대값: 0.15 (15%)")

print("\nFILTER_PIANO:")
if 'FILTER_PIANO' in wb.defined_names:
    for sheet, cell in wb.defined_names['FILTER_PIANO'].destinations:
        ws = wb[sheet]
        cell_clean = cell.replace('$', '')
        value = ws[cell_clean].value
        print(f"  {sheet}!{cell} = {value}")
        print(f"  타입: {type(value)}")
        print(f"  기대값: 0.25 (25%)")

print("\n" + "="*70)
print("📊 계산 추적")
print("="*70)

# 계산 추적
print("\n예상 계산:")
print("  TAM = 100,000,000,000")
print("  Step 1: TAM × 15% = 15,000,000,000 (₩150억)")
print("  Step 2: Step 1 × 25% = 3,750,000,000 (₩37.5억)")
print("  SAM = ₩37.5억")

print("\n실제 계산 (수식 추적):")
if 'Method_1_TopDown' in wb.sheetnames:
    ws = wb['Method_1_TopDown']
    print(f"  A5 = {ws['A5'].value}")
    print(f"  B5 = {ws['B5'].value}")
    print(f"  C5 = {ws['C5'].value}")
    print(f"  B6 = {ws['B6'].value} (= A5 × B5)")
    print(f"  C6 = {ws['C6'].value} (= B6 × C5 = SAM)")

