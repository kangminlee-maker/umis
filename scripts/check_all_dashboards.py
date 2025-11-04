#!/usr/bin/env python3
"""
모든 Excel의 Summary/Dashboard 시트 값 확인
주요 셀에 숫자가 제대로 출력되는지 검증
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
examples_dir = project_root / 'examples' / 'excel'

print("\n" + "="*70)
print("🔍 모든 Dashboard/Summary 시트 값 확인")
print("="*70)

issues = []

# 1. Market Sizing (Summary 시트)
print("\n1️⃣ Market Sizing - Summary 시트")
print("-"*70)

filepath = examples_dir / 'market_sizing_piano_subscription_CALCULATED_20251104.xlsx'
if filepath.exists():
    wb = load_workbook(filepath, data_only=True)
    
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        
        # 주요 셀 확인
        checks = [
            ('B5', 'TAM', 100_000_000_000),
            ('B6', 'SAM (평균)', 12_062_500_000),
            ('B10', 'Method 1 SAM', 3_750_000_000),
            ('B11', 'Method 2 SAM', 12_000_000_000),
            ('B12', 'Method 3 SAM', 7_500_000_000),
            ('B13', 'Method 4 SAM', 25_000_000_000),
            ('B16', 'Max/Min Ratio', 6.67),
            ('B23', 'Best Case Avg SAM', 12_062_500_000 * 1.15),
            ('B24', 'Base Case Avg SAM', 12_062_500_000),
            ('B25', 'Worst Case Avg SAM', 12_062_500_000 * 0.85),
        ]
        
        for cell, label, expected in checks:
            value = ws[cell].value
            
            if value is None:
                issues.append(f"❌ Market Sizing Summary!{cell} ({label}): 값 없음")
                print(f"❌ {cell} ({label}): None")
            elif isinstance(value, str):
                # 수식이 그대로 있음
                issues.append(f"❌ Market Sizing Summary!{cell} ({label}): 수식만 있음 ({value})")
                print(f"❌ {cell} ({label}): 수식 ({value[:30]}...)")
            else:
                # 숫자 확인
                if expected > 1000:
                    formatted = f"₩{value/1_0000_0000:.1f}억"
                    expected_fmt = f"₩{expected/1_0000_0000:.1f}억"
                else:
                    formatted = f"{value:.2f}"
                    expected_fmt = f"{expected:.2f}"
                
                # 오차 확인
                error = abs(value - expected) / abs(expected) if expected != 0 else abs(value - expected)
                
                if error < 0.02:  # 2% 허용
                    print(f"✅ {cell} ({label}): {formatted}")
                else:
                    issues.append(f"⚠️ Market Sizing Summary!{cell} ({label}): {formatted} ≠ {expected_fmt} (오차 {error*100:.1f}%)")
                    print(f"⚠️ {cell} ({label}): {formatted} ≠ {expected_fmt}")
    else:
        issues.append("❌ Market Sizing: Summary 시트 없음")
        print("❌ Summary 시트 없음")
else:
    issues.append("❌ Market Sizing: 파일 없음")
    print("❌ 파일 없음")

# 2. Unit Economics (Dashboard 시트)
print("\n2️⃣ Unit Economics - Dashboard 시트")
print("-"*70)

filepath = examples_dir / 'unit_economics_CALCULATED_20251104.xlsx'
if filepath.exists():
    wb = load_workbook(filepath, data_only=True)
    
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        
        checks = [
            ('B5', 'LTV', 78750),
            ('B6', 'CAC', 25000),
            ('B7', 'LTV/CAC Ratio', 3.15),
            ('B8', 'Payback Period', 7.94),
        ]
        
        for cell, label, expected in checks:
            value = ws[cell].value
            
            if value is None:
                issues.append(f"❌ Unit Economics Dashboard!{cell} ({label}): 값 없음")
                print(f"❌ {cell} ({label}): None")
            elif isinstance(value, str):
                issues.append(f"❌ Unit Economics Dashboard!{cell} ({label}): 수식만 있음")
                print(f"❌ {cell} ({label}): 수식 ({value[:30]}...)")
            else:
                if expected > 1000:
                    formatted = f"₩{value:,.0f}"
                    expected_fmt = f"₩{expected:,.0f}"
                else:
                    formatted = f"{value:.2f}"
                    expected_fmt = f"{expected:.2f}"
                
                error = abs(value - expected) / abs(expected) if expected != 0 else abs(value - expected)
                
                if error < 0.02:
                    print(f"✅ {cell} ({label}): {formatted}")
                else:
                    issues.append(f"⚠️ Unit Economics Dashboard!{cell}: {formatted} ≠ {expected_fmt}")
                    print(f"⚠️ {cell} ({label}): {formatted} ≠ {expected_fmt}")
    else:
        issues.append("❌ Unit Economics: Dashboard 시트 없음")
        print("❌ Dashboard 시트 없음")
else:
    issues.append("❌ Unit Economics: 파일 없음")
    print("❌ 파일 없음")

# 3. Financial Projection (Dashboard 시트)
print("\n3️⃣ Financial Projection - Dashboard 시트")
print("-"*70)

filepath = examples_dir / 'financial_projection_CALCULATED_20251104.xlsx'
if filepath.exists():
    wb = load_workbook(filepath, data_only=True)
    
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        
        checks = [
            ('B5', 'Revenue Year 5', 4295_0000_0000),
            ('B6', 'Net Income Year 5', 429_0000_0000),
            ('B7', 'CAGR', 0.28),
        ]
        
        for cell, label, expected in checks:
            value = ws[cell].value
            
            if value is None:
                issues.append(f"❌ Financial Projection Dashboard!{cell} ({label}): 값 없음")
                print(f"❌ {cell} ({label}): None")
            elif isinstance(value, str):
                issues.append(f"❌ Financial Projection Dashboard!{cell} ({label}): 수식만 있음")
                print(f"❌ {cell} ({label}): 수식 ({value[:30]}...)")
            else:
                if expected > 1000:
                    formatted = f"₩{value/1_0000_0000:.0f}억"
                    expected_fmt = f"₩{expected/1_0000_0000:.0f}억"
                else:
                    formatted = f"{value*100:.0f}%" if value < 1 else f"{value:.2f}"
                    expected_fmt = f"{expected*100:.0f}%" if expected < 1 else f"{expected:.2f}"
                
                error = abs(value - expected) / abs(expected) if expected != 0 else abs(value - expected)
                
                if error < 0.02:
                    print(f"✅ {cell} ({label}): {formatted}")
                else:
                    issues.append(f"⚠️ Financial Projection Dashboard!{cell}: {formatted} ≠ {expected_fmt}")
                    print(f"⚠️ {cell} ({label}): {formatted} ≠ {expected_fmt}")
    else:
        issues.append("❌ Financial Projection: Dashboard 시트 없음")
        print("❌ Dashboard 시트 없음")
else:
    issues.append("❌ Financial Projection: 파일 없음")
    print("❌ 파일 없음")

# 최종 결과
print("\n" + "="*70)
print("📊 최종 확인 결과")
print("="*70)

if not issues:
    print("\n✅ 모든 Dashboard/Summary 값 정상!")
    print("\n검증 완료:")
    print("  - Market Sizing: 10개 값 ✅")
    print("  - Unit Economics: 4개 값 ✅")
    print("  - Financial Projection: 3개 값 ✅")
    print("\n💡 즉시 확인 가능:")
    print("  - 모든 값이 계산되어 표시됨")
    print("  - Excel에서 열면 바로 확인")
    sys.exit(0)
else:
    print(f"\n❌ {len(issues)}개 문제 발견:\n")
    for issue in issues:
        print(f"  {issue}")
    
    print("\n📋 수정 필요:")
    print("  1. 값 없음 → populate 스크립트 재실행")
    print("  2. 수식만 있음 → data_only=False로 로드된 것 (정상일 수 있음)")
    print("  3. 값 불일치 → 계산 로직 확인")
    
    sys.exit(1)

