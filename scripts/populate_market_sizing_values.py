#!/usr/bin/env python3
"""
Market Sizing Excel에 계산된 값 직접 입력
수식은 주석으로 남기고, 값을 하드코딩
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def populate_values():
    """계산된 값을 직접 입력"""
    
    filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'
    
    print("\n" + "="*70)
    print("📊 Market Sizing 값 입력 (하드코딩)")
    print("="*70 + "\n")
    
    # 파일 열기
    wb = load_workbook(filepath)
    
    # === 1. Method_1_TopDown 값 입력 ===
    if 'Method_1_TopDown' in wb.sheetnames:
        ws = wb['Method_1_TopDown']
        
        print("1️⃣ Method_1_TopDown 계산")
        
        # TAM
        tam = 100_000_000_000  # ₩1,000억
        ws['A5'] = tam
        ws['A5'].number_format = '#,##0'
        ws['A5'].comment = Comment("원 수식: =TAM_VALUE", "System")
        print(f"  TAM (A5): ₩{tam/1_0000_0000:.0f}억")
        
        # 비율
        korea_ratio = 0.15
        piano_ratio = 0.25
        ws['B5'] = korea_ratio
        ws['B5'].number_format = '0.0%'
        ws['C5'] = piano_ratio
        ws['C5'].number_format = '0.0%'
        
        # Step 1: TAM × 한국
        step1 = tam * korea_ratio  # ₩150억
        ws['B6'] = step1
        ws['B6'].number_format = '#,##0'
        ws['B6'].comment = Comment("원 수식: =A5*B5", "System")
        print(f"  한국 시장 (B6): ₩{step1/1_0000_0000:.0f}억")
        
        # Step 2: Step 1 × 피아노 = SAM
        sam1 = step1 * piano_ratio  # ₩37.5억
        ws['C6'] = sam1
        ws['C6'].number_format = '#,##0'
        ws['C6'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws['C6'].font = Font(bold=True)
        ws['C6'].comment = Comment("원 수식: =B6*C5\nSAM (Method 1)", "System")
        print(f"  SAM (C6): ₩{sam1/1_0000_0000:.1f}억 ✅\n")
    
    # === 2. Method_2_BottomUp 값 입력 ===
    if 'Method_2_BottomUp' in wb.sheetnames:
        ws = wb['Method_2_BottomUp']
        
        print("2️⃣ Method_2_BottomUp 계산")
        
        # 세그먼트 계산
        customers = 100_000  # 명
        rate = 0.20  # 20%
        aov = 50_000  # 원
        freq = 12  # 회
        
        ws['B4'] = customers
        ws['C4'] = rate
        ws['C4'].number_format = '0.0%'
        ws['D4'] = aov
        ws['D4'].number_format = '#,##0'
        ws['E4'] = freq
        
        # SAM
        sam2 = customers * rate * aov * freq  # ₩12,000,000,000 = ₩120억
        ws['F4'] = sam2
        ws['F4'].number_format = '#,##0'
        ws['F4'].comment = Comment("원 수식: =B4*C4*D4*E4", "System")
        
        ws['F6'] = sam2
        ws['F6'].number_format = '#,##0'
        ws['F6'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws['F6'].font = Font(bold=True)
        ws['F6'].comment = Comment("원 수식: =SUM(F4:F4)\nSAM (Method 2)", "System")
        print(f"  SAM (F6): ₩{sam2/1_0000_0000:.0f}억 ✅\n")
    
    # === 3. Method_3_Proxy 값 입력 ===
    if 'Method_3_Proxy' in wb.sheetnames:
        ws = wb['Method_3_Proxy']
        
        print("3️⃣ Method_3_Proxy 계산")
        
        proxy_market = 50_000_000_000  # ₩500억
        correlation = 0.3
        application = 0.5
        
        ws['B3'] = proxy_market
        ws['B3'].number_format = '#,##0'
        ws['B4'] = correlation
        ws['B4'].number_format = '0.0%'
        ws['B5'] = application
        ws['B5'].number_format = '0.0%'
        
        # SAM
        sam3 = proxy_market * correlation * application  # ₩75억
        ws['B7'] = sam3
        ws['B7'].number_format = '#,##0'
        ws['B7'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws['B7'].font = Font(bold=True)
        ws['B7'].comment = Comment("원 수식: =B3*B4*B5\nSAM (Method 3)", "System")
        print(f"  SAM (B7): ₩{sam3/1_0000_0000:.0f}억 ✅\n")
    
    # === 4. Method_4_CompetitorRevenue 값 입력 ===
    if 'Method_4_CompetitorRevenue' in wb.sheetnames:
        ws = wb['Method_4_CompetitorRevenue']
        
        print("4️⃣ Method_4_CompetitorRevenue 계산")
        
        comp_revenue = 10_000_000_000  # ₩100억
        comp_share = 0.40  # 40%
        
        ws['B4'] = comp_revenue
        ws['B4'].number_format = '#,##0'
        ws['C4'] = comp_share
        ws['C4'].number_format = '0.0%'
        
        ws['B5'] = comp_revenue
        ws['B5'].number_format = '#,##0'
        ws['C5'] = comp_share
        ws['C5'].number_format = '0.0%'
        
        # SAM
        sam4 = comp_revenue / comp_share  # ₩250억
        ws['B7'] = sam4
        ws['B7'].number_format = '#,##0'
        ws['B7'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws['B7'].font = Font(bold=True)
        ws['B7'].comment = Comment("원 수식: =B5/C5\nSAM (Method 4)", "System")
        print(f"  SAM (B7): ₩{sam4/1_0000_0000:.0f}억 ✅\n")
    
    # === 5. Convergence_Analysis 값 입력 ===
    if 'Convergence_Analysis' in wb.sheetnames:
        ws = wb['Convergence_Analysis']
        
        print("5️⃣ Convergence_Analysis 계산")
        
        # 4가지 SAM
        ws['B4'] = sam1  # Method 1
        ws['B4'].number_format = '#,##0'
        
        ws['B5'] = sam2  # Method 2
        ws['B5'].number_format = '#,##0'
        
        ws['B6'] = sam3  # Method 3
        ws['B6'].number_format = '#,##0'
        
        ws['B7'] = sam4  # Method 4
        ws['B7'].number_format = '#,##0'
        
        # 평균
        avg_sam = (sam1 + sam2 + sam3 + sam4) / 4  # ₩120.6억
        ws['B8'] = avg_sam
        ws['B8'].number_format = '#,##0'
        ws['B8'].font = Font(bold=True)
        ws['B8'].comment = Comment("원 수식: =AVERAGE(B4:B7)", "System")
        print(f"  평균 SAM (B8): ₩{avg_sam/1_0000_0000:.1f}억")
        
        # 표준편차
        import math
        values = [sam1, sam2, sam3, sam4]
        mean = avg_sam
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        stdev = math.sqrt(variance)
        ws['B9'] = stdev
        ws['B9'].number_format = '#,##0'
        print(f"  표준편차 (B9): ₩{stdev/1_0000_0000:.1f}억")
        
        # CV%
        cv = stdev / mean * 100
        ws['B10'] = cv / 100
        ws['B10'].number_format = '0.0%'
        print(f"  변동계수 (B10): {cv:.1f}%")
        
        # Max/Min 비율
        max_min = max(values) / min(values)
        ws['B11'] = max_min
        ws['B11'].number_format = '0.00'
        print(f"  Max/Min (B11): {max_min:.2f}")
        
        # 수렴 여부
        if max_min <= 1.3:
            ws['B12'] = "✅ 통과 (±30% 수렴)"
        else:
            ws['B12'] = "❌ 재검토 필요"
        print(f"  수렴 여부 (B12): {ws['B12'].value}")
        
        # 차이 % 계산 (C4-C7)
        for idx, sam_value in enumerate([sam1, sam2, sam3, sam4], start=4):
            diff_pct = (sam_value - avg_sam) / avg_sam * 100
            ws[f'C{idx}'] = diff_pct / 100
            ws[f'C{idx}'].number_format = '0.0%'
        
        print()
    
    # === 6. Summary 값 입력 ===
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        
        print("6️⃣ Summary 계산")
        
        # TAM
        ws['B5'] = tam
        ws['B5'].number_format = '#,##0'
        print(f"  TAM (B5): ₩{tam/1_0000_0000:.0f}억")
        
        # SAM (평균)
        ws['B6'] = avg_sam
        ws['B6'].number_format = '#,##0'
        ws['B6'].font = Font(bold=True, color="0070C0")
        ws['B6'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        print(f"  SAM (B6): ₩{avg_sam/1_0000_0000:.1f}억 ✅")
        
        # 4가지 Method SAM
        ws['B10'] = sam1
        ws['B10'].number_format = '#,##0'
        
        ws['B11'] = sam2
        ws['B11'].number_format = '#,##0'
        
        ws['B12'] = sam3
        ws['B12'].number_format = '#,##0'
        
        ws['B13'] = sam4
        ws['B13'].number_format = '#,##0'
        
        print()
    
    # 저장
    output_path = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_CALCULATED_20251104.xlsx'
    wb.save(output_path)
    
    print("="*70)
    print("✅ 값 입력 완료!")
    print("="*70)
    print(f"\n📁 저장 위치: {output_path}")
    print(f"📏 파일 크기: {output_path.stat().st_size / 1024:.1f} KB\n")
    
    print("📊 입력된 값:")
    print(f"  Method 1 (Top-Down): ₩{sam1/1_0000_0000:.1f}억")
    print(f"  Method 2 (Bottom-Up): ₩{sam2/1_0000_0000:.0f}억")
    print(f"  Method 3 (Proxy): ₩{sam3/1_0000_0000:.0f}억")
    print(f"  Method 4 (Competitor): ₩{sam4/1_0000_0000:.0f}억")
    print(f"\n  평균 SAM: ₩{avg_sam/1_0000_0000:.1f}억")
    print(f"  Max/Min: {max_min:.2f} ❌ (수렴 실패)")
    
    print("\n💡 이제 이 파일을 열면:")
    print("   - 모든 값이 즉시 표시됩니다")
    print("   - 수식은 주석으로 남아있습니다")
    print("   - Summary, Convergence 모두 값 있음")
    
    return output_path


if __name__ == "__main__":
    filepath = populate_values()
    
    print("\n" + "="*70)
    print("🎉 완료!")
    print("="*70)
    print(f"\n다음 명령어로 열기:")
    print(f"open {filepath}")
    
    sys.exit(0)

