#!/usr/bin/env python3
"""
Excel 수식 진단
주요 셀의 수식을 출력하여 문제 확인
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Add project root
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


def diagnose_financial_projection():
    """Financial Projection 수식 진단"""
    
    filepath = project_root / 'examples' / 'excel' / 'financial_projection_korean_adult_education_example_20251104.xlsx'
    
    if not filepath.exists():
        print(f"❌ 파일 없음: {filepath}")
        return
    
    print("\n" + "="*70)
    print("🔍 Financial Projection 수식 진단")
    print("="*70)
    
    wb = load_workbook(filepath, data_only=False)
    
    # 1. Revenue_Buildup
    if 'Revenue_Buildup' in wb.sheetnames:
        ws = wb['Revenue_Buildup']
        
        print("\n1️⃣ Revenue_Buildup 시트")
        print("-"*70)
        
        # 세그먼트 Row 5
        print("\n세그먼트 1 (Row 5):")
        print(f"  A5: {ws['A5'].value}")
        print(f"  B5 (Year 0): {ws['B5'].value}")
        print(f"  C5 (Year 1): {ws['C5'].value}")
        print(f"  D5 (Year 2): {ws['D5'].value}")
        print(f"  H5 (Growth): {ws['H5'].value}")
        
        # 세그먼트 Row 6
        print("\n세그먼트 2 (Row 6):")
        print(f"  A6: {ws['A6'].value}")
        print(f"  B6 (Year 0): {ws['B6'].value}")
        print(f"  C6 (Year 1): {ws['C6'].value}")
        print(f"  H6 (Growth): {ws['H6'].value}")
        
        # Total Revenue
        print("\nTotal Revenue (Row 9):")
        print(f"  A9: {ws['A9'].value}")
        print(f"  B9 (Year 0): {ws['B9'].value}")
        print(f"  C9 (Year 1): {ws['C9'].value}")
        print(f"  G9 (Year 5): {ws['G9'].value}")
        
        # YoY Growth
        print("\nYoY Growth (Row 10):")
        print(f"  C10 (Year 1): {ws['C10'].value}")
    
    # 2. Cost_Structure
    if 'Cost_Structure' in wb.sheetnames:
        ws = wb['Cost_Structure']
        
        print("\n2️⃣ Cost_Structure 시트")
        print("-"*70)
        
        # Revenue 참조 행 찾기
        revenue_row = None
        cogs_row = None
        
        for row_idx in range(4, 10):
            cell = ws[f'A{row_idx}']
            if cell.value:
                if 'Revenue' in str(cell.value) and 'Total' not in str(cell.value):
                    revenue_row = row_idx
                elif 'COGS' in str(cell.value):
                    cogs_row = row_idx
        
        if revenue_row and cogs_row:
            print(f"\nRevenue (Row {revenue_row}):")
            print(f"  B{revenue_row} (Year 0): {ws[f'B{revenue_row}'].value}")
            print(f"  C{revenue_row} (Year 1): {ws[f'C{revenue_row}'].value}")
            
            print(f"\nCOGS (Row {cogs_row}):")
            print(f"  B{cogs_row} (Year 0): {ws[f'B{cogs_row}'].value}")
            print(f"  C{cogs_row} (Year 1): {ws[f'C{cogs_row}'].value}")
            
            # 수식 분석
            cogs_y0_formula = ws[f'B{cogs_row}'].value
            cogs_y1_formula = ws[f'C{cogs_row}'].value
            
            print(f"\n수식 분석:")
            if isinstance(cogs_y0_formula, str):
                if f'B{revenue_row}' in cogs_y0_formula:
                    print(f"  ✅ COGS Year 0: Revenue Row {revenue_row}의 B 컬럼 참조 (정상)")
                elif f'C{revenue_row}' in cogs_y0_formula:
                    print(f"  ❌ COGS Year 0: Revenue Row {revenue_row}의 C 컬럼 참조 (한 칸 밀림!)")
                else:
                    print(f"  ⚠️ COGS Year 0: {cogs_y0_formula}")
            
            if isinstance(cogs_y1_formula, str):
                if f'C{revenue_row}' in cogs_y1_formula:
                    print(f"  ✅ COGS Year 1: Revenue Row {revenue_row}의 C 컬럼 참조 (정상)")
                elif f'D{revenue_row}' in cogs_y1_formula:
                    print(f"  ❌ COGS Year 1: Revenue Row {revenue_row}의 D 컬럼 참조 (한 칸 밀림!)")
                else:
                    print(f"  ⚠️ COGS Year 1: {cogs_y1_formula}")
    
    # 3. PL_5Year
    if 'PL_5Year' in wb.sheetnames:
        ws = wb['PL_5Year']
        
        print("\n3️⃣ PL_5Year 시트")
        print("-"*70)
        
        # Revenue 행 찾기
        revenue_row = None
        for row_idx in range(4, 10):
            cell = ws[f'A{row_idx}']
            if cell.value and cell.value == 'Revenue':
                revenue_row = row_idx
                break
        
        if revenue_row:
            print(f"\nRevenue (Row {revenue_row}):")
            print(f"  B{revenue_row} (Year 0): {ws[f'B{revenue_row}'].value}")
            print(f"  C{revenue_row} (Year 1): {ws[f'C{revenue_row}'].value}")
            print(f"  G{revenue_row} (Year 5): {ws[f'G{revenue_row}'].value}")
    
    # 4. Dashboard
    if 'Dashboard' in wb.sheetnames:
        ws = wb['Dashboard']
        
        print("\n4️⃣ Dashboard 시트")
        print("-"*70)
        
        print("\n주요 셀:")
        print(f"  A1: {ws['A1'].value}")
        print(f"  A5: {ws['A5'].value}")
        print(f"  B5: {ws['B5'].value}")
        print(f"  B6: {ws['B6'].value}")
        print(f"  B7: {ws['B7'].value}")


def diagnose_unit_economics():
    """Unit Economics 수식 진단"""
    
    filepath = project_root / 'examples' / 'excel' / 'unit_economics_music_streaming_example_20251104.xlsx'
    
    if not filepath.exists():
        print(f"❌ 파일 없음: {filepath}")
        return
    
    print("\n" + "="*70)
    print("🔍 Unit Economics 수식 진단")
    print("="*70)
    
    wb = load_workbook(filepath, data_only=False)
    
    # 1. Inputs
    if 'Inputs' in wb.sheetnames:
        ws = wb['Inputs']
        
        print("\n1️⃣ Inputs 시트")
        print("-"*70)
        
        print("\n핵심 입력값:")
        print(f"  B5 (ARPU): {ws['B5'].value}")
        print(f"  B6 (CAC): {ws['B6'].value}")
        print(f"  B7 (Gross Margin): {ws['B7'].value}")
        print(f"  B8 (Churn): {ws['B8'].value}")
        print(f"  B9 (Lifetime): {ws['B9'].value}")
    
    # 2. LTV_Calculation
    if 'LTV_Calculation' in wb.sheetnames:
        ws = wb['LTV_Calculation']
        
        print("\n2️⃣ LTV_Calculation 시트")
        print("-"*70)
        
        # LTV 방법 1
        for row_idx in range(8, 12):
            cell = ws[f'A{row_idx}']
            if cell.value and 'LTV (방법 1)' in str(cell.value):
                print(f"\nLTV 방법 1 (Row {row_idx}):")
                print(f"  A{row_idx}: {ws[f'A{row_idx}'].value}")
                print(f"  B{row_idx}: {ws[f'B{row_idx}'].value}")
                break
        
        # LTV 평균
        for row_idx in range(16, 22):
            cell = ws[f'A{row_idx}']
            if cell.value and '최종 LTV' in str(cell.value):
                print(f"\n최종 LTV (Row {row_idx}):")
                print(f"  A{row_idx}: {ws[f'A{row_idx}'].value}")
                print(f"  B{row_idx}: {ws[f'B{row_idx}'].value}")
                break
    
    # 3. LTV_CAC_Ratio
    if 'LTV_CAC_Ratio' in wb.sheetnames:
        ws = wb['LTV_CAC_Ratio']
        
        print("\n3️⃣ LTV_CAC_Ratio 시트")
        print("-"*70)
        
        for row_idx in range(6, 10):
            cell = ws[f'A{row_idx}']
            if cell.value and 'LTV/CAC Ratio' in str(cell.value):
                print(f"\nLTV/CAC Ratio (Row {row_idx}):")
                print(f"  A{row_idx}: {ws[f'A{row_idx}'].value}")
                print(f"  B{row_idx}: {ws[f'B{row_idx}'].value}")
                break


if __name__ == "__main__":
    print("\n" + "="*70)
    print("🔍 Excel 수식 진단 도구")
    print("="*70)
    print("\n목적: 주요 셀의 수식을 출력하여 문제 확인\n")
    
    # 1. Financial Projection 진단
    diagnose_financial_projection()
    
    # 2. Unit Economics 진단
    diagnose_unit_economics()
    
    print("\n" + "="*70)
    print("📋 진단 완료")
    print("="*70)
    print("\n💡 확인 사항:")
    print("   1. 수식에 자기 참조 있는가? (C5 = =C5*...)")
    print("   2. 수식이 올바른 셀 참조하는가? (C5 = =B5*...)")
    print("   3. Named Range 정상 참조하는가?")
    
    print("\n📋 다음 단계:")
    print("   1. Excel 파일 직접 열기")
    print("   2. 위에서 출력된 셀 위치로 이동")
    print("   3. 값이 계산되었는지 확인")
    print("   4. Golden과 비교")

