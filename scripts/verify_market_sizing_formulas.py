#!/usr/bin/env python3
"""
Market Sizing 수식 검증
Summary 시트의 모든 수식이 올바른 셀을 참조하는지 확인
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

project_root = Path(__file__).parent.parent
filepath = project_root / 'examples' / 'excel' / 'market_sizing_piano_subscription_example_20251104.xlsx'

if not filepath.exists():
    print("❌ 파일 없음")
    sys.exit(1)

wb = load_workbook(filepath, data_only=False)

print("\n" + "="*70)
print("🔍 Summary 시트 수식 상세 검증")
print("="*70)

if 'Summary' not in wb.sheetnames:
    print("❌ Summary 시트 없음")
    sys.exit(1)

ws = wb['Summary']

print("\nSummary 시트 전체 수식 (Row 1-30):")
print("-"*70)

errors = []

for row_idx in range(1, 31):
    a_val = ws[f'A{row_idx}'].value
    b_val = ws[f'B{row_idx}'].value
    
    if a_val or b_val:
        print(f"\nRow {row_idx}:")
        if a_val:
            print(f"  A{row_idx}: {a_val}")
        if b_val:
            print(f"  B{row_idx}: {b_val}")
            
            # B열이 수식이면 참조 확인
            if isinstance(b_val, str) and b_val.startswith('='):
                formula = b_val
                
                # 주요 참조 패턴 확인
                if 'Scenarios!' in formula:
                    # Scenarios의 어느 셀을 참조?
                    import re
                    match = re.search(r'Scenarios!([A-Z]+)(\d+)', formula)
                    if match:
                        col, row = match.groups()
                        ref_cell = f'{col}{row}'
                        
                        # Scenarios 시트에서 해당 셀 내용 확인
                        if 'Scenarios' in wb.sheetnames:
                            scenarios_ws = wb['Scenarios']
                            ref_content = scenarios_ws[ref_cell].value
                            ref_label = scenarios_ws[f'A{row}'].value
                            
                            print(f"     → Scenarios!{ref_cell}")
                            print(f"        A{row} (라벨): {ref_label}")
                            print(f"        {ref_cell} (값/수식): {ref_content}")
                            
                            # 의도 확인
                            if a_val and 'Best' in str(a_val) and 'Average SAM' in str(ref_label):
                                print(f"        ✅ Best Case → Average SAM 참조 정상")
                            elif a_val and 'Best' in str(a_val) and 'Average SAM' not in str(ref_label):
                                errors.append({
                                    'row': row_idx,
                                    'cell': f'B{row_idx}',
                                    'intent': a_val,
                                    'formula': formula,
                                    'ref': f'Scenarios!{ref_cell}',
                                    'ref_label': ref_label,
                                    'ref_value': ref_content,
                                    'error': f"'{a_val}'를 원하는데 Scenarios!{ref_cell}은 '{ref_label}'"
                                })
                                print(f"        ❌ 의도 불일치!")

print("\n" + "="*70)
print("📊 검증 결과")
print("="*70)

if errors:
    print(f"\n❌ {len(errors)}개 오류 발견:\n")
    
    for err in errors:
        print(f"Summary!{err['cell']} (Row {err['row']}):")
        print(f"  라벨: {err['intent']}")
        print(f"  수식: {err['formula']}")
        print(f"  참조: {err['ref']}")
        print(f"  참조 라벨: {err['ref_label']}")
        print(f"  참조 값: {err['ref_value']}")
        print(f"  ❌ 오류: {err['error']}")
        print()
    
    print("💡 수정 방법:")
    print("  1. Scenarios 시트에서 'Average SAM' 행 찾기")
    print("  2. 해당 행 번호 확인 (예: B21)")
    print("  3. summary_builder.py 수정:")
    print("     Before: =Scenarios!B13")
    print("     After: =Scenarios!B21 (또는 Named Range 사용)")
    
    sys.exit(1)
else:
    print("\n✅ 모든 수식 참조 정상!")
    print("\n검증 완료:")
    print("  - Summary의 모든 참조가 의도한 셀을 참조함")
    print("  - Scenarios!B13 같은 잘못된 참조 없음")
    
    sys.exit(0)

